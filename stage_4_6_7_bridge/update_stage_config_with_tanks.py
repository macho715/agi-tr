#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Update stage_config_sample_4_7.json with tank data from Stage_Tanks sheet"""

import json
from pathlib import Path
from openpyxl import load_workbook

# Load Stage_Tanks data
excel_path = Path("Stage_Tanks_template.xlsx")
wb = load_workbook(excel_path, data_only=True)
ws = wb["Stage_Tanks"]

# Read tank data
stage_tanks_data = {}
for row in range(2, ws.max_row + 1):
    stage = ws.cell(row=row, column=1).value
    if not stage:
        continue
    
    tank_id = ws.cell(row=row, column=2).value
    percent_fill = ws.cell(row=row, column=3).value
    sg = ws.cell(row=row, column=4).value
    use_for_ballast = ws.cell(row=row, column=5).value
    
    if not stage or not tank_id:
        continue
    
    stage = str(stage)
    if stage not in stage_tanks_data:
        stage_tanks_data[stage] = []
    
    if percent_fill and percent_fill > 0:
        stage_tanks_data[stage].append({
            "tank_id": str(tank_id),
            "percent_fill": float(percent_fill),
            "sg": float(sg) if sg else None,
        })

wb.close()

# Load and update JSON
json_path = Path("stage_config_sample_4_7.json")
with open(json_path, "r", encoding="utf-8") as f:
    config = json.load(f)

# Update tanks for each stage
for stage in config["stages"]:
    stage_name = stage["name"]
    if stage_name in stage_tanks_data:
        stage["tanks"] = stage_tanks_data[stage_name]
        print(f"[OK] Updated {stage_name}: {len(stage['tanks'])} tanks")

# Save updated JSON
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(config, f, indent=2, ensure_ascii=False)

print(f"[OK] Updated {json_path}")

