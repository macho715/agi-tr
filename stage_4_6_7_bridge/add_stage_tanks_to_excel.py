#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Add Stage_Tanks sheet to existing Excel workbook"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load existing workbook
excel_path = Path("LCT_BUSHRA_AGI_TR.xlsx")
wb = load_workbook(excel_path)

# Styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
normal_font = Font(name="Calibri", size=11)
input_fill = PatternFill("solid", fgColor="FFF2CC")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(border_style="thin", color="C0C0C0"),
    right=Side(border_style="thin", color="C0C0C0"),
    top=Side(border_style="thin", color="C0C0C0"),
    bottom=Side(border_style="thin", color="C0C0C0")
)

# Remove existing Stage_Tanks sheet if present
if "Stage_Tanks" in wb.sheetnames:
    wb.remove(wb["Stage_Tanks"])

# Create Stage_Tanks sheet
ws_tanks = wb.create_sheet("Stage_Tanks")

# Headers
headers = ["Stage", "Tank_ID", "Percent_Fill", "SG", "UseForBallast", "Comment"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws_tanks.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Sample row (Stage 5A-2, CL.P)
sample_row = 2
ws_tanks.cell(row=sample_row, column=1, value="Stage 5A-2").fill = input_fill
ws_tanks.cell(row=sample_row, column=2, value="CL.P").fill = input_fill
ws_tanks.cell(row=sample_row, column=3, value=80.0).fill = input_fill
ws_tanks.cell(row=sample_row, column=4, value=1.025).fill = input_fill
ws_tanks.cell(row=sample_row, column=5, value="TRUE").fill = input_fill
ws_tanks.cell(row=sample_row, column=6, value="샘플 행 – Stage/Tank/Percent는 프로젝트 값으로 수정").fill = input_fill

for col in range(1, 7):
    ws_tanks.cell(row=sample_row, column=col).font = normal_font
    ws_tanks.cell(row=sample_row, column=col).border = thin_border

# Column widths
ws_tanks.column_dimensions['A'].width = 15
ws_tanks.column_dimensions['B'].width = 15
ws_tanks.column_dimensions['C'].width = 15
ws_tanks.column_dimensions['D'].width = 12
ws_tanks.column_dimensions['E'].width = 15
ws_tanks.column_dimensions['F'].width = 50

# Save
wb.save(excel_path)
print(f"[OK] Added Stage_Tanks sheet to: {excel_path}")

