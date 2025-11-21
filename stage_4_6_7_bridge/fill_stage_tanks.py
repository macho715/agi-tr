#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Fill Stage_Tanks sheet with tank selections for Stage 4/6/7"""

import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load master tanks to get available tank IDs
master_tanks_path = Path("bushra_stability/data/master_tanks.json")
with open(master_tanks_path, "r", encoding="utf-8") as f:
    master_data = json.load(f)

# Tank selection for each stage
# Using common ballast tanks (SW type) that are typically used for trim adjustment
# These are example selections - should be adjusted based on actual project requirements
STAGE_TANKS = {
    "Stage 4": [
        {"tank_id": "VOID3.P", "percent_fill": 80.0, "sg": 1.025, "use_for_ballast": True, "comment": "Port ballast for trim adjustment"},
        {"tank_id": "VOID3.S", "percent_fill": 80.0, "sg": 1.025, "use_for_ballast": True, "comment": "Starboard ballast for trim adjustment"},
        {"tank_id": "VOIDDB2.C", "percent_fill": 60.0, "sg": 1.025, "use_for_ballast": True, "comment": "Center ballast tank"},
    ],
    "Stage 6": [
        {"tank_id": "VOID3.P", "percent_fill": 70.0, "sg": 1.025, "use_for_ballast": True, "comment": "Port ballast"},
        {"tank_id": "VOID3.S", "percent_fill": 70.0, "sg": 1.025, "use_for_ballast": True, "comment": "Starboard ballast"},
        {"tank_id": "FWCARGO1.P", "percent_fill": 50.0, "sg": 1.0, "use_for_ballast": True, "comment": "Fresh water cargo tank"},
        {"tank_id": "FWCARGO1.S", "percent_fill": 50.0, "sg": 1.0, "use_for_ballast": True, "comment": "Fresh water cargo tank"},
    ],
    "Stage 7": [
        {"tank_id": "VOID3.P", "percent_fill": 60.0, "sg": 1.025, "use_for_ballast": True, "comment": "Port ballast"},
        {"tank_id": "VOID3.S", "percent_fill": 60.0, "sg": 1.025, "use_for_ballast": True, "comment": "Starboard ballast"},
        {"tank_id": "FWCARGO2.P", "percent_fill": 40.0, "sg": 1.0, "use_for_ballast": True, "comment": "Fresh water cargo tank"},
        {"tank_id": "FWCARGO2.S", "percent_fill": 40.0, "sg": 1.0, "use_for_ballast": True, "comment": "Fresh water cargo tank"},
    ],
}

# Load workbook (try template first, then main file)
excel_path = Path("Stage_Tanks_template.xlsx")
if not excel_path.exists():
    excel_path = Path("LCT_BUSHRA_AGI_TR.xlsx")

try:
    wb = load_workbook(excel_path)
except Exception as e:
    print(f"[ERROR] Cannot open {excel_path}: {e}")
    print("  File may be open in Excel. Please close it and try again.")
    exit(1)

if "Stage_Tanks" not in wb.sheetnames:
    print(f"[ERROR] Stage_Tanks sheet not found in {excel_path}")
    wb.close()
    exit(1)

ws = wb["Stage_Tanks"]
input_fill = PatternFill("solid", fgColor="FFF2CC")

# Clear existing data (keep header row)
for row in range(2, ws.max_row + 1):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.value = None
        if cell.fill:
            cell.fill = PatternFill()  # Reset to default fill

# Fill data
row = 2
for stage_name, tanks in STAGE_TANKS.items():
    for tank in tanks:
        ws.cell(row=row, column=1, value=stage_name).fill = input_fill
        ws.cell(row=row, column=2, value=tank["tank_id"]).fill = input_fill
        ws.cell(row=row, column=3, value=tank["percent_fill"]).fill = input_fill
        ws.cell(row=row, column=4, value=tank["sg"]).fill = input_fill
        ws.cell(row=row, column=5, value="TRUE" if tank["use_for_ballast"] else "FALSE").fill = input_fill
        ws.cell(row=row, column=6, value=tank["comment"]).fill = input_fill
        row += 1

# Save
wb.save(excel_path)
print(f"[OK] Filled Stage_Tanks sheet in {excel_path}")
print(f"  - Stage 4: {len(STAGE_TANKS['Stage 4'])} tanks")
print(f"  - Stage 6: {len(STAGE_TANKS['Stage 6'])} tanks")
print(f"  - Stage 7: {len(STAGE_TANKS['Stage 7'])} tanks")
print(f"  - Total rows: {row - 2}")

