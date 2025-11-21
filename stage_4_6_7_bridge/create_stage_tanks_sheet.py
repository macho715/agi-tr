#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Create Stage_Tanks Excel sheet template"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load master tanks
master_tanks_path = Path("bushra_stability/data/master_tanks.json")
with open(master_tanks_path, "r", encoding="utf-8") as f:
    master_data = json.load(f)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
normal_font = Font(name="Calibri", size=11)
input_fill = PatternFill("solid", fgColor="FFF2CC")
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(border_style="thin", color="C0C0C0"),
    right=Side(border_style="thin", color="C0C0C0"),
    top=Side(border_style="thin", color="C0C0C0"),
    bottom=Side(border_style="thin", color="C0C0C0")
)

# ========== Sheet 1: Stage_Tanks ==========
ws_tanks = wb.create_sheet("Stage_Tanks")

# Headers
headers = ["Stage", "Tank_ID", "Percent_Fill", "SG", "UseForBallast", "Comment"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws_tanks.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Sample row (Stage 5A-2, CL.P)
sample_row = 2
ws_tanks.cell(row=sample_row, column=1, value="Stage 5A-2").fill = input_fill
ws_tanks.cell(row=sample_row, column=2, value="CL.P").fill = input_fill
ws_tanks.cell(row=sample_row, column=3, value=80.0).fill = input_fill
ws_tanks.cell(row=sample_row, column=4, value=1.025).fill = input_fill
ws_tanks.cell(row=sample_row, column=5, value="TRUE").fill = input_fill
ws_tanks.cell(row=sample_row, column=6, value="샘플 행 – Stage/Tank/Percent는 프로젝트 값으로 수정").fill = input_fill

for col in range(1, 7):
    ws_tanks.cell(row=sample_row, column=col).font = normal_font
    ws_tanks.cell(row=sample_row, column=col).border = thin_border

# Column widths
ws_tanks.column_dimensions['A'].width = 15
ws_tanks.column_dimensions['B'].width = 15
ws_tanks.column_dimensions['C'].width = 15
ws_tanks.column_dimensions['D'].width = 12
ws_tanks.column_dimensions['E'].width = 15
ws_tanks.column_dimensions['F'].width = 50

# ========== Sheet 2: Tank_Master_Ref ==========
ws_ref = wb.create_sheet("Tank_Master_Ref")

# Headers
ref_headers = ["Tank_ID", "Type", "Capacity_m3", "SG_Master", "LCG_m", "VCG_m", "TCG_m", "FSM_full_tm", "Content", "Location"]
for col_idx, header in enumerate(ref_headers, start=1):
    cell = ws_ref.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Tank data
for row_idx, tank in enumerate(master_data["tanks"], start=2):
    ws_ref.cell(row=row_idx, column=1, value=tank["Tank_ID"])
    ws_ref.cell(row=row_idx, column=2, value=tank.get("Type", ""))
    ws_ref.cell(row=row_idx, column=3, value=tank["Capacity_m3"])
    ws_ref.cell(row=row_idx, column=4, value=tank["SG_Master"])
    ws_ref.cell(row=row_idx, column=5, value=tank["LCG_m"])
    ws_ref.cell(row=row_idx, column=6, value=tank["VCG_m"])
    ws_ref.cell(row=row_idx, column=7, value=tank["TCG_m"])
    ws_ref.cell(row=row_idx, column=8, value=tank["FSM_full_tm"])
    ws_ref.cell(row=row_idx, column=9, value=tank.get("Content", ""))
    ws_ref.cell(row=row_idx, column=10, value=tank.get("Location", ""))
    
    for col in range(1, 11):
        ws_ref.cell(row=row_idx, column=col).font = normal_font
        ws_ref.cell(row=row_idx, column=col).border = thin_border
        if col in [3, 4, 5, 6, 7, 8]:  # Numeric columns
            ws_ref.cell(row=row_idx, column=col).number_format = "0.00"

# Column widths
ws_ref.column_dimensions['A'].width = 15
ws_ref.column_dimensions['B'].width = 10
ws_ref.column_dimensions['C'].width = 12
ws_ref.column_dimensions['D'].width = 12
ws_ref.column_dimensions['E'].width = 12
ws_ref.column_dimensions['F'].width = 12
ws_ref.column_dimensions['G'].width = 12
ws_ref.column_dimensions['H'].width = 15
ws_ref.column_dimensions['I'].width = 35
ws_ref.column_dimensions['J'].width = 15

# Save
output_path = Path("Stage_Tanks_template.xlsx")
wb.save(output_path)
print(f"[OK] Created: {output_path}")
print(f"  - Stage_Tanks sheet: Sample row for Stage 5A-2")
print(f"  - Tank_Master_Ref sheet: {len(master_data['tanks'])} tanks")

