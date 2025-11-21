# -*- coding: utf-8 -*-
"""
통합 엑셀 검증 스크립트

기능:
- 원본 vs 생성 파일 종합 비교
- 시트/컬럼/헤더/값/수식 검증
- 자동 최신 파일 감지
- 명령줄 옵션 지원

사용법:
    python verify_excel_generation.py [--quick] [--detailed] [--formulas] [--original PATH] [--generated PATH]
"""

import argparse
from pathlib import Path
from glob import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_latest_generated_file(output_dir="output", pattern="LCT_BUSHRA_AGI_TR_from_scratch*.xlsx"):
    """최신 생성 파일 자동 감지"""
    generated_files = glob(f"{output_dir}/{pattern}")
    if generated_files:
        return Path(max(generated_files, key=lambda p: Path(p).stat().st_mtime))
    return None


def verify_sheet_count(wb_orig, wb_gen, issues):
    """시트 수 검증"""
    print("\n[1] 시트 수 검증")
    print("-" * 80)
    if len(wb_orig.sheetnames) != len(wb_gen.sheetnames):
        print(f"❌ 불일치: 원본 {len(wb_orig.sheetnames)}개, 생성 {len(wb_gen.sheetnames)}개")
        issues.append("시트 수 불일치")
        return False
    else:
        print(f"✅ 일치: {len(wb_orig.sheetnames)}개")
        for sheet_name in wb_orig.sheetnames:
            print(f"   - {sheet_name}")
        return True


def verify_column_count(wb_orig, wb_gen, issues):
    """각 시트별 컬럼 수 검증"""
    print("\n[2] 컬럼 수 검증")
    print("-" * 80)
    all_match = True
    for sheet_name in wb_orig.sheetnames:
        if sheet_name not in wb_gen.sheetnames:
            print(f"❌ {sheet_name}: 생성된 파일에 없음")
            issues.append(f"{sheet_name} 시트 없음")
            all_match = False
            continue
        
        orig_cols = wb_orig[sheet_name].max_column
        gen_cols = wb_gen[sheet_name].max_column
        if orig_cols == gen_cols:
            print(f"✅ {sheet_name}: {orig_cols}열")
        else:
            print(f"❌ {sheet_name}: 원본 {orig_cols}열, 생성 {gen_cols}열")
            issues.append(f"{sheet_name} 컬럼 수 불일치")
            all_match = False
    return all_match


def verify_headers(wb_orig, wb_gen, issues, detailed=False):
    """헤더 비교"""
    print("\n[3] 헤더 검증")
    print("-" * 80)
    all_match = True
    
    # December_Tide_2025 헤더
    if "December_Tide_2025" in wb_orig.sheetnames:
        ws_tide_orig = wb_orig["December_Tide_2025"]
        ws_tide_gen = wb_gen["December_Tide_2025"]
        print("\nDecember_Tide_2025:")
        for col in [1, 2]:
            orig_val = ws_tide_orig.cell(1, col).value
            gen_val = ws_tide_gen.cell(1, col).value if ws_tide_gen.max_column >= col else None
            col_letter = get_column_letter(col)
            match = orig_val == gen_val
            status = "✅" if match else "❌"
            print(f"  {col_letter}1: {status}")
            if detailed or not match:
                print(f"    원본: {repr(orig_val)}")
                print(f"    생성: {repr(gen_val)}")
            if not match:
                issues.append(f"December_Tide_2025 {col_letter}1 헤더 불일치")
                all_match = False
    
    # Hourly_FWD_AFT_Heights 헤더
    if "Hourly_FWD_AFT_Heights" in wb_orig.sheetnames:
        ws_hourly_orig = wb_orig["Hourly_FWD_AFT_Heights"]
        ws_hourly_gen = wb_gen["Hourly_FWD_AFT_Heights"]
        print("\nHourly_FWD_AFT_Heights:")
        max_col = min(ws_hourly_orig.max_column, ws_hourly_gen.max_column, 15)
        for col in range(1, max_col + 1):
            orig_val = ws_hourly_orig.cell(1, col).value
            gen_val = ws_hourly_gen.cell(1, col).value
            col_letter = get_column_letter(col)
            match = orig_val == gen_val
            status = "✅" if match else "❌"
            if detailed or not match:
                print(f"  {col_letter}1: {status}")
                if detailed:
                    print(f"    원본: {repr(orig_val)}")
                    print(f"    생성: {repr(gen_val)}")
            if not match:
                issues.append(f"Hourly_FWD_AFT_Heights {col_letter}1 헤더 불일치")
                all_match = False
    
    # RORO_Stage_Scenarios 헤더
    if "RORO_Stage_Scenarios" in wb_orig.sheetnames:
        ws_roro_orig = wb_orig["RORO_Stage_Scenarios"]
        ws_roro_gen = wb_gen["RORO_Stage_Scenarios"]
        print("\nRORO_Stage_Scenarios:")
        max_col = min(ws_roro_orig.max_column, ws_roro_gen.max_column, 21)
        for col in range(1, max_col + 1):
            orig_val = ws_roro_orig.cell(14, col).value
            gen_val = ws_roro_gen.cell(14, col).value if ws_roro_gen.max_column >= col else None
            col_letter = get_column_letter(col)
            match = orig_val == gen_val
            if orig_val or gen_val:  # 둘 중 하나라도 값이 있으면
                status = "✅" if match else "❌"
                if detailed or not match:
                    print(f"  {col_letter}14: {status}")
                    if detailed:
                        print(f"    원본: {repr(orig_val)}")
                        print(f"    생성: {repr(gen_val)}")
                if not match:
                    issues.append(f"RORO_Stage_Scenarios {col_letter}14 헤더 불일치")
                    all_match = False
    
    return all_match


def verify_calc_values(wb_orig, wb_gen, issues):
    """Calc 시트 값 검증"""
    print("\n[4] Calc 시트 값 검증")
    print("-" * 80)
    if "Calc" not in wb_orig.sheetnames:
        print("⚠️ Calc 시트 없음")
        return True
    
    ws_calc_orig = wb_orig["Calc"]
    ws_calc_gen = wb_gen["Calc"]
    all_match = True
    
    # 주요 파라미터 행들
    param_rows = [5, 6, 7, 8, 10, 11, 12, 14, 15, 16, 17]
    for row in param_rows:
        orig_val = ws_calc_orig.cell(row, 5).value
        gen_val = ws_calc_gen.cell(row, 5).value
        param = ws_calc_orig.cell(row, 3).value
        if orig_val != gen_val:
            print(f"❌ 행 {row} ({param}): 원본={orig_val}, 생성={gen_val}")
            issues.append(f"Calc 행 {row} ({param}) 값 불일치")
            all_match = False
        else:
            print(f"✅ 행 {row} ({param}): {orig_val}")
    
    return all_match


def verify_roro_values(wb_orig, wb_gen, issues, detailed=False):
    """RORO_Stage_Scenarios 시트 값 검증"""
    print("\n[5] RORO_Stage_Scenarios 시트 값 검증")
    print("-" * 80)
    if "RORO_Stage_Scenarios" not in wb_orig.sheetnames:
        print("⚠️ RORO_Stage_Scenarios 시트 없음")
        return True
    
    ws_roro_orig = wb_orig["RORO_Stage_Scenarios"]
    ws_roro_gen = wb_gen["RORO_Stage_Scenarios"]
    all_match = True
    
    # C12 값 (X_Ballast)
    c12_orig = ws_roro_orig.cell(12, 3).value
    c12_gen = ws_roro_gen.cell(12, 3).value
    if c12_orig != c12_gen:
        print(f"❌ C12 (X_Ballast): 원본={c12_orig}, 생성={c12_gen}")
        issues.append("RORO_Stage_Scenarios C12 값 불일치")
        all_match = False
    else:
        print(f"✅ C12 (X_Ballast): {c12_orig}")
    
    # G열 값 (Trim_target_cm) - 모든 Stage
    print("\nG열 (Trim_target_cm) 값:")
    for row in range(15, 25):
        stage = ws_roro_orig.cell(row, 1).value
        orig_val = ws_roro_orig.cell(row, 7).value
        gen_val = ws_roro_gen.cell(row, 7).value if ws_roro_gen.max_column >= 7 else None
        if orig_val != gen_val:
            print(f"❌ 행 {row} ({stage}): 원본={orig_val}, 생성={gen_val}")
            issues.append(f"RORO_Stage_Scenarios 행 {row} G열 값 불일치")
            all_match = False
        elif detailed:
            print(f"✅ 행 {row} ({stage}): {orig_val}")
    
    # B, C 열 값 (W_stage_t, x_stage_m)
    if detailed:
        print("\nB, C 열 (W_stage_t, x_stage_m) 값:")
        for row in range(15, 25):
            stage = ws_roro_orig.cell(row, 1).value
            b_orig = ws_roro_orig.cell(row, 2).value
            b_gen = ws_roro_gen.cell(row, 2).value if ws_roro_gen.max_column >= 2 else None
            c_orig = ws_roro_orig.cell(row, 3).value
            c_gen = ws_roro_gen.cell(row, 3).value if ws_roro_gen.max_column >= 3 else None
            
            b_match = b_orig == b_gen
            c_match = c_orig == c_gen
            
            if not b_match or not c_match:
                print(f"❌ 행 {row} ({stage}):")
                if not b_match:
                    print(f"   B열: 원본={b_orig}, 생성={b_gen}")
                    issues.append(f"RORO_Stage_Scenarios 행 {row} B열 값 불일치")
                if not c_match:
                    print(f"   C열: 원본={c_orig}, 생성={c_gen}")
                    issues.append(f"RORO_Stage_Scenarios 행 {row} C열 값 불일치")
                all_match = False
    
    return all_match


def verify_hourly_values(wb_orig, wb_gen, issues):
    """Hourly_FWD_AFT_Heights 시트 값 검증"""
    print("\n[6] Hourly_FWD_AFT_Heights 시트 값 검증")
    print("-" * 80)
    if "Hourly_FWD_AFT_Heights" not in wb_orig.sheetnames:
        print("⚠️ Hourly_FWD_AFT_Heights 시트 없음")
        return True
    
    ws_hourly_orig = wb_orig["Hourly_FWD_AFT_Heights"]
    ws_hourly_gen = wb_gen["Hourly_FWD_AFT_Heights"]
    
    # N2 값 (설명 텍스트)
    n2_orig = ws_hourly_orig.cell(2, 14).value if ws_hourly_orig.max_column >= 14 else None
    n2_gen = ws_hourly_gen.cell(2, 14).value if ws_hourly_gen.max_column >= 14 else None
    
    if n2_orig != n2_gen:
        print(f"❌ N2 값 불일치")
        print(f"   원본: {repr(n2_orig)}")
        print(f"   생성: {repr(n2_gen)}")
        issues.append("Hourly_FWD_AFT_Heights N2 값 불일치")
        return False
    else:
        print(f"✅ N2 값 일치")
        return True


def verify_formulas(wb_orig, wb_gen, issues):
    """수식 비교 (기본적인 수식 존재 여부만 확인)"""
    print("\n[7] 수식 검증")
    print("-" * 80)
    print("⚠️ 수식 검증은 기본적인 존재 여부만 확인합니다.")
    print("   상세 수식 비교는 --formulas 옵션을 사용하세요.")
    
    # RORO_Stage_Scenarios의 주요 수식 셀 확인
    if "RORO_Stage_Scenarios" in wb_orig.sheetnames:
        ws_roro_orig = wb_orig["RORO_Stage_Scenarios"]
        ws_roro_gen = wb_gen["RORO_Stage_Scenarios"]
        
        # 행 15의 주요 수식 셀들
        formula_cols = [4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]  # D~R
        formula_count_orig = 0
        formula_count_gen = 0
        
        for col in formula_cols:
            if ws_roro_orig.max_column >= col:
                cell_orig = ws_roro_orig.cell(15, col)
                if cell_orig.value and isinstance(cell_orig.value, str) and cell_orig.value.startswith('='):
                    formula_count_orig += 1
            
            if ws_roro_gen.max_column >= col:
                cell_gen = ws_roro_gen.cell(15, col)
                if cell_gen.value and isinstance(cell_gen.value, str) and cell_gen.value.startswith('='):
                    formula_count_gen += 1
        
        print(f"행 15 수식 개수: 원본 {formula_count_orig}개, 생성 {formula_count_gen}개")
        if formula_count_orig == formula_count_gen:
            print("✅ 수식 개수 일치")
            return True
        else:
            print("❌ 수식 개수 불일치")
            issues.append("RORO_Stage_Scenarios 수식 개수 불일치")
            return False
    
    return True


def main():
    parser = argparse.ArgumentParser(description="엑셀 생성 파일 검증")
    parser.add_argument("--quick", action="store_true", help="빠른 검증 (기본 검사만)")
    parser.add_argument("--detailed", action="store_true", help="상세 검증 (모든 값 출력)")
    parser.add_argument("--formulas", action="store_true", help="수식 상세 비교")
    parser.add_argument("--original", type=str, default="LCT_BUSHRA_AGI_TR.xlsx", help="원본 파일 경로")
    parser.add_argument("--generated", type=str, default=None, help="생성 파일 경로 (자동 감지 시 생략)")
    
    args = parser.parse_args()
    
    # 파일 경로 설정
    original_file = Path(args.original)
    if args.generated:
        generated_file = Path(args.generated)
    else:
        generated_file = find_latest_generated_file()
        if not generated_file:
            print("❌ 생성된 파일을 찾을 수 없습니다.")
            print("   output/ 디렉토리에서 LCT_BUSHRA_AGI_TR_from_scratch*.xlsx 파일을 찾을 수 없습니다.")
            return 1
    
    if not original_file.exists():
        print(f"❌ 원본 파일을 찾을 수 없습니다: {original_file}")
        return 1
    
    if not generated_file.exists():
        print(f"❌ 생성 파일을 찾을 수 없습니다: {generated_file}")
        return 1
    
    print("=" * 80)
    print("엑셀 생성 파일 검증")
    print("=" * 80)
    print(f"원본: {original_file}")
    print(f"생성: {generated_file}")
    if args.quick:
        print("모드: 빠른 검증")
    elif args.detailed:
        print("모드: 상세 검증")
    else:
        print("모드: 표준 검증")
    
    # 워크북 로드
    wb_orig = load_workbook(original_file, data_only=False)
    wb_gen = load_workbook(generated_file, data_only=False)
    
    issues = []
    all_match = True
    
    # 검증 실행
    all_match &= verify_sheet_count(wb_orig, wb_gen, issues)
    all_match &= verify_column_count(wb_orig, wb_gen, issues)
    all_match &= verify_headers(wb_orig, wb_gen, issues, detailed=args.detailed)
    all_match &= verify_calc_values(wb_orig, wb_gen, issues)
    all_match &= verify_roro_values(wb_orig, wb_gen, issues, detailed=args.detailed)
    all_match &= verify_hourly_values(wb_orig, wb_gen, issues)
    
    if args.formulas:
        all_match &= verify_formulas(wb_orig, wb_gen, issues)
    
    wb_orig.close()
    wb_gen.close()
    
    # 결과 출력
    print("\n" + "=" * 80)
    if all_match:
        print("✅ 모든 검증 통과!")
        return 0
    else:
        print(f"❌ 검증 실패: {len(issues)}개 문제 발견")
        print("\n발견된 문제:")
        for issue in issues:
            print(f"  - {issue}")
        return 1


if __name__ == "__main__":
    exit(main())

