# -*- coding: utf-8 -*-
# update_stage_values.py
# Update Stage W and x values in RORO_Stage_Scenarios sheet

import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excel 파일 경로 (상위 폴더 기준)
EXCEL_FILE = "../../output/LCT_BUSHRA_GateAB_v4_HYBRID.xlsx"  # Updated to v4 HYBRID

# 권장 1차 입력값 (midship=0, LCF=+29.29m 기준, Verified from Stability Book)
STAGE_VALUES = {
    1: {"W": 0, "x": None, "desc": "Empty condition"},
    2: {"W": 65, "x": -10, "desc": "SPMT 1st entry on ramp (약 30% 반력)"},
    3: {"W": 110, "x": -5, "desc": "~50% on ramp (반력 절반)"},
    4: {"W": 217, "x": -2, "desc": "Full on ramp / break-even (1기 전중량)"},
    5: {"W": 434, "x": 32.5, "desc": "Deck full load (217t × 2) → 합성중심"},
}

# 입력 셀 위치 (row0=13이므로 Stage 1은 row 14부터)
STAGE_ROWS = {
    1: 14,
    2: 15,
    3: 16,
    4: 17,
    5: 18,
}

def update_stage_values():
    """Update stage values in Excel file"""
    if not os.path.exists(EXCEL_FILE):
        print(f"✗ ERROR: Excel file not found: {EXCEL_FILE}")
        print(f"  Please run patch3.py first to generate the Excel file.")
        return False
    
    try:
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        
        if "RORO_Stage_Scenarios" not in wb.sheetnames:
            print(f"✗ ERROR: RORO_Stage_Scenarios sheet not found")
            wb.close()
            return False
        
        roro = wb["RORO_Stage_Scenarios"]
        input_fill = PatternFill("solid", fgColor="FFF2CC")  # Yellow fill
        
        print("\n" + "="*70)
        print("Updating Stage Values in RORO_Stage_Scenarios Sheet")
        print("="*70)
        print(f"File: {EXCEL_FILE}\n")
        
        # Update each stage
        for stage_num in [1, 2, 3, 4, 5]:
            row = STAGE_ROWS[stage_num]
            values = STAGE_VALUES[stage_num]
            
            # Update W_stage_t (column B)
            wb["RORO_Stage_Scenarios"].cell(row=row, column=2, value=values["W"])
            wb["RORO_Stage_Scenarios"].cell(row=row, column=2).fill = input_fill
            
            # Update x_stage_m (column C)
            if values["x"] is not None:
                wb["RORO_Stage_Scenarios"].cell(row=row, column=3, value=values["x"])
                wb["RORO_Stage_Scenarios"].cell(row=row, column=3).fill = input_fill
            else:
                # Clear cell if None
                wb["RORO_Stage_Scenarios"].cell(row=row, column=3, value="")
                wb["RORO_Stage_Scenarios"].cell(row=row, column=3).fill = input_fill
            
            # Print update
            x_str = f"{values['x']:.1f}" if values["x"] is not None else "–"
            print(f"Stage {stage_num}: W={values['W']} t, x={x_str} m  ({values['desc']})")
        
        # Save workbook
        wb.save(EXCEL_FILE)
        wb.close()
        
        print("\n" + "="*70)
        print("✓ Successfully updated stage values!")
        print("="*70)
        print(f"\nFile saved: {EXCEL_FILE}")
        print("\nNext steps:")
        print("1. Open the Excel file and verify the values")
        print("2. Check calculated columns (TM, Trim, Dfwd, Daft, Heights)")
        print("3. Review Trim_Check column for any 'EXCESSIVE' warnings")
        print("4. Adjust values if needed based on actual stowage plan coordinates")
        
        return True
    
    except PermissionError:
        print(f"✗ ERROR: Cannot save {EXCEL_FILE}. Close the file if open.")
        return False
    except Exception as e:
        print(f"✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

def print_summary():
    """Print summary table"""
    print("\n" + "="*70)
    print("STAGE VALUES SUMMARY")
    print("="*70)
    print(f"{'Stage':<10} {'W_stage_t (t)':<15} {'x_stage_m (m)':<15} {'Description':<40}")
    print("-"*70)
    
    for stage_num in [1, 2, 3, 4, 5]:
        values = STAGE_VALUES[stage_num]
        x_str = f"{values['x']:.1f}" if values["x"] is not None else "–"
        print(f"Stage {stage_num:<4} {values['W']:<15} {x_str:<15} {values['desc']:<40}")
    
    print("="*70)
    print("\nCoordinate System:")
    print("  - midship = 0")
    print("  - LCF = +29.29 m (aft direction +, Verified from Stability Book)")
    print("  - Forward direction: negative x")
    print("  - Aft direction: positive x")

def main():
    """Main execution"""
    print_summary()
    
    # Ask for confirmation
    print("\n" + "="*70)
    response = input("Update Excel file with these values? (y/n): ").strip().lower()
    
    if response == 'y' or response == 'yes':
        update_stage_values()
    else:
        print("Update cancelled.")

if __name__ == "__main__":
    main()

