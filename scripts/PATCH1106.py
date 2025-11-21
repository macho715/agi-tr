# 수정된: LCT BUSHRA — FWD/AFT 리포트 생성기 (견고화 버전)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

src_path = "LCT_BUSHRA_Package_RORO_FIXED.xlsx"
out_path = "Bushra_FWD_AFT_Report_NOW.xlsx"

# 파일 존재 / 로드 체크
try:
    wb = load_workbook(src_path, data_only=True)
except FileNotFoundError:
    raise FileNotFoundError(f"Source workbook not found: {src_path}")

# Helper: safe get sheet by substring
def get_ws(name_contains):
    for n in wb.sheetnames:
        if name_contains.lower() in n.lower():
            return wb[n]
    return None

calc = get_ws("Calc")
stage = get_ws("Stage_Heights") or get_ws("Stage")
tide = get_ws("December_Tide_2025") or get_ws("Tide")
hourly = get_ws("Hourly_FWD_AFT_Heights")

if calc is None:
    raise RuntimeError("Calc sheet not found")
if stage is None:
    raise RuntimeError("Stage_Heights (or Stage) sheet not found")

# --- Read constants from Calc (존재 여부만 체크)
L_ramp = calc["D4"].value
theta_max = calc["D5"].value
KminusZ = calc["D6"].value
min_fwd = calc["D9"].value if calc["D9"].value is not None else None
max_fwd = calc["D10"].value if calc["D10"].value is not None else None

# --- Build tide map (if available)
tide_map = {}
if tide is not None:
    for r in range(2, tide.max_row + 1):
        dt = tide.cell(row=r, column=1).value
        tv = tide.cell(row=r, column=2).value
        if dt is None or tv is None:
            continue
        # openpyxl often returns datetime already; accept that
        if isinstance(dt, str):
            # 유연한 파싱: 시도할 포맷 목록
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                try:
                    dt_parsed = datetime.strptime(dt, fmt)
                    dt = dt_parsed
                    break
                except Exception:
                    continue
            else:
                # 파싱 불가면 건너뜀
                continue
        if isinstance(dt, datetime):
            tide_map[dt] = tv

# --- 헤더 행 자동탐색 (첫 10행 내에서 'stage' 포함 행 찾기)
hdr_row = 1
headers = {}
found_header = False
search_rows = min(10, stage.max_row)
for rr in range(1, search_rows + 1):
    tmp = {}
    for c in range(1, stage.max_column + 1):
        v = stage.cell(row=rr, column=c).value
        if v:
            tmp[str(v).strip().lower()] = c
    if any("stage" in k for k in tmp.keys()):
        headers = tmp
        hdr_row = rr
        found_header = True
        break

if not found_header:
    # fallback: 1행 사용 (원래 코드와 호환)
    hdr_row = 1
    for c in range(1, stage.max_column + 1):
        v = stage.cell(row=hdr_row, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

def col(key):
    key = key.lower()
    for k, c in headers.items():
        if key in k:
            return c
    return None

col_stage = col("stage")
col_name  = col("stage name") or col("name")
col_time  = col("reference time") or col("time") or col("ref")
col_fwd   = col("fwd draft") or col("fwd")
col_aft   = col("aft draft") or col("aft")
col_trim  = col("trim")
col_angle = col("ramp angle") or col("angle")
col_status= col("status") or col("ok")

# --- Read stage rows
rows = []
for r in range(hdr_row + 1, stage.max_row + 1):
    stg = stage.cell(row=r, column=col_stage).value if col_stage else None
    nm  = stage.cell(row=r, column=col_name).value if col_name else ""
    ref = stage.cell(row=r, column=col_time).value if col_time else None
    # skip empty rows (both stage and name empty)
    if stg is None and (nm is None or str(nm).strip() == ""):
        continue
    fwd = stage.cell(row=r, column=col_fwd).value if col_fwd else None
    aft = stage.cell(row=r, column=col_aft).value if col_aft else None
    trm = stage.cell(row=r, column=col_trim).value if col_trim else None
    ang = stage.cell(row=r, column=col_angle).value if col_angle else None
    sts = stage.cell(row=r, column=col_status).value if col_status else None

    # tide lookup: 허용되는 datetime 키로 맞춤
    tide_val = None
    if ref and tide_map:
        key = ref
        if isinstance(key, str):
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                try:
                    key = datetime.strptime(key, fmt)
                    break
                except Exception:
                    key = None
        if isinstance(key, datetime):
            tide_val = tide_map.get(key)

    rows.append((stg, nm or "", ref, tide_val, fwd, aft, trm, ang, sts or ""))

# --- Create output workbook
out = Workbook()
ws = out.active
ws.title = "FWD_AFT_Report"

title_font = Font(name="Calibri", size=16, bold=True)
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
thin = Side(border_style="thin", color="A0A0A0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header title
ws.merge_cells("A1:J1")
ws["A1"] = "LCT BUSHRA – FWD/AFT Report (Current Values)"
ws["A1"].font = title_font
ws["A1"].alignment = center

# Header info (K–Z, L_ramp, theta_max, Dfwd limits)
hdr_info = [
    ("K–Z (m, Chart Datum)", KminusZ),
    ("Linkspan Length L_ramp (m)", L_ramp),
    ("θ_max (deg)", theta_max),
    ("Dfwd Limits (m)", f"{min_fwd} – {max_fwd}" if (min_fwd is not None and max_fwd is not None) else "N/A"),
]
r0 = 3
for label, val in hdr_info:
    ws.cell(row=r0, column=1, value=label).font = Font(bold=True)
    ws.cell(row=r0, column=1).alignment = left
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=4)
    ws.cell(row=r0, column=2, value=val)
    r0 += 1

# Table header (10 cols)
table_start = r0 + 1
cols = ["Stage","Stage Name","Reference Time (GST)","Tide (m, CD)",
        "FWD Draft (m)","AFT Draft (m)","Trim (m)","Ramp Angle (°)","Status","Remarks"]
for i, h in enumerate(cols, start=1):
    c = ws.cell(row=table_start, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Table rows — REMARKS 칸까지 채움
rr = table_start + 1
for (stg, nm, ref, tide_val, fwd, aft, trm, ang, sts) in rows:
    # Remarks 기본 빈 문자열
    remarks = ""
    values = [stg, nm, ref, tide_val, fwd, aft, trm, ang, sts, remarks]
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=rr, column=i, value=v)
        # 포맷: 시간은 readable string, 숫자는 2자리 포맷
        if isinstance(v, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"
            cell.alignment = center if i != 2 else left
        elif isinstance(v, (int, float)):
            cell.number_format = "0.00"
            cell.alignment = center if i != 2 else left
        else:
            cell.alignment = center if i != 2 else left
        cell.border = border
    rr += 1

# Column widths
widths = [8, 26, 22, 12, 14, 14, 12, 14, 12, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Save and report
out.save(out_path)
print(f"Report saved: {out_path}")
