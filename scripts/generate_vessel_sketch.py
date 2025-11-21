# -*- coding: utf-8 -*-
"""
LCT BUSHRA 선박 측면도 스케치 생성 (PDF Elevation View 스타일)
Mammoet DWG 업데이트용 FWD/AFT Height 표시
"""

import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import FancyBboxPatch, Rectangle, FancyArrowPatch
import numpy as np
from openpyxl import load_workbook
import os

# 한글 폰트 설정 (필요 시)
# plt.rcParams['font.family'] = 'Malgun Gothic'  # Windows
# plt.rcParams['font.family'] = 'AppleGothic'  # macOS

def load_height_data(excel_path):
    """Excel에서 Stage별 높이 데이터 읽기"""
    try:
        wb = load_workbook(excel_path)
        if "RoRo_Height_Report" not in wb.sheetnames:
            print(f"Warning: RoRo_Height_Report sheet not found in {excel_path}")
            return None

        ws = wb["RoRo_Height_Report"]
        stages = []

        # 헤더는 12행, 데이터는 13행부터
        for row in range(13, 18):  # 5개 스테이지
            stage = ws.cell(row=row, column=1).value
            if stage:
                stages.append({
                    "stage": stage,
                    "desc": ws.cell(row=row, column=2).value or "",
                    "fwd_draft": ws.cell(row=row, column=6).value or 0,
                    "aft_draft": ws.cell(row=row, column=7).value or 0,
                    "fwd_height": ws.cell(row=row, column=10).value or 0,
                    "aft_height": ws.cell(row=row, column=11).value or 0,
                    "tide": ws.cell(row=row, column=8).value or 0.5,
                })
        wb.close()
        return stages
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return None


def create_vessel_sketch(stage_data, output_path, stage_name="Stage 1"):
    """
    선박 측면도 스케치 생성 (PDF Elevation View 스타일)

    Parameters:
    - stage_data: Stage별 높이 데이터 (dict)
    - output_path: 출력 파일 경로
    - stage_name: Stage 이름 (예: "Stage 1", "Stage 2")
    """
    # Vessel Dimensions (PDF Page 3)
    LENGTH = 60.302  # m (Verified: Length between perpendiculars)
    BREADTH = 14.60  # m
    DEPTH = 3.65  # m (Keel ~ Deck) - LCT Bushra Moulded Depth (verified)
    # Source: RoRo Simulation_stowage plan_20251103.pdf (LCT SPECIFICATION: DEPTH (m) 3.65)
    #         Vessel_Stability_Booklet.pdf (Principal Particulars: Moulded Depth 3.65 m)
    #         Cross-verified: 5/5 documents match

    # Linkspan
    LINKSPAN_LENGTH = 12.00  # m (12000mm)

    # Sea Bed ~ WL (PDF Stage 2)
    SEA_BED_DEPTH = 2.40  # m (2400mm)

    # Stage 데이터에서 Draft/Height 추출
    if stage_data:
        fwd_draft = stage_data.get("fwd_draft", 2.0)
        aft_draft = stage_data.get("aft_draft", 2.0)
        fwd_height = stage_data.get("fwd_height", 3.35)
        aft_height = stage_data.get("aft_height", 3.35)
        tide = stage_data.get("tide", 0.5)
    else:
        # 기본값 (Initial)
        fwd_draft = 2.0
        aft_draft = 2.0
        fwd_height = 3.35
        aft_height = 3.35
        tide = 0.5

    # Waterline (WL) 기준 계산
    wl_level = 0  # 기준선
    keel_level = -SEA_BED_DEPTH  # Sea Bed 기준
    deck_level_fwd = wl_level + fwd_height
    deck_level_aft = wl_level + aft_height

    # Figure 생성
    fig, ax = plt.subplots(1, 1, figsize=(16, 10))

    # 배경 설정
    ax.set_facecolor('white')

    # Sea Bed 표시
    sea_bed = Rectangle(
        (0, keel_level - 0.5),
        LENGTH,
        0.5,
        facecolor='lightblue',
        edgecolor='navy',
        linewidth=2,
        alpha=0.3,
        label='Sea Bed'
    )
    ax.add_patch(sea_bed)
    ax.text(LENGTH / 2, keel_level - 0.25, f'Sea Bed (2400mm)',
            ha='center', va='center', fontsize=10, style='italic')

    # Waterline (WL) 표시
    ax.axhline(y=wl_level, color='blue', linestyle='--', linewidth=2, label='Waterline (WL)')
    ax.text(LENGTH + 2, wl_level, 'WL', fontsize=12, color='blue', fontweight='bold')

    # 선박 Hull (측면도)
    # Keel line
    keel_y = wl_level - (fwd_draft + aft_draft) / 2  # 평균 draft
    trim = fwd_draft - aft_draft  # Trim (FWD > AFT이면 +)

    # 선박 외형 (사다리꼴로 간소화)
    hull_points = [
        (0, keel_y),  # FWD Keel
        (0, deck_level_fwd),  # FWD Deck
        (LENGTH, deck_level_aft),  # AFT Deck
        (LENGTH, keel_y),  # AFT Keel
    ]

    hull = patches.Polygon(
        hull_points,
        facecolor='lightgray',
        edgecolor='black',
        linewidth=2,
        alpha=0.7,
        label='LCT Bushra Hull'
    )
    ax.add_patch(hull)

    # Deck line 표시
    ax.plot([0, LENGTH], [deck_level_fwd, deck_level_aft],
            'k-', linewidth=3, label='Deck Level')

    # FWD/AFT Height 표시 (수직선 + 텍스트)
    # FWD
    ax.plot([0, 0], [wl_level, deck_level_fwd],
            'r-', linewidth=2, linestyle=':', label='FWD Height')
    ax.text(-2, (wl_level + deck_level_fwd) / 2,
            f'FWD Height\n{fwd_height:.2f}m',
            ha='right', va='center', fontsize=11,
            bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.7),
            fontweight='bold')

    # AFT
    ax.plot([LENGTH, LENGTH], [wl_level, deck_level_aft],
            'g-', linewidth=2, linestyle=':', label='AFT Height')
    ax.text(LENGTH + 2, (wl_level + deck_level_aft) / 2,
            f'AFT Height\n{aft_height:.2f}m',
            ha='left', va='center', fontsize=11,
            bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.7),
            fontweight='bold')

    # Draft 표시
    ax.plot([0, 0], [keel_y, wl_level],
            'b-', linewidth=1.5, linestyle='--', alpha=0.5)
    ax.text(-2, (keel_y + wl_level) / 2,
            f'FWD Draft\n{fwd_draft:.2f}m',
            ha='right', va='center', fontsize=9, color='blue', style='italic')

    ax.plot([LENGTH, LENGTH], [keel_y, wl_level],
            'b-', linewidth=1.5, linestyle='--', alpha=0.5)
    ax.text(LENGTH + 2, (keel_y + wl_level) / 2,
            f'AFT Draft\n{aft_draft:.2f}m',
            ha='left', va='center', fontsize=9, color='blue', style='italic')

    # Linkspan 연결점 표시 (FWD Deck, 12m)
    linkspan_start = 0
    linkspan_end = linkspan_start + LINKSPAN_LENGTH
    linkspan_y = deck_level_fwd

    # Linkspan (간단한 직선)
    ax.plot([linkspan_start, linkspan_end],
            [linkspan_y, linkspan_y + 0.1],  # 약간의 경사
            'orange', linewidth=4, label='Linkspan (12m)')
    ax.text(linkspan_start + LINKSPAN_LENGTH / 2, linkspan_y + 0.2,
            'Linkspan 12m', ha='center', va='bottom', fontsize=10,
            bbox=dict(boxstyle='round', facecolor='orange', alpha=0.5))

    # RoRo 위치 표시 (Transformer, PDF Page 3)
    # Transformer TMG3: 4.20m W x 5.80m L x 4.59m H
    transformer_x = 5  # FWD에서 5m 위치
    transformer_y = deck_level_fwd
    transformer_w = 5.80
    transformer_h = 4.59

    transformer = Rectangle(
        (transformer_x, transformer_y),
        transformer_w,
        transformer_h,
        facecolor='red',
        edgecolor='darkred',
        linewidth=2,
        alpha=0.6,
        label='Transformer TMG3'
    )
    ax.add_patch(transformer)
    ax.text(transformer_x + transformer_w / 2, transformer_y + transformer_h / 2,
            'Transformer\n217t', ha='center', va='center', fontsize=9,
            fontweight='bold', color='white')

    # RoRo Plates (6m x 2m x 0.7m)
    roro_plate = Rectangle(
        (transformer_x - 1, transformer_y - 0.7),
        6,
        0.7,
        facecolor='brown',
        edgecolor='#654321',  # darkbrown → hex color
        linewidth=1.5,
        alpha=0.7,
        label='RoRo Plate'
    )
    ax.add_patch(roro_plate)

    # Vessel Dimensions 텍스트
    dim_text = f"""Vessel Dimensions (PDF Page 3):
Length: {LENGTH} m
Breadth: {BREADTH} m
Depth (D): {DEPTH} m

Stage: {stage_name}
FWD Draft: {fwd_draft:.2f} m
AFT Draft: {aft_draft:.2f} m
Trim: {trim:.2f} m (+Aft)
Tide: {tide:.2f} m

FWD Height: {fwd_height:.2f} m (Deck level from WL)
AFT Height: {aft_height:.2f} m (Deck level from WL)"""

    ax.text(LENGTH + 5, deck_level_fwd + 2, dim_text,
            fontsize=9, va='top', ha='left',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

    # 축 설정
    ax.set_xlim(-5, LENGTH + 10)
    ax.set_ylim(keel_level - 1, deck_level_fwd + 6)
    ax.set_xlabel('Distance from FWD (m)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Height from Waterline (m)', fontsize=12, fontweight='bold')
    ax.set_title(f'LCT BUSHRA — {stage_name} Elevation View (For Mammoet DWG Update)',
                 fontsize=14, fontweight='bold', pad=20)

    # 그리드
    ax.grid(True, linestyle='--', alpha=0.3)
    ax.axhline(y=0, color='k', linewidth=0.5)
    ax.axvline(x=0, color='k', linewidth=0.5)

    # 범례
    ax.legend(loc='upper left', fontsize=9, framealpha=0.9)

    # PDF 참조 노트
    note_text = "PDF Reference: RoRo Simulation Stowage Plan 2025-11-03\nGeneral Notes 6: LCT Captain to advise final height at FWD & AFT"
    ax.text(LENGTH / 2, keel_level - 0.8, note_text,
            ha='center', va='top', fontsize=8, style='italic',
            bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.7))

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"Vessel sketch saved: {output_path}")
    plt.close()


def main():
    """메인 실행 함수"""
    # 경로 설정
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "..", "output", "LCT_BUSHRA_GateAB_v4_HYBRID_generated.xlsx")
    output_dir = os.path.join(script_dir, "..", "output")

    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)

    # Excel에서 데이터 로드
    stages = load_height_data(excel_path)

    if not stages:
        print("Warning: Using default values (no Excel data found)")
        stages = [None] * 2  # Stage 1, Stage 2

    # Stage별 스케치 생성
    stage_names = ["Stage 1", "Stage 2"]
    for idx, stage_name in enumerate(stage_names):
        stage_data = stages[idx] if stages and idx < len(stages) else None
        output_path = os.path.join(output_dir, f"vessel_sketch_{stage_name.lower().replace(' ', '_')}.png")
        create_vessel_sketch(stage_data, output_path, stage_name)

    print("\n" + "=" * 80)
    print("Vessel sketch generation completed!")
    print(f"Output directory: {output_dir}")
    print("=" * 80)


if __name__ == "__main__":
    main()

