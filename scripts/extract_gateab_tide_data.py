# -*- coding: utf-8 -*-
# Bushra_GateAB_Updated_v3.xlsx에서 조수 데이터 추출

from openpyxl import load_workbook
import json

print("=" * 80)
print("GateAB v3 조수 데이터 추출")
print("=" * 80)

wb = load_workbook("../backup/Bushra_GateAB_Updated_v3.xlsx", data_only=True)
tide_sheet = wb["December_Tide_2025"]

tide_data = []
for row in range(2, tide_sheet.max_row + 1):
    datetime_val = tide_sheet.cell(row, 1).value
    tide_val = tide_sheet.cell(row, 2).value
    if datetime_val and tide_val:
        tide_data.append(
            {
                "datetime": str(datetime_val),
                "tide_m": (
                    float(tide_val) if isinstance(tide_val, (int, float)) else tide_val
                ),
            }
        )

print(f"\n추출된 조수 데이터: {len(tide_data)}개")
print(f"시작: {tide_data[0] if tide_data else 'N/A'}")
print(f"종료: {tide_data[-1] if tide_data else 'N/A'}")

# JSON으로 저장
with open("../data/gateab_v3_tide_data.json", "w", encoding="utf-8") as f:
    json.dump(tide_data, f, indent=2, ensure_ascii=False)

print(f"\n✓ 조수 데이터 저장: ../data/gateab_v3_tide_data.json")
print("=" * 80)
