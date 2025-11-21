# -*- coding: utf-8 -*-
# build_bushra_v4_standalone.py
# 독립 실행 가능한 버전 - JSON 의존성 없음
# Generates: LCT_BUSHRA_GateAB_v4_HYBRID_generated.xlsx

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import math

out_path = "LCT_BUSHRA_GateAB_v4_HYBRID_generated.xlsx"

wb = Workbook()

# Styles
title_font = Font(name="Calibri", size=14, bold=True)
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
sec_font = Font(name="Calibri", size=11, bold=True)
note_fill = PatternFill("solid", fgColor="F2F2F2")
header_fill = PatternFill("solid", fgColor="1F4E78")
input_fill = PatternFill("solid", fgColor="FFF2CC")
warning_fill = PatternFill("solid", fgColor="FFFF00")
error_fill = PatternFill("solid", fgColor="FF0000")
pass_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
thin = Side(border_style="thin", color="C0C0C0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------- 1) Calc sheet ----------
ws = wb.active
ws.title = "Calc"

# Title
ws["A1"] = "LCT BUSHRA — RORO FWD/AFT Draft Calculator v4 INTEGRATED (Core Constants)"
ws["A1"].font = title_font
ws.merge_cells("A1:E1")

# Coordinate Standard (최우선 명시)
ws["A2"] = "⚠️ COORDINATE STANDARD (READ FIRST):"
ws["A2"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
ws["A2"].fill = note_fill
ws.merge_cells("A2:E2")

ws["A3"] = "x_stage: Distance from midship (m). Negative = forward, Positive = aft"
ws["A3"].fill = note_fill
ws.merge_cells("A3:E3")

ws["A4"] = (
    "LCF: Distance from midship (m) - MUST match Stability Booklet basis. If your Booklet uses FP or AP reference, convert to midship first!"
)
ws["A4"].fill = note_fill
ws.merge_cells("A4:E4")

# Header row
r = 6
headers = ["SECTION", "PARAMETER", "UNIT", "VALUE", "NOTES"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.border = border
    c.alignment = center

# 셀 매핑 (patch.md 기준): Row 7 → D8, Row 8 → D9, Row 9 → D10, ...
# 실제 D열 셀: D8, D9, D10, D11, D12, D13, D14, (D15 빈행), D16, D17, D18
mapping = [
    ("INPUT CONSTANTS", "", None, "", True),  # Section header
    ("L_ramp_m", "m", 12, "Linkspan length (Mammoet)", False),  # Row 7 → D8
    ("theta_max_deg", "deg", 6, "Max ramp angle (Harbour Master)", False),  # Row 8 → D9
    ("KminusZ_m", "m", 3, "K - Z (SITE MEASURE) ⚠️", False),  # Row 9 → D10
    ("D_vessel_m", "m", 3.65, 
     "LCT Bushra Moulded Depth: 3.65m (verified: RoRo Simulation & Stability Booklet, 5/5 documents match)", 
     False),  # Row 10 → D11
    ("LIMITS & OPS", "", None, "", True),  # Section header
    ("min_fwd_draft_m", "m", 1.5, "Operational lower limit", False),  # Row 11 → D12
    ("max_fwd_draft_m", "m", 3.5, "Operational upper limit", False),  # Row 12 → D13
    ("pump_rate_tph", "t/h", 10, "Ballast pump rate (t/h)", False),  # Row 13 → D14
    ("", "", None, "", False),  # spacer (Row 14 → D15 빈행)
    ("STABILITY", "", None, "", True),  # Section header
    (
        "MTC_t_m_per_m",
        "t·m/m",
        4072,  # Verified: 40.72 t·m/cm * 100 = 4072 t·m/m (from Stability Book, Draft ~2.50m)
        "MTC from Stability Book (cm to m: *100)",
        False,
    ),  # Row 15 → D16
    ("LCF_m_from_midship", "m", 29.29, "LCF from midship (Verified from Stability Book, Draft ~2.50m)", False),  # Row 16 → D17
    (
        "TPC_t_per_cm",
        "t/cm",
        None,
        "Tonnes per cm immersion (opt)",
        False,
    ),  # Row 17 → D18
]

r = 7
for param, unit, val, note, is_header in mapping:
    if is_header:
        ws.cell(row=r, column=1, value=param).font = sec_font
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
    elif param:
        ws.cell(row=r, column=2, value=param)
        ws.cell(row=r, column=3, value=unit).alignment = center
        vcell = ws.cell(row=r, column=4, value=val)
        if val is not None:
            vcell.fill = input_fill
            vcell.protection = Protection(locked=False)
        ws.cell(row=r, column=5, value=note)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
    else:
        # 빈 행
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
    r += 1

# Critical notes
ws.cell(
    row=r,
    column=1,
    value="⚠ CRITICAL: Enter K-Z measured on site into Calc!D10 (KminusZ_m).",
)
ws.cell(
    row=r + 1,
    column=1,
    value="Ensure MTC/LCF (Calc!D16/D17) match BV stability booklet; use midship ref.",
)
for rr in range(r, r + 2):
    ws.cell(row=rr, column=1).fill = note_fill
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)

for i, wid in enumerate([26, 24, 10, 12, 70], 1):
    ws.column_dimensions[get_column_letter(i)].width = wid

# 셀 보호: 입력 셀만 언락 (D8~D18, Row 7~17)
for row_idx in range(7, 18):
    if ws.cell(row=row_idx, column=4).value is not None:
        ws.cell(row=row_idx, column=4).protection = Protection(locked=False)
ws.protection.sheet = True
ws.protection.selectLockedCells = True
ws.protection.selectUnlockedCells = True

# ---------- 2) December_Tide_2025 (template: 744 rows) ----------
tide = wb.create_sheet("December_Tide_2025")
tide["A1"] = "datetime_gst"
tide["B1"] = "tide_m (Chart Datum)"
for c in ("A1", "B1"):
    tide[c].font = hdr_font
    tide[c].fill = header_fill
    tide[c].alignment = center
    tide[c].border = border

start = datetime(2025, 12, 1, 0, 0)
for i in range(744):
    row = i + 2
    tide.cell(
        row=row, column=1, value=(start + timedelta(hours=i)).strftime("%Y-%m-%d %H:%M")
    )
    tide.cell(row=row, column=2, value="")
    for c in range(1, 3):
        tide.cell(row=row, column=c).border = border
tide.column_dimensions["A"].width = 22
tide.column_dimensions["B"].width = 12

# ---------- 3) Hourly_FWD_AFT_Heights ----------
out = wb.create_sheet("Hourly_FWD_AFT_Heights")
hdrs = [
    "DateTime (GST)",
    "Tide_m",
    "Dfwd_req_m",
    "Daft_req_m",
    "Status",
    "Actual_Dfwd_m",
    "Actual_Daft_m",
    "Ramp_Angle_deg",
    "Actual_Angle_deg",
    "FWD_Height_m",
    "AFT_Height_m",
    "Notes",
]
for i, h in enumerate(hdrs, 1):
    c = out.cell(row=1, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Build 744 rows with formulas (셀 매핑: D8, D9, D10, D12, D13, D14)
for i in range(2, 746):
    out.cell(
        row=i,
        column=1,
        value=f'=IF(December_Tide_2025!A{i}="","",December_Tide_2025!A{i})',
    )
    out.cell(
        row=i,
        column=2,
        value=f'=IF(December_Tide_2025!B{i}="","",December_Tide_2025!B{i})',
    )
    # Dfwd_req = KminusZ + Tide - L_ramp * TAN(RADIANS(theta_max))
    # Calc!D10 = KminusZ, Calc!D8 = L_ramp, Calc!D9 = theta_max
    out.cell(
        row=i,
        column=3,
        value=f'=IF(A{i}="","",Calc!$D$10 + B{i} - Calc!$D$8 * TAN(RADIANS(Calc!$D$9)))',
    )
    out.cell(row=i, column=4, value=f'=IF(C{i}="","",C{i})')
    # Ramp_Angle = DEGREES(ATAN((KminusZ - Dfwd + Tide) / L_ramp))
    out.cell(
        row=i,
        column=8,
        value=f'=IF(C{i}="","",DEGREES(ATAN((Calc!$D$10 - C{i} + B{i})/Calc!$D$8)))',
    )
    out.cell(
        row=i,
        column=9,
        value=f'=IF(OR(F{i}="",Calc!$D$8=0),"",DEGREES(ATAN((Calc!$D$10 - F{i} + B{i})/Calc!$D$8)))',
    )
    # Status = IF(AND(Dfwd >= min, Dfwd <= max, Angle <= theta_max), "OK", "CHECK")
    # Calc!D12 = min_draft, Calc!D13 = max_draft, Calc!D9 = theta_max
    out.cell(
        row=i,
        column=5,
        value=f'=IF(C{i}="","",IF(AND(C{i}>=Calc!$D$12, C{i}<=Calc!$D$13, H{i}<=Calc!$D$9),"OK","CHECK"))',
    )
    out.cell(row=i, column=6, value="")  # Actual Dfwd input
    out.cell(row=i, column=7, value="")  # Actual Daft input
    # Height = Depth - Draft + Tide (Calc!D11 = Depth, C/D = Draft, B = Tide)
    out.cell(
        row=i,
        column=10,
        value=f'=IF(OR(C{i}="",Calc!$D$11=""),"",Calc!$D$11 - C{i} + B{i})',
    )  # FWD_Height_m
    out.cell(
        row=i,
        column=11,
        value=f'=IF(OR(D{i}="",Calc!$D$11=""),"",Calc!$D$11 - D{i} + B{i})',
    )  # AFT_Height_m
    out.cell(row=i, column=12, value="")  # Notes
    for c in range(1, 13):
        out.cell(row=i, column=c).border = border

for col, w in zip(range(1, 13), [22, 10, 12, 12, 10, 14, 14, 14, 14, 12, 12, 28]):
    out.column_dimensions[get_column_letter(col)].width = w

# Conditional formatting
try:
    red_rule = FormulaRule(formula=["$H2>Calc!$D$9"], fill=error_fill, stopIfTrue=True)
    out.conditional_formatting.add("H2:H745", red_rule)
    yellow_rule = FormulaRule(
        formula=['$E2="CHECK"'], fill=warning_fill, stopIfTrue=True
    )
    out.conditional_formatting.add("A2:L745", yellow_rule)
except Exception:
    pass

# ---------- 4) RORO_Stage_Scenarios ----------
roro = wb.create_sheet("RORO_Stage_Scenarios")
roro["A1"] = "RORO STAGE-BY-STAGE LOADING ANALYSIS"
roro["A1"].font = title_font
roro.merge_cells("A1:L1")

roro["A3"] = "INPUTS (yellow)"
roro["A3"].font = sec_font
roro["A4"] = "Tmean baseline (m)"
roro["C4"] = 2.33
roro["C4"].fill = input_fill
roro["C4"].border = border

roro["A6"] = "CONSTANTS (from Calc sheet)"
roro["A6"].font = sec_font
# 셀 매핑 정정: D16=MTC, D17=LCF, D18=TPC, D14=pump_rate
roro["A7"] = "MTC (t·m/cm)"
roro["B7"] = "=Calc!D16"  # D16 = MTC
roro["A8"] = "LCF (m, midship=0)"
roro["B8"] = "=Calc!D17"  # D17 = LCF
roro["A9"] = "TPC (t/cm)"
roro["B9"] = "=Calc!D18"  # D18 = TPC
roro["A10"] = "Pump rate (t/h)"
roro["B10"] = "=Calc!D14"  # D14 = pump_rate
for rr in range(7, 11):
    roro.cell(row=rr, column=1).border = border
    roro.cell(row=rr, column=2).border = border

hdrs2 = [
    "Stage",
    "W_stage_t",
    "x_stage_m (midship=0)",
    "TM (t·m)",
    "Trim_cm",
    "Trim_m",
    "Dfwd_m",
    "Daft_m",
    "Ballast_t (≈Δmean)",
    "Ballast_time_h",
    "FWD_Height_m",
    "AFT_Height_m",
]
row0 = 12
for i, h in enumerate(hdrs2, 1):
    c = roro.cell(row=row0, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i in range(13):
    r = row0 + 1 + i
    roro.cell(row=r, column=1, value=f"Stage {i+1}")
    w_val = 217.0 if i < 2 else None
    x_val = -5.0 if i < 2 else None
    cW = roro.cell(row=r, column=2, value=w_val)
    cW.fill = input_fill
    cX = roro.cell(row=r, column=3, value=x_val)
    cX.fill = input_fill
    roro.cell(row=r, column=4, value=f'=IF(OR(B{r}="",C{r}=""),"",B{r}*(C{r}-$B$8))')
    roro.cell(row=r, column=5, value=f'=IF(OR(D{r}="",$B$7=0),"",D{r}/$B$7)')
    roro.cell(row=r, column=6, value=f'=IF(E{r}="","",E{r}/100)')
    roro.cell(row=r, column=7, value=f'=IF(OR($C$4="",F{r}=""),"",$C$4-F{r}/2)')
    roro.cell(row=r, column=8, value=f'=IF(OR($C$4="",F{r}=""),"",$C$4+F{r}/2)')
    roro.cell(row=r, column=9, value=f'=IF(OR($B$9="",F{r}=""),"",ABS(F{r})*50*$B$9)')
    roro.cell(row=r, column=10, value=f'=IF(OR(I{r}="",$B$10=0),"",I{r}/$B$10)')
    # Height = Depth - Draft + Tide_ref (Calc!D11 = Depth, G/H = Draft, Tide_ref는 평균값 가정)
    # Tide_ref는 Tmean baseline (C4) 또는 평균 조수값 사용 가능
    roro.cell(
        row=r,
        column=11,
        value=f'=IF(OR(G{r}="",Calc!$D$11=""),"",Calc!$D$11 - G{r} + $C$4)',
    )  # FWD_Height_m (Tmean baseline 사용)
    roro.cell(
        row=r,
        column=12,
        value=f'=IF(OR(H{r}="",Calc!$D$11=""),"",Calc!$D$11 - H{r} + $C$4)',
    )  # AFT_Height_m
    for c in range(1, 13):
        roro.cell(row=r, column=c).border = border

for col, w in zip(range(1, 13), [12, 12, 18, 12, 10, 10, 10, 10, 14, 14, 12, 12]):
    roro.column_dimensions[get_column_letter(col)].width = w

# ---------- 5) Formula_Test 시트 추가 ----------
test = wb.create_sheet("Formula_Test")
test["A1"] = "FORMULA VALIDATION TEST CASES"
test["A1"].font = title_font
test.merge_cells("A1:K1")

test["A2"] = (
    "⚠ This sheet validates all formulas with known test cases. All tests must show PASS."
)
test["A2"].font = sec_font
test["A2"].fill = note_fill
test.merge_cells("A2:K2")

# 테스트 헤더
test_hdrs = [
    "Test",
    "Description",
    "KminusZ",
    "Tide",
    "L_ramp",
    "theta",
    "Expected_Dfwd",
    "Calc_Dfwd",
    "Expected_Angle",
    "Calc_Angle",
    "Result",
]
row_t = 4
for i, h in enumerate(test_hdrs, 1):
    c = test.cell(row=row_t, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# 테스트 케이스 A: KminusZ=3.00, Tide=1.50 → Dfwd≈3.239, Angle≈6.0°
row_t += 1
test.cell(row=row_t, column=1, value="A")
test.cell(row=row_t, column=2, value="Boundary test (theta=max)")
test.cell(row=row_t, column=3, value=3.00)
test.cell(row=row_t, column=4, value=1.50)
test.cell(row=row_t, column=5, value=12)
test.cell(row=row_t, column=6, value=6)
test.cell(row=row_t, column=7, value=3.239)
# Calc_Dfwd = KminusZ + Tide - L*TAN(RADIANS(theta))
test.cell(
    row=row_t, column=8, value=f"=C{row_t}+D{row_t}-E{row_t}*TAN(RADIANS(F{row_t}))"
)
test.cell(row=row_t, column=9, value=6.0)
# Calc_Angle = DEGREES(ATAN((KminusZ - Dfwd + Tide)/L))
test.cell(
    row=row_t, column=10, value=f"=DEGREES(ATAN((C{row_t}-H{row_t}+D{row_t})/E{row_t}))"
)
# Result = IF(AND(ABS(H-G)<=0.01, ABS(J-I)<=0.1), "PASS", "FAIL")
test.cell(
    row=row_t,
    column=11,
    value=f'=IF(AND(ABS(H{row_t}-G{row_t})<=0.01,ABS(J{row_t}-I{row_t})<=0.1),"PASS","FAIL")',
)

# 테스트 케이스 B: KminusZ=3.00, Tide=0.50 → Dfwd≈2.239, Angle≈3.0°
row_t += 1
test.cell(row=row_t, column=1, value="B")
test.cell(row=row_t, column=2, value="Normal operation (mid-range)")
test.cell(row=row_t, column=3, value=3.00)
test.cell(row=row_t, column=4, value=0.50)
test.cell(row=row_t, column=5, value=12)
test.cell(row=row_t, column=6, value=6)
test.cell(row=row_t, column=7, value=2.239)
test.cell(
    row=row_t, column=8, value=f"=C{row_t}+D{row_t}-E{row_t}*TAN(RADIANS(F{row_t}))"
)
test.cell(row=row_t, column=9, value=3.0)
test.cell(
    row=row_t, column=10, value=f"=DEGREES(ATAN((C{row_t}-H{row_t}+D{row_t})/E{row_t}))"
)
test.cell(
    row=row_t,
    column=11,
    value=f'=IF(AND(ABS(H{row_t}-G{row_t})<=0.01,ABS(J{row_t}-I{row_t})<=0.1),"PASS","FAIL")',
)

# 테스트 케이스 C: Stage TM 검증
row_t += 2
test.cell(row=row_t, column=1, value="C")
test.cell(row=row_t, column=2, value="Stage TM calculation")
test.merge_cells(f"B{row_t}:F{row_t}")

test_hdrs_c = ["W (t)", "x (m)", "LCF (m)", "Expected_TM", "Calc_TM", "Result"]
row_t += 1
for i, h in enumerate(test_hdrs_c, 1):
    c = test.cell(row=row_t, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

row_t += 1
test.cell(row=row_t, column=1, value=217)
test.cell(row=row_t, column=2, value=-5.0)
test.cell(row=row_t, column=3, value=29.29)  # Verified LCF
test.cell(row=row_t, column=4, value=-7443.93)  # 217 * (-5 - 29.29) = -7443.93
test.cell(row=row_t, column=5, value=f"=A{row_t}*(B{row_t}-C{row_t})")
test.cell(row=row_t, column=6, value=f'=IF(ABS(E{row_t}-D{row_t})<=1,"PASS","FAIL")')

# 조건부 서식: PASS=녹색, FAIL=빨간색
try:
    # 전체 테스트 영역 K5:K6으로 확대 (A/B 케이스)
    pass_rule = FormulaRule(
        formula=['$K$5:$K$6="PASS"'], fill=pass_fill, stopIfTrue=True
    )
    fail_rule = FormulaRule(
        formula=['$K$5:$K$6="FAIL"'], fill=fail_fill, stopIfTrue=True
    )
    test.conditional_formatting.add("K5:K6", pass_rule)
    test.conditional_formatting.add("K5:K6", fail_rule)

    # Test C Result 컬럼 (컬럼 6, F9)
    pass_rule2 = FormulaRule(formula=['$F$9="PASS"'], fill=pass_fill, stopIfTrue=True)
    fail_rule2 = FormulaRule(formula=['$F$9="FAIL"'], fill=fail_fill, stopIfTrue=True)
    test.conditional_formatting.add("F9:F9", pass_rule2)
    test.conditional_formatting.add("F9:F9", fail_rule2)
    print("[INFO] Conditional formatting applied successfully.")
except Exception as e:
    print(f"[WARNING] Conditional formatting failed: {e}")

for col, w in zip(range(1, 12), [8, 25, 10, 10, 10, 10, 14, 12, 14, 12, 10]):
    test.column_dimensions[get_column_letter(col)].width = w

# 모든 테스트 행에 border 적용
for r in range(4, row_t + 1):
    for c in range(1, 12):
        if test.cell(row=r, column=c).value is not None or c <= 11:
            test.cell(row=r, column=c).border = border

# ---------- 6) RoRo_Height_Report 시트 생성 ----------
height_report = wb.create_sheet("RoRo_Height_Report")
height_report["A1"] = "LCT BUSHRA — RORO FWD/AFT HEIGHT REPORT (For Mammoet DWG Update)"
height_report["A1"].font = title_font
height_report.merge_cells("A1:L1")

# PDF 참조 정보
height_report["A2"] = (
    "PDF Reference: RoRo Simulation Stowage Plan 2025-11-03, General Notes 6"
)
height_report["A2"].font = sec_font
height_report["A2"].fill = note_fill
height_report.merge_cells("A2:L2")

# 물리적 상수 섹션
height_report["A4"] = "PHYSICAL CONSTANTS"
height_report["A4"].font = sec_font
height_report["A5"] = "A_wp (Waterplane Area, m²)"
height_report["B5"] = "=64*14.6*0.85"  # 공식화, Cb=0.85
height_report["A6"] = "ρ (Seawater Density, t/m³)"
height_report["B6"] = 1.025
height_report["A7"] = "MTC (Moment to Change Trim, t·m/m)"
height_report["B7"] = 40.96  # m 단위 일치
height_report["A8"] = "Lightship (t)"
height_report["B8"] = 1400  # 초기 Delta=1629t 일치 (Deadweight 229t 가정)
height_report["A9"] = "Deadweight_initial (t)"
height_report["B9"] = 100
height_report["A10"] = "Depth D (m)"
height_report["B10"] = "=Calc!D11"

for r in range(4, 11):
    for c in range(1, 3):
        if height_report.cell(row=r, column=c).value:
            height_report.cell(row=r, column=c).border = border

# Stage별 높이 테이블 헤더
height_hdrs = [
    "Stage",
    "Description",
    "Displacement Δ (t)",
    "Mean Draft Td (m)",
    "Trim t (m, +Aft)",
    "Draft FWD (m)",
    "Draft AFT (m)",
    "Tide (m)",
    "Depth D (m)",
    "FWD Height (m)",
    "AFT Height (m)",
    "Notes (PDF Ref)",
]
row_h = 12
for i, h in enumerate(height_hdrs, 1):
    c = height_report.cell(row=row_h, column=i, value=h)
    c.font = hdr_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Stage 데이터 (계획의 샘플 값 사용)
stage_data = [
    ("Initial", "Pre-Loadout (Even Keel)", 900, 1.17, 0, 1.17, 1.17, 0.5),
    ("Stage 1", "SPMT Entry, 1st Transformer", 1200, 1.56, 0.30, 1.71, 1.41, 0.5),
    ("Stage 2", "Mid-Loadout, 50% Loaded", 1500, 1.95, 0.45, 2.18, 1.72, 0.5),
    ("Final", "Post-Loadout, Full 434t", 1534, 2.00, 0.60, 2.30, 1.70, 0.5),
    ("RoRo Specific", "Jetty-LCT Distance Confirm", 1100, 1.43, 0.20, 1.53, 1.33, 0.5),
]
notes_data = [
    "Stage 1 Prep; Even Keel (#msg-28)",
    "RoRo Plates; SPMT Entry (Page 2)",
    "FWB Tank #2; Timber Packing (Page 3)",
    "Sail-away; Deck 10t/m² OK (#msg-30)",
    "Note 5/6: Captain advise heights (Page 3)",
]

row_h += 1
for idx, (stage, desc, delta, td, trim, dfwd, daft, tide) in enumerate(stage_data):
    height_report.cell(row=row_h, column=1, value=stage)
    height_report.cell(row=row_h, column=2, value=desc)
    height_report.cell(row=row_h, column=3, value=delta)
    height_report.cell(row=row_h, column=4, value=td)
    height_report.cell(row=row_h, column=5, value=trim)
    height_report.cell(row=row_h, column=6, value=dfwd)
    height_report.cell(row=row_h, column=7, value=daft)
    height_report.cell(row=row_h, column=8, value=tide)
    # Depth = Calc!D11
    height_report.cell(row=row_h, column=9, value=f"=Calc!D11")
    # FWD Height = Depth - Draft FWD + Tide
    height_report.cell(row=row_h, column=10, value=f"=I{row_h} - F{row_h} + H{row_h}")
    # AFT Height = Depth - Draft AFT + Tide
    height_report.cell(row=row_h, column=11, value=f"=I{row_h} - G{row_h} + H{row_h}")
    height_report.cell(row=row_h, column=12, value=notes_data[idx])
    for c in range(1, 13):
        height_report.cell(row=row_h, column=c).border = border
    row_h += 1

# 컬럼 너비 설정
for col, w in zip(range(1, 13), [12, 35, 14, 12, 12, 12, 12, 10, 12, 14, 14, 50]):
    height_report.column_dimensions[get_column_letter(col)].width = w

# ---------- 7) README ----------
readme = wb.create_sheet("README")
readme["A1"] = "LCT BUSHRA FWD/AFT Draft Calculator - QUICK GUIDE"
readme["A1"].font = title_font

readme_content = [
    "Usage: 1) Update Calc values; 2) Paste tide data; 3) Check Hourly Status; 4) Stage planning",
    "",
    "⚠️ CRITICAL CELL MAPPING (must memorize):",
    "Calc!D8 = L_ramp_m (m) — 기본 12.0",
    "Calc!D9 = theta_max_deg (deg) — 기본 6.0",
    "Calc!D10 = KminusZ_m (m) — 현장 실측값(필수) ⚠️",
    "Calc!D11 = D_vessel_m (m) — 3.65m (Moulded Depth, verified: RoRo Simulation & Stability Booklet, 5/5 documents match)",
    "Calc!D12 = min_fwd_draft_m (m) — 운용 하한",
    "Calc!D13 = max_fwd_draft_m (m) — 운용 상한",
    "Calc!D14 = pump_rate_tph (t/h) — 기본 10",
    "Calc!D16 = MTC_t_m_per_m — 4072 (Verified, m 단위)",
    "Calc!D17 = LCF_m_from_midship — 29.29 (Verified from Stability Book, Draft ~2.50m) midship 기준",
    "Calc!D18 = TPC_t_per_cm — optional",
    "",
    "Formulas:",
    "Dfwd_req = Calc!D10 + Tide_m - Calc!D8 * TAN(RADIANS(Calc!D9))",
    "RampAngle = DEGREES(ATAN((Calc!D10 - Dfwd_req + Tide_m) / Calc!D8))",
    "FWD_Height = Calc!D11 - Draft_FWD + Tide",
    "AFT_Height = Calc!D11 - Draft_AFT + Tide",
    "",
    "Test Cases (Formula_Test sheet):",
    "Test A → Dfwd ≈ 3.239 m, Angle ≈ 6.0°",
    "Test B → Dfwd ≈ 2.239 m, Angle ≈ ~3.0°",
    "Test C → TM = -7443.93 t·m (W=217, x=-5, LCF=29.29, Verified)",
    "",
    "RoRo_Height_Report: Stage별 높이 테이블 (Mammoet DWG 업데이트용)",
]

for i, line in enumerate(readme_content, 3):
    readme.cell(row=i, column=1, value=line)

readme.column_dimensions["A"].width = 120

# Save workbook
try:
    wb.save(out_path)
    wb.close()
    print(f"[SUCCESS] Excel saved to {out_path}")
except Exception as e:
    print(f"[ERROR] Save failed: {e}")
    raise

print("=" * 80)
print("\n[SUCCESS] Generation complete!")
print("\nNext steps:")
print("1. Open Excel file")
print("2. Enter K-Z measured on site into Calc!D10 (REQUIRED)")
print("3. December_Tide_2025 sheet: Paste 744 tide values into column B")
print("4. Formula_Test sheet: Verify all tests show PASS")
print("5. Hourly_FWD_AFT_Heights sheet: Check Status=OK time periods")
print("=" * 80)
