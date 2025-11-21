# -*- coding: utf-8 -*-
"""Excel/CSV 파일을 JSON으로 변환하는 스크립트"""
import json
import os
from pathlib import Path

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("pandas not available, trying openpyxl...")

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def excel_to_json_with_openpyxl(excel_file, json_file):
    """openpyxl을 사용하여 Excel 파일을 JSON으로 변환"""
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        # 첫 번째 행을 헤더로 사용
        headers = [cell.value for cell in ws[1]]

        # 데이터 읽기
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # 빈 행 제외
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value
                if row_dict:
                    data.append(row_dict)

        # JSON으로 저장
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

        print(f"Successfully converted {excel_file} to {json_file}")
        print(f"  Rows: {len(data)}")
        if data:
            print(f"  Columns: {list(data[0].keys())}")
        return True
    except Exception as e:
        print(f"Error with openpyxl: {e}")
        return False

def csv_to_json_with_pandas(csv_file, json_file):
    """pandas를 사용하여 CSV를 JSON으로 변환"""
    encodings = ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr', 'latin-1']

    for encoding in encodings:
        try:
            df = pd.read_csv(csv_file, encoding=encoding)
            # NaN 값을 None으로 변환
            df = df.where(pd.notnull(df), None)
            data = df.to_dict('records')

            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)

            print(f"Successfully converted {csv_file} to {json_file} (encoding: {encoding})")
            print(f"  Rows: {len(data)}")
            if data:
                print(f"  Columns: {list(data[0].keys())}")
            return True
        except Exception as e:
            continue

    return False

if __name__ == "__main__":
    base_dir = Path(__file__).parent
    data_dir = base_dir / "data"

    # 변환할 파일들
    files = [
        ("Tank 좌표.csv", "tank_coordinates.json"),
        ("Tank Capacity.csv", "tank_capacity.json")
    ]

    for csv_name, json_name in files:
        csv_path = data_dir / csv_name
        json_path = data_dir / json_name

        if not csv_path.exists():
            print(f"File not found: {csv_path}")
            continue

        # 먼저 Excel로 읽기 시도
        if HAS_OPENPYXL:
            print(f"\nTrying to read {csv_name} as Excel file...")
            if excel_to_json_with_openpyxl(csv_path, json_path):
                continue

        # CSV로 읽기 시도
        if HAS_PANDAS:
            print(f"\nTrying to read {csv_name} as CSV file...")
            if csv_to_json_with_pandas(csv_path, json_path):
                continue

        print(f"Failed to convert {csv_name}")

