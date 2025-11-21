# -*- coding: utf-8 -*-
# generate_mammoet_package.py
# Generates submission package for Mammoet (RORO operations contractor)

import os
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Protection
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.patches as mpatches

# Configuration
EXCEL_INPUT = "../output/LCT_BUSHRA_GateAB_v4_HYBRID.xlsx"  # Updated to v4 HYBRID
OUTPUT_DIR = "MAMMOET_PACKAGE"
PDF_REPORT = "LCT_BUSHRA_FWD_AFT_Report_for_Mammoet.pdf"
LOCKED_EXCEL = "LCT_BUSHRA_FWD_AFT_Calculator_COMPLETE.xlsx"

# Submission metadata
SUBMISSION_INFO = {
    "vessel_name": "LCT BUSHRA",
    "project": "HVDC Transformer Transportation",
    "route": "Mina Zayed Port → DAS Island",
    "operator": "Samsung C&T / Mammoet",
    "date_prepared": datetime.now().strftime("%Y-%m-%d"),
    "prepared_by": "Samsung Logistics Team",
    "recipient": "Mammoet / Aries Marine",
}

def create_output_directory():
    """Create output directory structure"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        os.makedirs(os.path.join(OUTPUT_DIR, "01_PDF_Report"))
        os.makedirs(os.path.join(OUTPUT_DIR, "02_Working_Excel"))
        os.makedirs(os.path.join(OUTPUT_DIR, "03_Supporting_Evidence"))
        print(f"[OK] Created directory: {OUTPUT_DIR}")
    else:
        print(f"[OK] Directory exists: {OUTPUT_DIR}")

def safe_float(value, default=0.0):
    """Safely convert value to float"""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default

def load_excel_data():
    """Load data from Excel workbook"""
    try:
        wb = load_workbook(EXCEL_INPUT, data_only=True)
        
        # Read constants
        calc_ws = wb["Calc"]
        constants = {
            "KminusZ_m": safe_float(calc_ws["D10"].value),  # v4 HYBRID: D10
            "L_ramp_m": safe_float(calc_ws["D8"].value),    # v4 HYBRID: D8
            "theta_max_deg": safe_float(calc_ws["D9"].value),  # v4 HYBRID: D9
            "D_vessel_m": safe_float(calc_ws["D11"].value),  # v4 HYBRID: D11
            "Lpp_m": safe_float(calc_ws["D16"].value) if calc_ws["D16"].value else None,
            "min_fwd_draft": safe_float(calc_ws["D13"].value),  # v4 HYBRID: D13
            "max_fwd_draft": safe_float(calc_ws["D14"].value),  # v4 HYBRID: D14
        }
        
        # Read hourly data (all 744 hours)
        hourly_ws = wb["Hourly_FWD_AFT_Heights"]
        hourly_data = []
        for row in range(2, 746):
            try:
                dt = hourly_ws.cell(row=row, column=1).value
                tide = hourly_ws.cell(row=row, column=2).value
                dfwd_req = hourly_ws.cell(row=row, column=3).value
                trim = hourly_ws.cell(row=row, column=4).value
                dfwd_adj = hourly_ws.cell(row=row, column=5).value
                daft_adj = hourly_ws.cell(row=row, column=6).value
                angle = hourly_ws.cell(row=row, column=7).value
                status = hourly_ws.cell(row=row, column=8).value
                
                if dt and tide is not None:
                    # Calculate trim from drafts if available
                    trim_val = None
                    if dfwd_adj and daft_adj:
                        trim_val = safe_float(daft_adj) - safe_float(dfwd_adj)
                    elif dfwd_req and daft_req:
                        trim_val = safe_float(daft_req) - safe_float(dfwd_req)
                    
                    hourly_data.append({
                        "datetime": dt,
                        "tide": safe_float(tide),
                        "dfwd_req": safe_float(dfwd_req) if dfwd_req else None,
                        "daft_req": safe_float(daft_req) if daft_req else None,
                        "trim": trim_val,
                        "dfwd_adj": safe_float(dfwd_adj) if dfwd_adj else None,
                        "daft_adj": safe_float(daft_adj) if daft_adj else None,
                        "angle": safe_float(angle) if angle else None,
                        "status": str(status) if status else "",
                    })
            except:
                continue
        
        wb.close()
        print(f"[OK] Loaded {len(hourly_data)} hourly records")
        return constants, hourly_data
    
    except FileNotFoundError:
        print(f"[ERROR] {EXCEL_INPUT} not found!")
        print(f"  Run build_bushra_gateab_v4_hybrid.py first to generate the Excel file.")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] loading Excel: {e}")
        sys.exit(1)

def find_operation_windows(hourly_data, min_duration_hours=2):
    """Find continuous OK operation windows"""
    windows = []
    current_window = []
    
    for record in hourly_data:
        if record['status'] == 'OK':
            current_window.append(record)
        else:
            if len(current_window) >= min_duration_hours:
                windows.append(current_window)
            current_window = []
    
    # Check last window
    if len(current_window) >= min_duration_hours:
        windows.append(current_window)
    
    return windows

def generate_mammoet_pdf(constants, hourly_data, operation_windows):
    """Generate concise PDF report for Mammoet (1-2 pages)"""
    pdf_path = os.path.join(OUTPUT_DIR, "01_PDF_Report", PDF_REPORT)
    
    try:
        with PdfPages(pdf_path) as pdf:
            # Page 1: Summary + Recommended Windows
            fig, ax = plt.subplots(figsize=(11.69, 8.27))  # A4 landscape
            ax.axis('off')
            
            # Title
            y = 0.98
            ax.text(0.5, y, "LCT BUSHRA — FWD/AFT DRAFT REPORT", 
                   ha='center', va='top', fontsize=16, weight='bold')
            y -= 0.03
            ax.text(0.5, y, "For Mammoet RORO Operations", 
                   ha='center', va='top', fontsize=12, style='italic')
            y -= 0.05
            
            # Metadata
            meta_text = f"""
Vessel: {SUBMISSION_INFO['vessel_name']}
Project: {SUBMISSION_INFO['project']}
Route: {SUBMISSION_INFO['route']}
Prepared: {SUBMISSION_INFO['date_prepared']} | By: {SUBMISSION_INFO['prepared_by']}
Recipient: {SUBMISSION_INFO['recipient']}
            """
            ax.text(0.02, y, meta_text.strip(), ha='left', va='top', 
                   fontsize=9, family='monospace')
            y -= 0.12
            
            # Critical parameters box (highlighted)
            const_text = f"""
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃  CRITICAL PARAMETERS (Site Measured)                     ┃
┣━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┫
┃  K–Z Distance (measured):  {constants['KminusZ_m']:.2f} m                     ┃
┃  Linkspan Length:          {constants['L_ramp_m']:.1f} m                        ┃
┃  Maximum Ramp Angle:       {constants['theta_max_deg']:.1f}° (Operational Limit)        ┃
┃  Vessel Molded Depth:      {constants['D_vessel_m']:.2f} m                      ┃
┃  Draft Range (Operational): {constants['min_fwd_draft']:.1f} - {constants['max_fwd_draft']:.1f} m                ┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
            """
            ax.text(0.5, y, const_text.strip(), ha='center', va='top',
                   fontsize=9, family='monospace', weight='bold',
                   bbox=dict(boxstyle='round', facecolor='#FFE6CC', edgecolor='#FF6600', linewidth=2))
            y -= 0.16
            
            # Recommended operation windows
            ax.text(0.02, y, "RECOMMENDED OPERATION WINDOWS (≥2h continuous OK status)",
                   ha='left', va='top', fontsize=11, weight='bold')
            y -= 0.03
            
            # Show top 3 windows
            if operation_windows:
                window_text = ""
                for i, window in enumerate(operation_windows[:3], 1):
                    start = window[0]['datetime']
                    end = window[-1]['datetime']
                    duration = len(window)
                    avg_angle = sum(safe_float(r['angle']) for r in window if r['angle']) / len(window)
                    avg_tide = sum(safe_float(r['tide']) for r in window) / len(window)
                    
                    start_str = start if isinstance(start, str) else start.strftime("%Y-%m-%d %H:%M")
                    end_str = end if isinstance(end, str) else end.strftime("%Y-%m-%d %H:%M")
                    
                    window_text += f"""
Window {i}: {start_str} → {end_str}
  Duration: {duration} hours | Avg Tide: {avg_tide:.2f}m | Avg Ramp Angle: {avg_angle:.1f}°
                    """
                
                ax.text(0.02, y, window_text.strip(), ha='left', va='top',
                       fontsize=9, family='monospace',
                       bbox=dict(boxstyle='round', facecolor='#E8F5E9'))
                y -= 0.10
            else:
                ax.text(0.02, y, "[WARNING] No continuous OK windows found. Review hourly table.",
                       ha='left', va='top', fontsize=9, color='red')
                y -= 0.05
            
            # Detailed hourly table header
            ax.text(0.02, y, "DETAILED HOURLY SCHEDULE (First Window Expanded)",
                   ha='left', va='top', fontsize=10, weight='bold')
            y -= 0.03
            
            # Table for first window (or first 24h if no windows)
            table_data = [["Date", "Time\n(GST)", "Tide\n(m CD)", "Dfwd_req\n(m)", 
                          "Daft_req\n(m)", "Trim\n(m)", "Ramp∠\n(deg)", "Status", "Remark"]]
            
            records_to_show = operation_windows[0][:24] if operation_windows else hourly_data[:24]
            
            for record in records_to_show:
                dt = record['datetime']
                if isinstance(dt, str):
                    date_str = dt.split()[0]
                    time_str = dt.split()[1] if len(dt.split()) > 1 else ""
                else:
                    date_str = dt.strftime("%m-%d")
                    time_str = dt.strftime("%H:%M")
                
                trim_str = f"{record['trim']:.2f}" if record['trim'] != 0 else "Even-keel"
                remark = "OK for RORO" if record['status'] == 'OK' else "Verify conditions"
                
                table_data.append([
                    date_str,
                    time_str,
                    f"{safe_float(record['tide']):.2f}",
                    f"{safe_float(record['dfwd_req']):.2f}" if record['dfwd_req'] else "-",
                    f"{safe_float(record['daft_adj']):.2f}" if record['daft_adj'] else "-",
                    trim_str,
                    f"{safe_float(record['angle']):.1f}" if record['angle'] else "-",
                    record['status'] if record['status'] else "-",
                    remark
                ])
            
            # Create table
            table = ax.table(cellText=table_data,
                           colWidths=[0.08, 0.08, 0.09, 0.10, 0.10, 0.11, 0.09, 0.09, 0.16],
                           loc='center',
                           bbox=[0.02, 0.08, 0.96, y-0.11])
            
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            table.scale(1, 1.2)
            
            # Style header
            for i in range(9):
                cell = table[(0, i)]
                cell.set_facecolor('#1F4E78')
                cell.set_text_props(weight='bold', color='white')
            
            # Color code status
            for i in range(1, len(table_data)):
                status_cell = table[(i, 7)]
                if table_data[i][7] == "OK":
                    status_cell.set_facecolor('#90EE90')
                elif table_data[i][7] == "CHECK":
                    status_cell.set_facecolor('#FFD700')
            
            # Footer notes
            ax.text(0.02, 0.04,
                   "Notes: Dfwd_req = Even-keel draft requirement. Daft_req accounts for trim if applied.",
                   ha='left', va='top', fontsize=7, style='italic')
            ax.text(0.02, 0.02,
                   "[WARNING] This calculation is based on site-measured K-Z and official tide predictions. Verify actual conditions before operations.",
                   ha='left', va='top', fontsize=7, style='italic', color='red')
            ax.text(0.02, 0.005,
                   f"Data Source: Tide from [AD Ports/ADNOC]. Chart Datum reference. K-Z measured on [Date to be filled].",
                   ha='left', va='bottom', fontsize=7, style='italic')
            
            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)
            
            # Page 2: Additional windows (if more than 24h in first window)
            if operation_windows and len(operation_windows[0]) > 24:
                fig2, ax2 = plt.subplots(figsize=(11.69, 8.27))
                ax2.axis('off')
                
                y2 = 0.98
                ax2.text(0.5, y2, "LCT BUSHRA — FWD/AFT DRAFT REPORT (Continued)", 
                        ha='center', va='top', fontsize=16, weight='bold')
                y2 -= 0.06
                
                ax2.text(0.02, y2, "DETAILED HOURLY SCHEDULE (Continuation)",
                        ha='left', va='top', fontsize=10, weight='bold')
                y2 -= 0.03
                
                # Table for hours 25-48 of first window
                table_data2 = [["Date", "Time\n(GST)", "Tide\n(m CD)", "Dfwd_req\n(m)", 
                               "Daft_req\n(m)", "Trim\n(m)", "Ramp∠\n(deg)", "Status", "Remark"]]
                
                for record in operation_windows[0][24:48]:
                    dt = record['datetime']
                    if isinstance(dt, str):
                        date_str = dt.split()[0]
                        time_str = dt.split()[1] if len(dt.split()) > 1 else ""
                    else:
                        date_str = dt.strftime("%m-%d")
                        time_str = dt.strftime("%H:%M")
                    
                    trim_str = f"{safe_float(record['trim']):.2f}" if record['trim'] != 0 else "Even-keel"
                    remark = "OK for RORO" if record['status'] == 'OK' else "Verify conditions"
                    
                    table_data2.append([
                        date_str,
                        time_str,
                        f"{safe_float(record['tide']):.2f}",
                        f"{safe_float(record['dfwd_req']):.2f}" if record['dfwd_req'] else "-",
                        f"{safe_float(record['daft_adj']):.2f}" if record['daft_adj'] else "-",
                        trim_str,
                        f"{safe_float(record['angle']):.1f}" if record['angle'] else "-",
                        record['status'] if record['status'] else "-",
                        remark
                    ])
                
                table2 = ax2.table(cellText=table_data2,
                                 colWidths=[0.08, 0.08, 0.09, 0.10, 0.10, 0.11, 0.09, 0.09, 0.16],
                                 loc='center',
                                 bbox=[0.02, 0.15, 0.96, y2-0.18])
                
                table2.auto_set_font_size(False)
                table2.set_fontsize(7)
                table2.scale(1, 1.2)
                
                # Style
                for i in range(9):
                    cell = table2[(0, i)]
                    cell.set_facecolor('#1F4E78')
                    cell.set_text_props(weight='bold', color='white')
                
                for i in range(1, len(table_data2)):
                    status_cell = table2[(i, 7)]
                    if table_data2[i][7] == "OK":
                        status_cell.set_facecolor('#90EE90')
                    elif table_data2[i][7] == "CHECK":
                        status_cell.set_facecolor('#FFD700')
                
                # Contact info
                ax2.text(0.02, 0.10, "CONTACT INFORMATION", fontsize=10, weight='bold')
                contact_text = f"""
Prepared by: {SUBMISSION_INFO['prepared_by']}
For questions: [Contact Email/Phone]

Mammoet Operations Coordinator: [To be filled]
Aries Marine Superintendent: [To be filled]
Samsung Project Manager: [To be filled]
                """
                ax2.text(0.02, 0.09, contact_text.strip(), ha='left', va='top', fontsize=8)
                
                pdf.savefig(fig2, bbox_inches='tight')
                plt.close(fig2)
        
        print(f"[OK] PDF report generated: {pdf_path}")
        return pdf_path
    
    except Exception as e:
        print(f"[ERROR] generating PDF: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_locked_excel():
    """Create locked version with formula protection"""
    try:
        wb = load_workbook(EXCEL_INPUT)
        
        # Protect sheets
        for sheet_name in ["Calc", "Hourly_FWD_AFT_Heights", "RORO_Stage_Scenarios"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Unlock yellow input cells only
                for row in ws.iter_rows():
                    for cell in row:
                        try:
                            if cell.fill.fgColor and cell.fill.fgColor.rgb == "FFF2CC":
                                cell.protection = Protection(locked=False)
                            else:
                                cell.protection = Protection(locked=True)
                        except:
                            cell.protection = Protection(locked=True)
                
                # Protect sheet
                ws.protection.sheet = True
                ws.protection.password = "MAMMOET2025"
                ws.protection.formatCells = False
                ws.protection.selectLockedCells = True
                ws.protection.selectUnlockedCells = True
        
        locked_path = os.path.join(OUTPUT_DIR, "02_Working_Excel", LOCKED_EXCEL)
        wb.save(locked_path)
        wb.close()
        
        print(f"[OK] Locked Excel created: {locked_path}")
        print(f"  Password: MAMMOET2025")
        return locked_path
    
    except Exception as e:
        print(f"[ERROR] creating locked Excel: {e}")
        return None

def create_supporting_evidence():
    """Create supporting evidence templates"""
    support_dir = os.path.join(OUTPUT_DIR, "03_Supporting_Evidence")
    
    # 1. K-Z Measurement Photo Template
    kz_photo_guide = """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
K-Z DISTANCE MEASUREMENT EVIDENCE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

For: Mammoet RORO Operations
Vessel: LCT BUSHRA
Project: HVDC Transformer Transportation

MEASUREMENT DETAILS:
+------------------------------------------------------+
| Measurement Date: _______________                    |
| Time (GST):       _______________                    |
| Location:         Mina Zayed Port, Berth [__]       |
| Tide at Time:     _______ m (Chart Datum)           |
| Weather:          _______________                    |
+------------------------------------------------------+
| Measured K-Z:     _______ m                          |
| Verification:     _______ m (repeat measurement)    |
| Difference:       _______ m (should be < 0.05m)     |
+------------------------------------------------------+

MEASUREMENT METHOD:
1. Vessel positioned parallel to jetty
2. Measured from jetty deck to linkspan contact point
3. Corrections applied for:
   - Jetty elevation above CD: _______ m
   - Tide height at measurement: _______ m
   - Vessel trim (if applicable): _______ m

PERSONNEL PRESENT:
- Surveyor: _______________________
- LCT Bushra Master: _______________________
- Mammoet Representative: _______________________
- Samsung Representative: _______________________

PHOTO ATTACHED:
□ Overall measurement setup (wide view)
□ Close-up of contact point with measuring tape/laser
□ Digital readout of measurement equipment

VERIFICATION STATEMENT:
The K-Z distance of _______ m has been measured on site and 
verified. This value has been entered into Calc!D6 of the 
working Excel file.

Measured by: _______________________
Verified by: _______________________
Date: _______________________

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
NOTE: Attach actual measurement photo (min 1, recommended 3) 
      when submitting this package to Mammoet.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    with open(os.path.join(support_dir, "KZ_Measurement_Evidence.txt"), 'w', encoding='utf-8') as f:
        f.write(kz_photo_guide)
    
    # 2. Tide Data Source
    tide_source = """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TIDE DATA SOURCE DOCUMENTATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

For: Mammoet RORO Operations
Project: HVDC Transformer Transportation
Period: December 2025

OFFICIAL SOURCE:
□ Abu Dhabi Ports Authority (AD Ports)
   Website: https://www.adports.ae
   
□ ADNOC Marine & Offshore Services
   Contact: [_______________]
   
□ Other Official Source: _______________

SOURCE DETAILS:
+------------------------------------------------------+
| Document/Table Name: _______________                 |
| Publication Date:    _______________                 |
| Reference Number:    _______________                 |
| Datum:               Chart Datum (CD)                |
| Location:            Mina Zayed Port                 |
| Validity:            December 2025                   |
+------------------------------------------------------+

DATA CHARACTERISTICS:
- Temporal Resolution: Hourly predictions
- Type: Predicted (astronomical tide)
- Format: Excel spreadsheet / PDF table
- Units: Meters above Chart Datum

VERIFICATION:
Data reviewed and approved by:
Name: _______________________
Position: _______________________
Company: Samsung C&T / Mammoet
Signature: _______________________
Date: _______________________

ATTACHMENT:
□ Screenshot or PDF of original tide table
□ Excel file with tide data (if available)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
NOTE: Attach screenshot or copy of the official tide table 
      showing the data source clearly.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    with open(os.path.join(support_dir, "Tide_Data_Source.txt"), 'w', encoding='utf-8') as f:
        f.write(tide_source)
    
    print(f"[OK] Supporting evidence templates created")
    return support_dir

def create_readme():
    """Create package README for Mammoet"""
    readme = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LCT BUSHRA - MAMMOET SUBMISSION PACKAGE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Project: {SUBMISSION_INFO['project']}
Vessel: {SUBMISSION_INFO['vessel_name']}
Recipient: {SUBMISSION_INFO['recipient']}

PACKAGE CONTENTS (3 items):

01_PDF_Report/
  └── LCT_BUSHRA_FWD_AFT_Report_for_Mammoet.pdf
      Primary technical document for RORO operations planning
      Contains:
      [OK] Critical parameters (K-Z, linkspan, max angle)
      [OK] Recommended operation windows (>=2h continuous)
      [OK] Detailed hourly schedule with status indicators
      [OK] Tide, draft, and ramp angle data

02_Working_Excel/
  └── LCT_BUSHRA_FWD_AFT_Calculator_COMPLETE.xlsx
      Reference calculation workbook (formula protected)
      Features:
      [OK] Sheet protection password: MAMMOET2025
      [OK] Yellow cells editable (K-Z, tide data, trim inputs)
      [OK] Real-time recalculation enabled
      [OK] All 744 hours of December 2025 data

03_Supporting_Evidence/
  ├── KZ_Measurement_Evidence.txt
  │   Template for K-Z measurement documentation
  │   **ACTION: Attach actual measurement photo(s)**
  │
  └── Tide_Data_Source.txt
      Template for tide data source declaration
      **ACTION: Attach official tide table screenshot/PDF**

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SUBMISSION WORKFLOW
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STEP 1: PRE-SUBMISSION CHECKLIST
□ K-Z measured on site and entered in Excel (Calc!D6)
□ Official tide data pasted in Excel (December_Tide_2025 sheet)
□ K-Z measurement photo taken (min 1, recommended 3)
□ Tide table source documented (AD Ports or ADNOC)
□ Package regenerated after data updates

STEP 2: EMAIL TO MAMMOET
To: [Mammoet Operations Coordinator]
Cc: [Aries Marine], [Samsung Project Team]
Subject: LCT BUSHRA FWD/AFT Draft Calculation - HVDC Project

Attachments:
1. LCT_BUSHRA_FWD_AFT_Report_for_Mammoet.pdf
2. LCT_BUSHRA_FWD_AFT_Calculator_COMPLETE.xlsx (optional)
3. KZ_measurement_photo.jpg (or .pdf)
4. Tide_table_source.pdf (screenshot or copy)

Email Body Template:
───────────────────────────────────────────────────────
Dear Mammoet Team,

Please find attached the FWD/AFT draft calculations for 
LCT BUSHRA RORO operations at Mina Zayed Port.

Key Information:
• K-Z Distance (site measured): [X.XX] m
• Recommended Operation Window: [Date/Time] ([X] hours)
• Maximum Ramp Angle: [X.X]° (within 6° limit)
• Tide Data Source: [AD Ports / ADNOC]

The PDF report highlights continuous operation windows 
with favorable conditions. The working Excel file is 
provided for reference and can be used for real-time 
adjustments during operations.

Supporting evidence includes K-Z measurement photo and 
official tide data source.

Please review and let us know if any clarifications or 
additional information is required.

Best regards,
[Your Name]
Samsung C&T Logistics Team
───────────────────────────────────────────────────────

STEP 3: FOLLOW-UP
□ Confirm receipt with Mammoet
□ Address any questions or clarifications
□ Coordinate final operation scheduling
□ Share with Aries Marine for MSRA/Intact Stability input

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CRITICAL REMINDERS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠ K-Z VALUE: Default 3.0m is PLACEHOLDER ONLY
   → Must measure on site before operations
   → Update Calc!D6 in Excel immediately after measurement
   → Regenerate PDF report after update

⚠ TIDE DATA: Must be from official source
   → AD Ports or ADNOC only
   → Chart Datum reference required
   → Paste into December_Tide_2025!B2:B745 in Excel

⚠ OPERATION WINDOWS: "OK" status means:
   → Draft within operational limits (1.5 - 3.5m)
   → Ramp angle ≤ 6°
   → Tide data available for that hour
   → Still verify actual conditions on day of operation

⚠ MAMMOET USE: This package provides data for:
   → RORO operation scheduling
   → Linkspan positioning calculations
   → Intact Stability assessment (Aries Marine)
   → MSRA (Marine Spread Risk Assessment)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TECHNICAL SUPPORT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

For questions about this package:

Samsung C&T Logistics:
• Project Manager: [Name, Email, Phone]
• Marine Coordinator: [Name, Email, Phone]

LCT BUSHRA:
• Master: [Name, Phone, VHF Channel]
• Chief Officer: [Name, Phone]

Mammoet:
• Operations Coordinator: [Name, Email, Phone]
• Project Engineer: [Name, Email, Phone]

Aries Marine:
• Marine Superintendent: [Name, Email, Phone]

MACHO-GPT Technical Support:
• [Support contact details]

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DOCUMENT VERSION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Excel Calculator: Patch 3 (2025-11-06)
PDF Generator: Mammoet Version 1.0
Package Structure: Mammoet Submission Standard

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONFIDENTIAL: This package contains operational data for 
HVDC project. Distribute only to authorized project 
personnel (Samsung, Mammoet, Aries Marine).
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    readme_path = os.path.join(OUTPUT_DIR, "README_MAMMOET.txt")
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(readme)
    
    print(f"[OK] README created: {readme_path}")
    return readme_path

def main():
    """Main execution"""
    print("\n" + "="*70)
    print("LCT BUSHRA - MAMMOET SUBMISSION PACKAGE GENERATOR")
    print("="*70)
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Vessel: {SUBMISSION_INFO['vessel_name']}")
    print(f"Recipient: {SUBMISSION_INFO['recipient']}")
    print("="*70 + "\n")
    
    # Step 1: Create directory
    print("[1/6] Creating output directory...")
    create_output_directory()
    
    # Step 2: Load data
    print("\n[2/6] Loading Excel data...")
    constants, hourly_data = load_excel_data()
    
    # Step 3: Find operation windows
    print("\n[3/6] Analyzing operation windows...")
    operation_windows = find_operation_windows(hourly_data, min_duration_hours=2)
    print(f"  → Found {len(operation_windows)} continuous operation windows (≥2h)")
    if operation_windows:
        for i, window in enumerate(operation_windows[:3], 1):
            start = window[0]['datetime']
            duration = len(window)
            start_str = start if isinstance(start, str) else start.strftime("%m-%d %H:%M")
            print(f"     Window {i}: {start_str}, {duration}h duration")
    
    # Step 4: Generate PDF
    print("\n[4/6] Generating PDF report for Mammoet...")
    pdf_path = generate_mammoet_pdf(constants, hourly_data, operation_windows)
    
    # Step 5: Create locked Excel
    print("\n[5/6] Creating protected Excel workbook...")
    excel_path = create_locked_excel()
    
    # Step 6: Create supporting evidence
    print("\n[6/6] Creating supporting evidence templates...")
    support_dir = create_supporting_evidence()
    
    # Create README
    print("\n[7/7] Creating README...")
    readme_path = create_readme()
    
    # Summary
    print("\n" + "="*70)
    print("PACKAGE GENERATION COMPLETE")
    print("="*70)
    print(f"\nOutput directory: {OUTPUT_DIR}/")
    print("\nGenerated files:")
    print(f"  [OK] PDF Report: {PDF_REPORT}")
    print(f"  [OK] Working Excel: {LOCKED_EXCEL} (Password: MAMMOET2025)")
    print(f"  [OK] Supporting evidence: 2 templates")
    print(f"  [OK] README: README_MAMMOET.txt")
    
    print("\n" + "-"*70)
    print("NEXT STEPS - BEFORE SENDING TO MAMMOET:")
    print("-"*70)
    print("1. [WARNING] Measure K-Z distance on site")
    print("2. Update Calc!D6 in original Excel with actual K-Z")
    print("3. Paste official tide data (AD Ports/ADNOC)")
    print("4. Re-run this script to regenerate with updated data")
    print("5. Take K-Z measurement photo (attach to email)")
    print("6. Get tide table screenshot/PDF (attach to email)")
    print("7. Email package to Mammoet (see README for template)")
    print("-"*70)
    
    print(f"\n[EMAIL] Package Contents:")
    print("  - LCT_BUSHRA_FWD_AFT_Report_for_Mammoet.pdf (Primary)")
    print("  - LCT_BUSHRA_FWD_AFT_Calculator_COMPLETE.xlsx (Optional)")
    print("  - KZ_measurement_photo.jpg")
    print("  - Tide_table_source.pdf\n")

if __name__ == "__main__":
    main()
