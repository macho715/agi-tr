# -*- coding: utf-8 -*-
# LCT BUSHRA — 통합 운영 스크립트
# 패치, 검증, 분석 기능 통합
# Usage: python bushra_operations.py --patch|--validate|--analyze [options]

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import math
import pandas as pd
import argparse
import sys


class BushraOperations:
    """LCT BUSHRA 통합 운영 클래스"""

    def __init__(self):
        self.hybrid_file = "../output/LCT_BUSHRA_GateAB_v4_HYBRID.xlsx"
        self.final_file = "../output/LCT_BUSHRA_GateAB_v4_HYBRID_with_Dropdown.xlsx"
        self.v3_file = "../backup/Bushra_GateAB_Updated_v3.xlsx"

    # ========== PATCH FUNCTIONS ==========

    def patch_stage_heights(self):
        """Stage_Heights 시트 추가"""
        print("=" * 80)
        print("Stage_Heights 시트 패치 - v4 HYBRID")
        print("=" * 80)

        wb = load_workbook(self.hybrid_file)
        print("[OK] v4 HYBRID file loaded")

        if "Stage_Heights" in wb.sheetnames:
            del wb["Stage_Heights"]
            print("  [INFO] Existing Stage_Heights removed")

        if "Controls" not in wb.sheetnames:
            controls = wb.create_sheet("Controls")
            controls["A1"] = "Stage"
            controls["B1"] = "Reference Time (GST)"
            controls["C1"] = "Helper"

            stages = [
                (1, "", "Before Load-out (Empty Condition)"),
                (2, "", "SPMT 1st Entry on Ramp"),
                (3, "", "50% on Ramp"),
                (4, "", "Full on Ramp (Break-even)"),
                (5, "", "Deck Full Load (217t × 2 on Deck)"),
            ]

            for i, (stage_num, ref_time, helper) in enumerate(stages, 2):
                controls.cell(row=i, column=1, value=stage_num)
                controls.cell(row=i, column=2, value=ref_time)
                controls.cell(row=i, column=3, value=helper)

            print("  [OK] Controls sheet created")

        ws = wb.create_sheet("Stage_Heights")
        print("  [OK] Stage_Heights sheet created")

        hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        thin = Side(border_style="thin", color="C0C0C0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [
            "Stage",
            "Stage Name",
            "Reference Time (GST)",
            "FWD Draft (m)",
            "AFT Draft (m)",
            "Trim (m)",
            "Ramp Angle (°)",
            "Target Trim (m)",
            "FWD (m) with Trim",
            "AFT (m) with Trim",
            "Notes",
        ]

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hdr_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        def nearest_formula_v4(col_letter, r):
            rngA = "Hourly_FWD_AFT_Heights!A:A"
            idx = f"MATCH(C{r},{rngA},1)"
            t1 = f"INDEX({rngA},{idx})"
            t2 = f"INDEX({rngA},{idx}+1)"
            cond_choose_left = f"ABS(C{r}-{t1})<=ABS({t2}-C{r})"
            pick_left_val = f'IF(ABS(C{r}-{t1})<=TIME(0,30,0),INDEX(Hourly_FWD_AFT_Heights!{col_letter}:{col_letter},{idx}),"")'
            pick_right_val = f'IF(ABS({t2}-C{r})<=TIME(0,30,0),INDEX(Hourly_FWD_AFT_Heights!{col_letter}:{col_letter},{idx}+1),"")'
            return f'=IF(C{r}="","",IF({cond_choose_left},{pick_left_val},{pick_right_val}))'

        stages = [
            (
                1,
                "Before Load-out (Empty Condition)",
                "",
                "Baseline from Initial Condition (26-Oct-2025)",
            ),
            (
                2,
                "SPMT 1st Entry on Ramp",
                "2025-12-01 00:00:00",
                "Ref time set to placeholder; update if different",
            ),
            (
                3,
                "50% on Ramp",
                "2025-12-01 01:00:00",
                "Ref time set to placeholder; update if different",
            ),
            (
                4,
                "Full on Ramp (Break-even)",
                "2025-12-01 04:00:00",
                "Ref time set to placeholder; update if different",
            ),
            (
                5,
                "Deck Full Load (217t × 2 on Deck)",
                "2025-12-01 05:00:00",
                "Ref time set to placeholder; update if different",
            ),
        ]

        for i, (stage_num, stage_name, ref_time, notes) in enumerate(stages, 2):
            ws.cell(row=i, column=1, value=stage_num)
            ws.cell(row=i, column=2, value=stage_name)
            ws.cell(row=i, column=3, value=ref_time if ref_time else "")
            ws.cell(row=i, column=3).number_format = "yyyy-mm-dd hh:mm"

            if i == 2:
                ws.cell(
                    row=i,
                    column=4,
                    value=f"=IF(C{i}=\"\",1.93,{nearest_formula_v4('C', i)[1:]})",
                )
            else:
                ws.cell(row=i, column=4, value=nearest_formula_v4("C", i))
            ws.cell(row=i, column=4).number_format = "0.00"

            if i == 2:
                ws.cell(
                    row=i,
                    column=5,
                    value=f"=IF(C{i}=\"\",1.93,{nearest_formula_v4('D', i)[1:]})",
                )
            else:
                ws.cell(row=i, column=5, value=nearest_formula_v4("D", i))
            ws.cell(row=i, column=5).number_format = "0.00"

            ws.cell(row=i, column=6, value=f'=IF(AND(D{i}<>"",E{i}<>""),E{i}-D{i},"")')
            ws.cell(row=i, column=6).number_format = "0.00"

            ws.cell(row=i, column=7, value=nearest_formula_v4("E", i))
            ws.cell(row=i, column=7).number_format = "0.00"

            ws.cell(row=i, column=8, value="")
            ws.cell(row=i, column=8).number_format = "0.00"

            mean_expr = f'IF(AND(D{i}<>"",E{i}<>""),(D{i}+E{i})/2,"")'
            ws.cell(
                row=i,
                column=9,
                value=f'=IF(H{i}="",D{i},IF({mean_expr}="",D{i},{mean_expr}-H{i}/2))',
            )
            ws.cell(row=i, column=9).number_format = "0.00"

            ws.cell(
                row=i,
                column=10,
                value=f'=IF(H{i}="",E{i},IF({mean_expr}="",E{i},{mean_expr}+H{i}/2))',
            )
            ws.cell(row=i, column=10).number_format = "0.00"

            ws.cell(row=i, column=11, value=notes)

            for c in range(1, 12):
                ws.cell(row=i, column=c).border = border

        col_widths = [8, 30, 22, 14, 14, 10, 14, 14, 16, 16, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(self.hybrid_file)
        print("\n[SUCCESS] Stage_Heights 시트 패치 완료!")
        print("=" * 80)

    def add_dropdown(self):
        """드롭다운 추가"""
        print("=" * 80)
        print("Stage_Heights 드롭다운 추가 - v4 HYBRID")
        print("=" * 80)

        wb = load_workbook(self.hybrid_file)
        print("[OK] v4 HYBRID file loaded")

        if "Stage_Heights" not in wb.sheetnames:
            print("[ERROR] Stage_Heights sheet not found")
            return False

        ws = wb["Stage_Heights"]
        print("[OK] Stage_Heights sheet found")

        print("\n[Adding Data Validation]")

        dv = DataValidation(
            type="list",
            formula1="Hourly_FWD_AFT_Heights!$A$2:$A$745",
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True,
            errorTitle="Invalid Time",
            error="Please select a time from the dropdown list (December 2025, 744 hours)",
            promptTitle="Select Reference Time",
            prompt="Choose a time from December 2025 tide table (744 hours available)",
        )

        dv.add("C2:C6")
        ws.add_data_validation(dv)
        print("  [OK] Data validation added to C2:C6")

        print("\n[Date Format Application]")
        for row in range(2, 7):
            cell = ws.cell(row=row, column=3)
            cell.number_format = "yyyy-mm-dd hh:mm"
            print(f"  [OK] C{row}: yyyy-mm-dd hh:mm format applied")

        print("\n[Number Format Application]")
        for row in range(2, 7):
            for col_letter in ["D", "E", "F", "G", "H", "I", "J"]:
                col = ord(col_letter) - ord("A") + 1
                cell = ws.cell(row=row, column=col)
                cell.number_format = "0.00"
        print("  [OK] Columns D-J: 0.00 format applied")

        wb.save(self.final_file)
        print("\n" + "=" * 80)
        print(f"[SUCCESS] 드롭다운 추가 완료! 저장: {self.final_file}")
        print("=" * 80)
        return True

    # ========== VALIDATION FUNCTIONS ==========

    def validate_hybrid(self):
        """HYBRID 파일 기본 검증"""
        print("=" * 80)
        print("LCT BUSHRA GateAB v4 HYBRID - Validation Script")
        print("=" * 80)

        try:
            wb = load_workbook(self.hybrid_file, data_only=False)
            print("[OK] File loaded successfully")
        except FileNotFoundError:
            print("[ERROR] LCT_BUSHRA_GateAB_v4_HYBRID.xlsx not found")
            return False

        errors = []
        warnings = []
        info = []

        calc = wb["Calc"]
        STANDARD_MAPPING = {
            "D8": {"param": "L_ramp_m", "expected_range": (10, 15), "critical": True},
            "D9": {
                "param": "theta_max_deg",
                "expected_range": (5, 7),
                "critical": True,
            },
            "D10": {"param": "KminusZ_m", "expected_range": (2, 5), "critical": True},
            "D11": {"param": "D_vessel_m", "expected_range": (3, 4), "critical": False},
            "D13": {
                "param": "min_fwd_draft_m",
                "expected_range": (1, 2),
                "critical": True,
            },
            "D14": {
                "param": "max_fwd_draft_m",
                "expected_range": (3, 4),
                "critical": True,
            },
            "D15": {
                "param": "pump_rate_tph",
                "expected_range": (8, 15),
                "critical": False,
            },
            "D17": {
                "param": "MTC_t_m_per_cm",
                "expected_range": (30, 40),
                "critical": True,
            },
            "D18": {
                "param": "LCF_m_from_midship",
                "expected_range": (-50, 50),
                "critical": True,
            },
            "D19": {
                "param": "TPC_t_per_cm",
                "expected_range": (0, 10),
                "critical": False,
            },
        }

        print("\n[Standard Cell Mapping Check]")
        for cell_ref, mapping in STANDARD_MAPPING.items():
            col = ord(cell_ref[0]) - ord("A") + 1
            row = int(cell_ref[1:])
            cell = calc.cell(row, col)
            value = cell.value
            param_name = mapping["param"]
            expected_range = mapping["expected_range"]
            is_critical = mapping["critical"]

            if value is None:
                if is_critical:
                    errors.append(
                        f"Calc: {cell_ref} ({param_name}) is EMPTY - CRITICAL PARAMETER"
                    )
                    print(f"  [ERROR] {cell_ref} ({param_name}): EMPTY [CRITICAL]")
                else:
                    warnings.append(
                        f"Calc: {cell_ref} ({param_name}) is empty (optional)"
                    )
                    print(f"  [WARNING] {cell_ref} ({param_name}): EMPTY [optional]")
            elif isinstance(value, (int, float)):
                if expected_range[0] <= value <= expected_range[1]:
                    print(f"  [OK] {cell_ref} ({param_name}): {value} [within range]")
                    info.append(f"Calc: {cell_ref} = {value} ({param_name})")
                else:
                    msg = f"Calc: {cell_ref} ({param_name}) = {value} [OUT OF RANGE {expected_range}]"
                    if is_critical:
                        errors.append(msg)
                        print(
                            f"  [ERROR] {cell_ref} ({param_name}): {value} [OUT OF RANGE] [CRITICAL]"
                        )
                    else:
                        warnings.append(msg)
                        print(f"  [WARNING] {cell_ref} ({param_name}): {value} [OUT OF RANGE]")
            else:
                errors.append(
                    f"Calc: {cell_ref} ({param_name}) is not numeric: {value}"
                )
                print(f"  [ERROR] {cell_ref} ({param_name}): {value} [NOT NUMERIC]")

        print("\n" + "=" * 80)
        print("FINAL VALIDATION REPORT")
        print("=" * 80)
        print(f"\n[Summary]")
        print(f"  Information: {len(info)}")
        print(f"  Warnings: {len(warnings)}")
        print(f"  Errors: {len(errors)}")

        if errors:
            print(f"\n[ERRORS - Must Fix Before Use]")
            for i, error in enumerate(errors, 1):
                print(f"  {i}. {error}")

        if warnings:
            print(f"\n[WARNINGS - Review Recommended]")
            for i, warning in enumerate(warnings, 1):
                print(f"  {i}. {warning}")

        report_filename = f"../output/validation_report_gateab_hybrid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(report_filename, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("LCT BUSHRA GateAB v4 HYBRID — Validation Report\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            f.write(
                f"SUMMARY\n  Information: {len(info)}\n  Warnings: {len(warnings)}\n  Errors: {len(errors)}\n\n"
            )
            if errors:
                f.write("ERRORS (Must Fix)\n")
                for i, error in enumerate(errors, 1):
                    f.write(f"  {i}. {error}\n")
            if warnings:
                f.write("WARNINGS (Review Recommended)\n")
                for i, warning in enumerate(warnings, 1):
                    f.write(f"  {i}. {warning}\n")

        print(f"\n[OK] Validation report saved: {report_filename}")
        print("\n" + "=" * 80)
        return len(errors) == 0

    def validate_dropdown(self):
        """드롭다운 기능 검증"""
        print("=" * 80)
        print("Stage_Heights 드롭다운 검증")
        print("=" * 80)

        try:
            wb = load_workbook(self.final_file, data_only=False)
            print("[OK] File loaded")
        except FileNotFoundError:
            print("[ERROR] Final file not found")
            return False

        ws = wb["Stage_Heights"]
        print(f"[OK] Stage_Heights: {ws.max_row}행 x {ws.max_column}열")

        print("\n[Data Validation Check]")
        if ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            print(f"  [OK] Data validations found: {dv_count}")
            for dv in ws.data_validations.dataValidation:
                print(
                    f"    Type: {dv.type}, Formula: {dv.formula1}, Applies to: {dv.sqref}"
                )
        else:
            print("  [WARNING] No data validations found")

        print("\n" + "=" * 80)
        print("[SUCCESS] 검증 완료")
        print("=" * 80)
        return True

    def comprehensive_validation(self):
        """종합 검증 리포트"""
        print("=" * 80)
        print("[INFO] MACHO-GPT LogiMaster - Comprehensive Validation Report")
        print("=" * 80)
        print(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        try:
            wb = load_workbook(self.hybrid_file, data_only=False)
            wb_data = load_workbook(self.hybrid_file, data_only=True)
            print("[OK] Files loaded successfully")
        except FileNotFoundError:
            print("[ERROR] File not found")
            return False

        validation_results = {
            "timestamp": datetime.now().isoformat(),
            "file_name": "LCT_BUSHRA_GateAB_v4_HYBRID.xlsx",
            "version": "v4 HYBRID",
            "sections": {},
        }

        errors = []
        warnings = []
        passed_tests = []

        expected_sheets = [
            "Calc",
            "December_Tide_2025",
            "Hourly_FWD_AFT_Heights",
            "RORO_Stage_Scenarios",
            "Formula_Test",
            "README",
            "Summary_요약",
            "실행_방법",
            "시트_구성_수식",
            "제출물_검수체크리스트",
            "STANDARD_좌표기준",
        ]

        actual_sheets = wb.sheetnames
        missing_sheets = set(expected_sheets) - set(actual_sheets)
        extra_sheets = set(actual_sheets) - set(expected_sheets)

        if not missing_sheets:
            passed_tests.append("File Structure: All sheets present")
        else:
            errors.append(f"Missing sheets: {', '.join(missing_sheets)}")

        validation_results["sections"]["file_structure"] = {
            "expected_sheets": len(expected_sheets),
            "actual_sheets": len(actual_sheets),
            "status": "PASS" if not missing_sheets else "FAIL",
        }

        # Calc 파라미터 검증
        calc = wb["Calc"]
        calc_data = wb_data["Calc"]
        STANDARD_MAPPING = {
            "D8": {"param": "L_ramp_m", "expected": 12, "tolerance": 1},
            "D9": {"param": "theta_max_deg", "expected": 6, "tolerance": 1},
            "D10": {"param": "KminusZ_m", "expected": 3, "tolerance": 2},
        }

        calc_validation = {}
        for cell_ref, mapping in STANDARD_MAPPING.items():
            col = ord(cell_ref[0]) - ord("A") + 1
            row = int(cell_ref[1:])
            value = calc_data.cell(row, col).value
            if value and isinstance(value, (int, float)):
                deviation = abs(value - mapping["expected"])
                if deviation <= mapping["tolerance"]:
                    passed_tests.append(f"Calc!{cell_ref} = {value}")
                    calc_validation[cell_ref] = {"status": "PASS", "value": value}
                else:
                    warnings.append(
                        f"Calc!{cell_ref} = {value} [deviation: {deviation:.2f}]"
                    )

        validation_results["sections"]["calc_parameters"] = calc_validation

        # 조수 데이터 검증
        tide = wb["December_Tide_2025"]
        expected_rows = 745
        actual_rows = tide.max_row
        if actual_rows == expected_rows:
            passed_tests.append(f"Tide data: {actual_rows} rows complete")
        else:
            errors.append(
                f"Tide data row count mismatch: {actual_rows} vs {expected_rows}"
            )

        validation_results["sections"]["tide_data"] = {
            "rows": actual_rows,
            "expected_rows": expected_rows,
            "status": "PASS" if actual_rows == expected_rows else "FAIL",
        }

        # Formula_Test 검증
        try:
            test_data = wb_data["Formula_Test"]
            test_results = {}
            for row in range(5, 8):
                test_name = test_data.cell(row, 1).value
                result_cell = test_data.cell(row, 11).value
                if test_name and result_cell:
                    test_results[test_name] = result_cell
                    if result_cell == "PASS":
                        passed_tests.append(f"Formula_Test: Test {test_name} PASS")
                    else:
                        errors.append(f"Formula_Test: Test {test_name} FAILED")
        except KeyError:
            errors.append("Formula_Test sheet not found")

        total_checks = len(passed_tests) + len(warnings) + len(errors)
        score_percentage = (
            (len(passed_tests) / total_checks * 100) if total_checks > 0 else 0
        )

        if score_percentage >= 95 and len(errors) == 0:
            grade = "A+ (EXCELLENT)"
        elif score_percentage >= 90 and len(errors) == 0:
            grade = "A (VERY GOOD)"
        else:
            grade = "B (GOOD)"

        validation_results["summary"] = {
            "total_checks": total_checks,
            "passed": len(passed_tests),
            "warnings": len(warnings),
            "errors": len(errors),
            "score_percentage": score_percentage,
            "grade": grade,
        }

        report_filename = f"../output/comprehensive_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_filename, "w", encoding="utf-8") as f:
            json.dump(validation_results, f, indent=2, ensure_ascii=False)

        text_report_filename = f"../output/comprehensive_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(text_report_filename, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("LCT BUSHRA GateAB v4 HYBRID — COMPREHENSIVE VALIDATION REPORT\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Overall Grade: {grade}\n")
            f.write(
                f"Score: {score_percentage:.1f}% ({len(passed_tests)}/{total_checks} checks passed)\n"
            )
            f.write(f"Errors: {len(errors)}\n")
            f.write(f"Warnings: {len(warnings)}\n\n")
            if errors:
                f.write("CRITICAL ERRORS\n")
                for i, error in enumerate(errors, 1):
                    f.write(f"{i}. {error}\n")

        print(f"\n[OK] JSON report saved: {report_filename}")
        print(f"[OK] Text report saved: {text_report_filename}")
        print("\n" + "=" * 80)
        return len(errors) == 0

    # ========== ANALYSIS FUNCTIONS ==========

    def realtime_analysis(self):
        """실시간 Draft 분석"""
        print("=" * 80)
        print("[INFO] MACHO-GPT LogiMaster - Stage 실시간 Draft 분석")
        print("=" * 80)
        print(f"분석 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        try:
            wb = load_workbook(self.final_file, data_only=True)
            print("[OK] v4 HYBRID 파일 로드 완료")
        except FileNotFoundError:
            print("[ERROR] Final file not found")
            return False

        calc = wb["Calc"]
        params = {
            "L_ramp": calc.cell(8, 4).value,
            "theta_max": calc.cell(9, 4).value,
            "KminusZ": calc.cell(10, 4).value,
            "min_draft": calc.cell(13, 4).value,
            "max_draft": calc.cell(14, 4).value,
            "MTC": calc.cell(17, 4).value,
            "LCF": calc.cell(18, 4).value,
        }

        tide_sheet = wb["December_Tide_2025"]
        tide_data = []
        for row in range(2, tide_sheet.max_row + 1):
            dt = tide_sheet.cell(row, 1).value
            tide = tide_sheet.cell(row, 2).value
            if dt and tide:
                tide_data.append(
                    {
                        "datetime": dt,
                        "tide_m": float(tide) if isinstance(tide, (int, float)) else 0,
                    }
                )

        hourly_data = []
        for tide_entry in tide_data:
            dt = tide_entry["datetime"]
            tide = tide_entry["tide_m"]
            dfwd = (
                params["KminusZ"]
                + tide
                - params["L_ramp"] * math.tan(math.radians(params["theta_max"]))
            )
            daft = dfwd
            angle = math.degrees(
                math.atan((params["KminusZ"] - dfwd + tide) / params["L_ramp"])
            )
            status = (
                "OK"
                if (
                    params["min_draft"] <= dfwd <= params["max_draft"]
                    and angle <= params["theta_max"]
                )
                else "CHECK"
            )

            hourly_data.append(
                {
                    "datetime": str(dt),
                    "tide_m": tide,
                    "dfwd_m": dfwd,
                    "daft_m": daft,
                    "status": status,
                    "angle_deg": angle,
                }
            )

        df = pd.DataFrame(hourly_data)
        ok_windows = df[df["status"] == "OK"].copy()
        optimal = (
            ok_windows[(ok_windows["dfwd_m"] >= 2.0) & (ok_windows["dfwd_m"] <= 3.0)]
            if len(ok_windows) > 0
            else pd.DataFrame()
        )

        print(f"\n[Status=OK 시간대]")
        print(
            f"  전체: {len(ok_windows)}/{len(df)} ({len(ok_windows)/len(df)*100:.1f}%)"
        )

        if len(ok_windows) > 0:
            print(f"  Dfwd 최소: {ok_windows['dfwd_m'].min():.2f}m")
            print(f"  Dfwd 최대: {ok_windows['dfwd_m'].max():.2f}m")
            print(f"  Dfwd 평균: {ok_windows['dfwd_m'].mean():.2f}m")

        results = {
            "analysis_time": datetime.now().isoformat(),
            "parameters": params,
            "tide_data_count": len(tide_data),
            "ok_hours": len(ok_windows),
            "optimal_hours": len(optimal),
            "kpi": {
                "workable_percentage": len(ok_windows) / 744 * 100,
                "optimal_percentage": len(optimal) / 744 * 100,
            },
        }

        json_file = f"../data/stage_realtime_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2, ensure_ascii=False)

        csv_file = f"../data/stage_analysis_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        df.to_csv(csv_file, index=False, encoding="utf-8-sig")

        print(f"\n[OK] 분석 결과 저장: {json_file}")
        print(f"[OK] 데이터 저장: {csv_file}")
        print("\n" + "=" * 80)
        return True

    def analyze_v3(self):
        """v3 원본 파일 분석"""
        print("=" * 80)
        print("Bushra_GateAB_Updated_v3.xlsx 상세 분석")
        print("=" * 80)

        try:
            wb = load_workbook(self.v3_file, data_only=False)
        except FileNotFoundError:
            print("[ERROR] v3 file not found")
            return False

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'='*80}")
            print(f"시트: {sheet_name}")
            print(f"크기: {ws.max_row}행 x {ws.max_column}열")
            print(f"{'='*80}")

            if ws.max_row <= 20:
                print("\n[전체 내용]")
                for row in range(1, ws.max_row + 1):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row, col)
                        val = cell.value
                        if val is not None:
                            if isinstance(val, str) and len(val) > 50:
                                val = val[:50] + "..."
                            row_data.append(f"{get_column_letter(col)}{row}:{val}")
                    if row_data:
                        print(f"  행 {row}: {' | '.join(row_data)}")

        print("\n" + "=" * 80)
        print("분석 완료")
        print("=" * 80)
        return True


def main():
    parser = argparse.ArgumentParser(description="LCT BUSHRA 통합 운영 스크립트")
    parser.add_argument("--patch", action="store_true", help="Stage_Heights 시트 패치")
    parser.add_argument("--dropdown", action="store_true", help="드롭다운 추가")
    parser.add_argument("--validate", action="store_true", help="기본 검증")
    parser.add_argument(
        "--validate-dropdown", action="store_true", help="드롭다운 검증"
    )
    parser.add_argument("--comprehensive", action="store_true", help="종합 검증")
    parser.add_argument("--analyze", action="store_true", help="실시간 분석")
    parser.add_argument("--analyze-v3", action="store_true", help="v3 원본 분석")

    args = parser.parse_args()

    if not any(vars(args).values()):
        parser.print_help()
        return

    ops = BushraOperations()

    if args.patch:
        ops.patch_stage_heights()

    if args.dropdown:
        ops.add_dropdown()

    if args.validate:
        ops.validate_hybrid()

    if args.validate_dropdown:
        ops.validate_dropdown()

    if args.comprehensive:
        ops.comprehensive_validation()

    if args.analyze:
        ops.realtime_analysis()

    if args.analyze_v3:
        ops.analyze_v3()


if __name__ == "__main__":
    main()
