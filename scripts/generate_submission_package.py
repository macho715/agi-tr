# -*- coding: utf-8 -*-
# generate_submission_package.py
# Generates complete submission package for Harbor Master / Port Authority

import os
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Protection
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.patches as mpatches
from matplotlib.table import Table

# Configuration
EXCEL_INPUT = "../output/LCT_BUSHRA_GateAB_v4_HYBRID.xlsx"  # Updated to v4 HYBRID
OUTPUT_DIR = "SUBMISSION_PACKAGE"
PDF_REPORT = "LCT_BUSHRA_FWD_AFT_Report.pdf"
LOCKED_EXCEL = "LCT_BUSHRA_Calculator_LOCKED.xlsx"

# Submission metadata
SUBMISSION_INFO = {
    "vessel_name": "LCT BUSHRA",
    "imo_number": "IMO XXXXXXX",  # Update with actual IMO
    "project": "HVDC Transformer Transportation",
    "route": "Mina Zayed Port → DAS Island",
    "operator": "Samsung C&T Engineering & Construction",
    "date_prepared": datetime.now().strftime("%Y-%m-%d"),
    "prepared_by": "Samsung Logistics Team",
    "approved_by": "[Name - Chief Officer]",  # To be filled
}

def create_output_directory():
    """Create output directory structure"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        os.makedirs(os.path.join(OUTPUT_DIR, "01_PDF_Report"))
        os.makedirs(os.path.join(OUTPUT_DIR, "02_Working_Excel"))
        os.makedirs(os.path.join(OUTPUT_DIR, "03_Supporting_Documents"))
        print(f"✓ Created directory structure: {OUTPUT_DIR}")
    else:
        print(f"✓ Output directory exists: {OUTPUT_DIR}")

def load_excel_data():
    """Load data from Excel workbook"""
    try:
        wb = load_workbook(EXCEL_INPUT, data_only=True)
        
        # Read Calc sheet constants
        calc_ws = wb["Calc"]
        constants = {
            "KminusZ_m": calc_ws["D6"].value,
            "L_ramp_m": calc_ws["D4"].value,
            "theta_max_deg": calc_ws["D5"].value,
            "D_vessel_m": calc_ws["D7"].value,
            "Lpp_m": calc_ws["D16"].value,
        }
        
        # Read Hourly data (first 48 hours for report)
        hourly_ws = wb["Hourly_FWD_AFT_Heights"]
        hourly_data = []
        for row in range(2, 50):  # First 48 hours
            try:
                dt = hourly_ws.cell(row=row, column=1).value
                tide = hourly_ws.cell(row=row, column=2).value
                dfwd_req = hourly_ws.cell(row=row, column=3).value
                trim = hourly_ws.cell(row=row, column=4).value
                dfwd_adj = hourly_ws.cell(row=row, column=5).value
                daft_adj = hourly_ws.cell(row=row, column=6).value
                angle = hourly_ws.cell(row=row, column=7).value
                status = hourly_ws.cell(row=row, column=8).value
                
                if dt and tide is not None:
                    hourly_data.append({
                        "datetime": dt,
                        "tide": tide,
                        "dfwd_req": dfwd_req,
                        "trim": trim if trim else 0,
                        "dfwd_adj": dfwd_adj,
                        "daft_adj": daft_adj,
                        "angle": angle,
                        "status": status,
                    })
            except Exception as e:
                continue
        
        wb.close()
        print(f"✓ Loaded {len(hourly_data)} hourly records from Excel")
        return constants, hourly_data
    
    except FileNotFoundError:
        print(f"✗ ERROR: {EXCEL_INPUT} not found!")
        print(f"  Please run patch3.py first to generate the Excel file.")
        sys.exit(1)
    except Exception as e:
        print(f"✗ ERROR loading Excel: {e}")
        sys.exit(1)

def generate_pdf_report(constants, hourly_data):
    """Generate FWD/AFT Report PDF (1-2 pages)"""
    pdf_path = os.path.join(OUTPUT_DIR, "01_PDF_Report", PDF_REPORT)
    
    try:
        with PdfPages(pdf_path) as pdf:
            # Page 1: Header + First 24 hours
            fig, ax = plt.subplots(figsize=(11.69, 8.27))  # A4 landscape
            ax.axis('off')
            
            # Title
            y = 0.98
            ax.text(0.5, y, "LCT BUSHRA — FWD/AFT DRAFT REPORT", 
                   ha='center', va='top', fontsize=16, weight='bold')
            y -= 0.04
            
            # Metadata table
            metadata_text = f"""
Vessel: {SUBMISSION_INFO['vessel_name']} | IMO: {SUBMISSION_INFO['imo_number']}
Project: {SUBMISSION_INFO['project']}
Route: {SUBMISSION_INFO['route']}
Prepared: {SUBMISSION_INFO['date_prepared']} | By: {SUBMISSION_INFO['prepared_by']}
            """
            ax.text(0.02, y, metadata_text.strip(), ha='left', va='top', 
                   fontsize=9, family='monospace')
            y -= 0.12
            
            # Constants section
            const_text = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CRITICAL PARAMETERS (Measured on site)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
K–Z Distance (measured): {constants['KminusZ_m']:.2f} m
Linkspan Length: {constants['L_ramp_m']:.1f} m
Maximum Ramp Angle: {constants['theta_max_deg']:.1f}° (Harbour Master Limit)
Vessel Molded Depth: {constants['D_vessel_m']:.2f} m
Length Between Perpendiculars: {constants['Lpp_m']:.1f} m
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            """
            ax.text(0.02, y, const_text.strip(), ha='left', va='top',
                   fontsize=8, family='monospace',
                   bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.5))
            y -= 0.18
            
            # Table header
            ax.text(0.02, y, "HOURLY FWD/AFT DRAFT SCHEDULE (First 24 hours)",
                   ha='left', va='top', fontsize=10, weight='bold')
            y -= 0.03
            
            # Prepare table data (first 24 hours)
            table_data = [["Time (GST)", "Tide\n(m)", "Dfwd_req\n(m)", "Trim\n(m)", 
                          "Dfwd_adj\n(m)", "Daft_adj\n(m)", "Ramp∠\n(deg)", "Status", "Notes"]]
            
            for i, record in enumerate(hourly_data[:24]):
                dt_str = record['datetime'] if isinstance(record['datetime'], str) else record['datetime'].strftime("%m-%d %H:%M")
                trim_str = f"{record['trim']:.2f}" if record['trim'] and record['trim'] != 0 else "-"
                notes = "Even-keel" if not record['trim'] or record['trim'] == 0 else "Trim adjusted"
                
                # Safe number formatting
                def safe_float_format(val, fmt=".2f"):
                    if val is None:
                        return "-"
                    try:
                        return f"{float(val):{fmt}}"
                    except (TypeError, ValueError):
                        return str(val) if val else "-"
                
                table_data.append([
                    dt_str,
                    safe_float_format(record['tide']),
                    safe_float_format(record['dfwd_req']),
                    trim_str,
                    safe_float_format(record['dfwd_adj']),
                    safe_float_format(record['daft_adj']),
                    safe_float_format(record['angle'], ".1f"),
                    record['status'] if record['status'] else "-",
                    notes
                ])
            
            # Create table
            table = ax.table(cellText=table_data, 
                           colWidths=[0.11, 0.07, 0.09, 0.07, 0.09, 0.09, 0.08, 0.08, 0.14],
                           loc='center',
                           bbox=[0.02, 0.05, 0.96, y-0.08])
            
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            table.scale(1, 1.3)
            
            # Style header row
            for i in range(9):
                cell = table[(0, i)]
                cell.set_facecolor('#1F4E78')
                cell.set_text_props(weight='bold', color='white')
            
            # Color code status column
            for i in range(1, len(table_data)):
                status_cell = table[(i, 7)]
                if table_data[i][7] == "OK":
                    status_cell.set_facecolor('#90EE90')
                elif table_data[i][7] == "CHECK":
                    status_cell.set_facecolor('#FFD700')
            
            # Footer
            ax.text(0.02, 0.01, 
                   "⚠ This report is based on calculated values. Actual conditions may vary. Always verify with real-time measurements.",
                   ha='left', va='bottom', fontsize=7, style='italic', color='red')
            
            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)
            
            # Page 2: Next 24 hours (if available)
            if len(hourly_data) > 24:
                fig2, ax2 = plt.subplots(figsize=(11.69, 8.27))
                ax2.axis('off')
                
                y2 = 0.98
                ax2.text(0.5, y2, "LCT BUSHRA — FWD/AFT DRAFT REPORT (Continued)", 
                        ha='center', va='top', fontsize=16, weight='bold')
                y2 -= 0.06
                
                ax2.text(0.02, y2, "HOURLY FWD/AFT DRAFT SCHEDULE (Hours 25-48)",
                        ha='left', va='top', fontsize=10, weight='bold')
                y2 -= 0.03
                
                # Table for hours 25-48
                table_data2 = [["Time (GST)", "Tide\n(m)", "Dfwd_req\n(m)", "Trim\n(m)", 
                               "Dfwd_adj\n(m)", "Daft_adj\n(m)", "Ramp∠\n(deg)", "Status", "Notes"]]
                
                for record in hourly_data[24:48]:
                    dt_str = record['datetime'] if isinstance(record['datetime'], str) else record['datetime'].strftime("%m-%d %H:%M")
                    trim_str = f"{record['trim']:.2f}" if record['trim'] and record['trim'] != 0 else "-"
                    notes = "Even-keel" if not record['trim'] or record['trim'] == 0 else "Trim adjusted"
                    
                    # Safe number formatting
                    def safe_float_format(val, fmt=".2f"):
                        if val is None:
                            return "-"
                        try:
                            return f"{float(val):{fmt}}"
                        except (TypeError, ValueError):
                            return str(val) if val else "-"
                    
                    table_data2.append([
                        dt_str,
                        safe_float_format(record['tide']),
                        safe_float_format(record['dfwd_req']),
                        trim_str,
                        safe_float_format(record['dfwd_adj']),
                        safe_float_format(record['daft_adj']),
                        safe_float_format(record['angle'], ".1f"),
                        record['status'] if record['status'] else "-",
                        notes
                    ])
                
                table2 = ax2.table(cellText=table_data2,
                                 colWidths=[0.11, 0.07, 0.09, 0.07, 0.09, 0.09, 0.08, 0.08, 0.14],
                                 loc='center',
                                 bbox=[0.02, 0.10, 0.96, y2-0.13])
                
                table2.auto_set_font_size(False)
                table2.set_fontsize(7)
                table2.scale(1, 1.3)
                
                # Style header
                for i in range(9):
                    cell = table2[(0, i)]
                    cell.set_facecolor('#1F4E78')
                    cell.set_text_props(weight='bold', color='white')
                
                # Color code status
                for i in range(1, len(table_data2)):
                    status_cell = table2[(i, 7)]
                    if table_data2[i][7] == "OK":
                        status_cell.set_facecolor('#90EE90')
                    elif table_data2[i][7] == "CHECK":
                        status_cell.set_facecolor('#FFD700')
                
                # Signature section
                ax2.text(0.02, 0.06, "APPROVALS", fontsize=10, weight='bold')
                approval_text = f"""
Prepared by: {SUBMISSION_INFO['prepared_by']}
Signature: ________________    Date: ___________

Reviewed by: {SUBMISSION_INFO['approved_by']}
Signature: ________________    Date: ___________

Harbor Master Approval:
Signature: ________________    Date: ___________
                """
                ax2.text(0.02, 0.05, approval_text.strip(), ha='left', va='top', fontsize=8)
                
                pdf.savefig(fig2, bbox_inches='tight')
                plt.close(fig2)
        
        print(f"✓ PDF report generated: {pdf_path}")
        return pdf_path
    
    except Exception as e:
        print(f"✗ ERROR generating PDF: {e}")
        return None

def create_locked_excel():
    """Create locked version of Excel with formulas protected"""
    try:
        # Load workbook
        wb = load_workbook(EXCEL_INPUT)
        
        # Protect calculation sheets
        for sheet_name in ["Calc", "Hourly_FWD_AFT_Heights", "RORO_Stage_Scenarios"]:
            ws = wb[sheet_name]
            
            # Unlock input cells (yellow cells)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill.fgColor.rgb == "FFFFF2CC":  # Yellow fill
                        cell.protection = Protection(locked=False)
                    else:
                        cell.protection = Protection(locked=True)
            
            # Protect sheet (allow selecting but not editing formulas)
            ws.protection.sheet = True
            ws.protection.password = "BUSHRA2025"  # Set password
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False
            ws.protection.insertColumns = False
            ws.protection.insertRows = False
            ws.protection.deleteColumns = False
            ws.protection.deleteRows = False
        
        # Save locked version
        locked_path = os.path.join(OUTPUT_DIR, "02_Working_Excel", LOCKED_EXCEL)
        wb.save(locked_path)
        wb.close()
        
        print(f"✓ Locked Excel created: {locked_path}")
        print(f"  Password: BUSHRA2025 (for unprotecting sheets)")
        return locked_path
    
    except Exception as e:
        print(f"✗ ERROR creating locked Excel: {e}")
        return None

def create_supporting_documents():
    """Create templates for supporting documents"""
    support_dir = os.path.join(OUTPUT_DIR, "03_Supporting_Documents")
    
    # 1. K-Z Measurement Record Template
    kz_template = """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
K-Z DISTANCE MEASUREMENT RECORD
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Vessel: LCT BUSHRA
Location: Mina Zayed Port, Berth [___]
Date: ________________
Time (GST): ________________

MEASUREMENT TEAM:
- Surveyor: ________________
- Chief Officer: ________________
- Samsung Representative: ________________

ENVIRONMENTAL CONDITIONS:
- Tide Level: ________ m (CD)
- Weather: ________________
- Sea State: ________________

MEASUREMENT METHOD:
1. Position vessel parallel to jetty
2. Measure from jetty deck level to linkspan contact point
3. Account for:
   - Jetty elevation above Chart Datum: ________ m
   - Vessel freeboard at measurement point: ________ m
   - Tide height at measurement time: ________ m

MEASUREMENTS:
┌─────────────────────────────────────────────────┐
│ Measurement Point  │ Reading (m) │ Notes       │
├────────────────────┼─────────────┼─────────────┤
│ Jetty Deck Level   │             │             │
│ Contact Point (K)  │             │             │
│ Calculated K-Z     │             │ ◄── FINAL   │
└─────────────────────────────────────────────────┘

VERIFICATION:
- Repeat measurement: ________ m
- Difference: ________ m (should be < 0.05m)

PHOTO DOCUMENTATION:
- Photo 1: Overall view of measurement setup
- Photo 2: Close-up of contact point
- Photo 3: Measuring equipment with reading
- Photos attached: □ Yes  □ No

FINAL K-Z VALUE: ________ m

SIGNATURES:
Surveyor: ________________  Date: ________
Chief Officer: ________________  Date: ________
Harbor Master (if present): ________________  Date: ________

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
IMPORTANT: This value must be entered in Calc!D6 before operations.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    with open(os.path.join(support_dir, "KZ_Measurement_Record_TEMPLATE.txt"), 'w', encoding='utf-8') as f:
        f.write(kz_template)
    
    # 2. Tide Data Source Declaration
    tide_source = """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TIDE DATA SOURCE DECLARATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Project: HVDC Transformer Transportation
Vessel: LCT BUSHRA
Location: Mina Zayed Port → DAS Island
Period: December 2025

OFFICIAL TIDE DATA SOURCE:
□ Abu Dhabi Ports Authority (AD Ports)
□ ADNOC Marine & Offshore Services
□ Other: ____________________

Source Details:
- Document Reference: ____________________
- Publication Date: ____________________
- Validity Period: ____________________
- Datum: Chart Datum (CD) / Other: ____________

DATA CHARACTERISTICS:
- Temporal Resolution: Hourly / Other: ____________
- Predicted / Observed: ____________
- Corrections Applied: Yes □  No □
  If Yes, specify: ____________________

VERIFICATION:
- Cross-checked with: ____________________
- Verified by: ____________________
- Date: ____________________

ATTACHMENT:
- Original tide table attached: □ Yes  □ No
- Format: PDF / Excel / Other: ____________

DECLARATION:
I hereby declare that the tide data used in the calculation workbook
(December_Tide_2025 sheet) is sourced from the official source stated
above and represents the best available prediction for the operational
period.

Name: ____________________
Position: ____________________
Signature: ____________________
Date: ____________________

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    with open(os.path.join(support_dir, "Tide_Data_Source_Declaration_TEMPLATE.txt"), 'w', encoding='utf-8') as f:
        f.write(tide_source)
    
    # 3. Submission Checklist
    checklist = """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HARBOR MASTER SUBMISSION CHECKLIST
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Vessel: LCT BUSHRA
Submission Date: ____________________
Operation Date: ____________________

REQUIRED DOCUMENTS:

01. PDF REPORT (Mandatory)
    □ LCT_BUSHRA_FWD_AFT_Report.pdf
    □ Contains first 48 hours of operations
    □ K-Z value clearly stated
    □ All parameters verified
    □ Signatures obtained

02. WORKING EXCEL (Reference)
    □ LCT_BUSHRA_Calculator_LOCKED.xlsx
    □ Calc!D6 (K-Z) updated with site measurement
    □ December_Tide_2025 sheet populated with official data
    □ All formulas verified (no #DIV/0! or #REF! errors)
    □ Sheet protection enabled

03. SUPPORTING DOCUMENTS
    □ K-Z Measurement Record
      - Measurement date/time: ____________________
      - Surveyor name: ____________________
      - Photos attached (min 3): □ Yes  □ No
    
    □ Tide Data Source Declaration
      - Source: ____________________
      - Date range: ____________________
      - Official stamp/signature: □ Yes  □ No
    
    □ Vessel Documentation
      - Stability Booklet: □ Attached  □ On file
      - Certificate of Class: □ Attached  □ On file
      - Crew List: □ Attached  □ On file

TECHNICAL VERIFICATION:

□ K-Z distance verified on site
□ Linkspan specifications confirmed (12m length)
□ Maximum ramp angle ≤ 6° in all calculated scenarios
□ All "Status" cells in Hourly sheet show "OK" for intended operation window
□ Trim Check in RORO sheet: No "EXCESSIVE" warnings (or explained if present)
□ Ballast operations feasible within timeline

OPERATIONAL APPROVAL:

□ Chief Officer review completed
□ Samsung C&T approval obtained
□ Harbor Master preliminary review: □ Pass  □ Needs revision

SUBMISSION METHOD:
□ Email to: ____________________
□ Physical copy delivered to: ____________________
□ Online portal: ____________________

SUBMITTED BY:
Name: ____________________
Company: Samsung C&T Engineering & Construction
Position: ____________________
Contact: ____________________
Signature: ____________________
Date: ____________________

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FOR HARBOR MASTER USE ONLY:

Received: ____________________
Reviewed by: ____________________
Approval Status: □ Approved  □ Approved with conditions  □ Rejected
Conditions/Comments:
_________________________________________________________________
_________________________________________________________________

Harbor Master Signature: ____________________
Date: ____________________
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    with open(os.path.join(support_dir, "Submission_Checklist.txt"), 'w', encoding='utf-8') as f:
        f.write(checklist)
    
    print(f"✓ Supporting document templates created in: {support_dir}")
    return support_dir

def create_readme():
    """Create README for submission package"""
    readme_content = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LCT BUSHRA - RORO OPERATION SUBMISSION PACKAGE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Project: {SUBMISSION_INFO['project']}
Vessel: {SUBMISSION_INFO['vessel_name']}

PACKAGE CONTENTS:

01_PDF_Report/
  └── LCT_BUSHRA_FWD_AFT_Report.pdf
      Primary submission document for Harbor Master
      Contains:
      - Critical parameters (K-Z, linkspan length, max angle)
      - 48-hour hourly draft schedule
      - Status indicators (OK/CHECK)
      - Approval signature blocks

02_Working_Excel/
  └── LCT_BUSHRA_Calculator_LOCKED.xlsx
      Reference calculation workbook
      Features:
      - All calculation sheets protected (password: BUSHRA2025)
      - Yellow cells remain editable (inputs)
      - Formulas locked to prevent accidental changes
      - Real-time recalculation enabled

03_Supporting_Documents/
  ├── KZ_Measurement_Record_TEMPLATE.txt
  │   Template for documenting K-Z site measurement
  │   **ACTION REQUIRED:** Fill this out during site survey
  │
  ├── Tide_Data_Source_Declaration_TEMPLATE.txt
  │   Template for declaring official tide data source
  │   **ACTION REQUIRED:** Complete with actual source details
  │
  └── Submission_Checklist.txt
      Master checklist for submission preparation
      Use this to verify all documents are ready

SUBMISSION WORKFLOW:

STEP 1: PRE-SUBMISSION (Before operations)
  □ Conduct K-Z measurement on site
  □ Fill KZ_Measurement_Record_TEMPLATE.txt
  □ Take photos (minimum 3)
  □ Update Calc!D6 in Excel with measured K-Z value
  
STEP 2: DATA PREPARATION
  □ Obtain official tide data (AD Ports/ADNOC)
  □ Paste into December_Tide_2025 sheet (B2:B745)
  □ Fill Tide_Data_Source_Declaration_TEMPLATE.txt
  □ Verify all calculations (check for errors)

STEP 3: DOCUMENT GENERATION
  □ Run generate_submission_package.py
  □ Review PDF report for accuracy
  □ Verify Excel workbook functionality
  
STEP 4: FINAL SUBMISSION
  □ Complete Submission_Checklist.txt
  □ Obtain internal approvals (Chief Officer, Samsung)
  □ Submit package to Harbor Master:
     - Email PDF report
     - Provide Excel workbook (optional, for reference)
     - Attach supporting documents

CRITICAL REMINDERS:

⚠ K-Z DISTANCE: Must be measured on site. The default value (3.0m)
  in the Excel is a placeholder ONLY. Using incorrect K-Z can result
  in dangerous ramp angles.

⚠ TIDE DATA: Must be from official source (AD Ports or ADNOC). Using
  unofficial or inaccurate tide data may result in operational delays
  or safety incidents.

⚠ VERIFICATION: Before submission, verify that:
  - All "Status" cells show "OK" for intended operation window
  - Ramp angles are ≤ 6° (Harbor Master limit)
  - No #DIV/0! or #REF! errors in Excel
  - K-Z measurement is documented with photos

TECHNICAL SUPPORT:

For questions or issues with this package:
- Samsung C&T Logistics Team: [contact]
- MACHO-GPT System: [contact]
- Marine Operations Coordinator: [contact]

DOCUMENT VERSION:
- Excel Calculator: Patch 3 (2025-11-06)
- PDF Report Generator: Version 1.0
- Package Structure: Version 1.0

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONFIDENTIAL: This package contains operational data for HVDC project.
Do not distribute outside authorized project personnel and port authority.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    """
    
    readme_path = os.path.join(OUTPUT_DIR, "README.txt")
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print(f"✓ README created: {readme_path}")
    return readme_path

def main():
    """Main execution"""
    print("\n" + "="*70)
    print("LCT BUSHRA - SUBMISSION PACKAGE GENERATOR")
    print("="*70)
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Vessel: {SUBMISSION_INFO['vessel_name']}")
    print(f"Project: {SUBMISSION_INFO['project']}")
    print("="*70 + "\n")
    
    # Step 1: Create directory structure
    print("[1/6] Creating output directory structure...")
    create_output_directory()
    
    # Step 2: Load Excel data
    print("\n[2/6] Loading data from Excel workbook...")
    constants, hourly_data = load_excel_data()
    
    # Step 3: Generate PDF report
    print("\n[3/6] Generating PDF report...")
    pdf_path = generate_pdf_report(constants, hourly_data)
    
    # Step 4: Create locked Excel
    print("\n[4/6] Creating locked Excel workbook...")
    excel_path = create_locked_excel()
    
    # Step 5: Create supporting documents
    print("\n[5/6] Creating supporting document templates...")
    support_dir = create_supporting_documents()
    
    # Step 6: Create README
    print("\n[6/6] Creating package README...")
    readme_path = create_readme()
    
    # Summary
    print("\n" + "="*70)
    print("PACKAGE GENERATION COMPLETE")
    print("="*70)
    print(f"\nOutput directory: {OUTPUT_DIR}/")
    print("\nGenerated files:")
    print(f"  ✓ PDF Report: {PDF_REPORT}")
    print(f"  ✓ Locked Excel: {LOCKED_EXCEL}")
    print(f"  ✓ Supporting docs: 3 templates")
    print(f"  ✓ README: README.txt")
    
    print("\n" + "-"*70)
    print("NEXT STEPS:")
    print("-"*70)
    print("1. Conduct K-Z measurement on site")
    print("2. Update KZ_Measurement_Record_TEMPLATE.txt")
    print("3. Update Calc!D6 in Excel with actual K-Z value")
    print("4. Paste official tide data into December_Tide_2025 sheet")
    print("5. Fill Tide_Data_Source_Declaration_TEMPLATE.txt")
    print("6. Re-run this script to regenerate PDF with updated data")
    print("7. Complete Submission_Checklist.txt")
    print("8. Submit package to Harbor Master")
    print("-"*70)
    
    print(f"\n⚠ CRITICAL: Default K-Z = {constants['KminusZ_m']}m is a placeholder!")
    print("   Must be updated with site measurement before operations.\n")

if __name__ == "__main__":
    main()
