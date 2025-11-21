# -*- coding: utf-8 -*-
# build_bushra_agi_tr_from_scratch.py
# LCT_BUSHRA_AGI_TR.xlsx 파일을 프로그래밍 방식으로 Excel 함수를 생성하여 만드는 스크립트
# 원본 파일을 복사하지 않고, 수식을 직접 생성하여 동일한 구조의 Excel 파일 생성

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
import os
import sys
import json
from datetime import datetime

OUTPUT_FILE = "../../output/LCT_BUSHRA_AGI_TR_from_scratch.xlsx"

# ============================================================================
# Helper Functions
# ============================================================================

# 스타일 정의
def get_styles():
    """공통 스타일 정의"""
    return {
        "title_font": Font(name="Calibri", size=18, bold=True),
        "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "normal_font": Font(name="Calibri", size=11),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "ok_fill": PatternFill("solid", fgColor="C6E0B4"),
        "ng_fill": PatternFill("solid", fgColor="F8CBAD"),
        "thin_border": Side(border_style="thin", color="C0C0C0"),
        "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }

# 함수 생성 헬퍼 함수
def create_index_match_formula(lookup_value, lookup_range, return_range):
    """INDEX/MATCH 조합 수식 생성"""
    return f'=INDEX({return_range}, MATCH("{lookup_value}", {lookup_range}, 0))'

def create_if_formula(condition, true_value, false_value=""):
    """IF 수식 생성"""
    if isinstance(false_value, str) and false_value:
        return f'=IF({condition}, {true_value}, "{false_value}")'
    elif false_value == "":
        return f'=IF({condition}, {true_value}, "")'
    else:
        return f'=IF({condition}, {true_value}, {false_value})'

def create_if_or_formula(conditions, true_value, false_value=""):
    """IF/OR 조합 수식 생성"""
    or_conditions = "OR(" + ", ".join(conditions) + ")"
    if isinstance(false_value, str) and false_value:
        return f'=IF({or_conditions}, {true_value}, "{false_value}")'
    else:
        return f'=IF({or_conditions}, {true_value}, "")'


# ============================================================================
# Calc Sheet Creation
# ============================================================================

def create_calc_sheet(wb):
    """Calc 시트 생성 - 원본 파일 구조와 동일하게 생성"""
    ws = wb.create_sheet("Calc")
    styles = get_styles()
    
    # Row 2: 제목
    ws.cell(row=2, column=2).value = "LCT BUSHRA — RORO Calculator"
    ws.cell(row=2, column=2).font = styles["title_font"]
    
    # Row 3: 헤더 (원본과 동일: SECTION | PARAMETER | UNIT | VALUE | NOTES)
    headers_row3 = ["", "SECTION", "PARAMETER", "UNIT", "VALUE", "NOTES"]
    for col_idx, header in enumerate(headers_row3, start=1):
        if header:  # 빈 문자열은 건너뛰기
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"]
            )
    
    # INPUT CONSTANTS 섹션 (Row 5-8)
    ws.cell(row=5, column=2).value = "INPUT CONSTANTS"
    ws.cell(row=5, column=3).value = "L_ramp_m"
    ws.cell(row=5, column=4).value = "m"
    ws.cell(row=5, column=5).value = 12.0
    ws.cell(row=5, column=5).fill = styles["input_fill"]
    ws.cell(row=5, column=6).value = "Linkspan length. Calc!D4"
    
    ws.cell(row=6, column=3).value = "theta_max_deg"
    ws.cell(row=6, column=4).value = "deg"
    ws.cell(row=6, column=5).value = 6.0
    ws.cell(row=6, column=5).fill = styles["input_fill"]
    ws.cell(row=6, column=6).value = "Max ramp angle. Calc!D5"
    
    ws.cell(row=7, column=3).value = "KminusZ_m"
    ws.cell(row=7, column=4).value = "m"
    ws.cell(row=7, column=5).value = 3.0
    ws.cell(row=7, column=5).fill = styles["input_fill"]
    ws.cell(row=7, column=6).value = "K - Z (UPDATE!). Calc!D6"
    
    ws.cell(row=8, column=3).value = "D_vessel_m"
    ws.cell(row=8, column=4).value = "m"
    ws.cell(row=8, column=5).value = 3.65  # Stability Booklet: Moulded Depth (기존: 4.85)
    ws.cell(row=8, column=5).fill = styles["input_fill"]
    ws.cell(row=8, column=6).value = "Molded depth. Calc!D7"
    
    # LIMITS & OPS 섹션 (Row 10-12)
    ws.cell(row=10, column=2).value = "LIMITS & OPS"
    ws.cell(row=10, column=3).value = "min_fwd_draft_m"
    ws.cell(row=10, column=4).value = "m"
    ws.cell(row=10, column=5).value = 1.5
    ws.cell(row=10, column=5).fill = styles["input_fill"]
    ws.cell(row=10, column=6).value = "Min draft. Calc!D9"
    
    ws.cell(row=11, column=3).value = "max_fwd_draft_m"
    ws.cell(row=11, column=4).value = "m"
    ws.cell(row=11, column=5).value = 3.5
    ws.cell(row=11, column=5).fill = styles["input_fill"]
    ws.cell(row=11, column=6).value = "Max draft. Calc!D10"
    
    ws.cell(row=12, column=3).value = "pump_rate_tph"
    ws.cell(row=12, column=4).value = "t/h"
    ws.cell(row=12, column=5).value = 10.0  # 원본 파일 값
    ws.cell(row=12, column=5).fill = styles["input_fill"]
    ws.cell(row=12, column=6).value = "Pump rate. Calc!D11"
    
    # STABILITY 섹션 (Row 14-17) - Stability Booklet 기준 값으로 업데이트
    ws.cell(row=14, column=2).value = "STABILITY"
    ws.cell(row=14, column=3).value = "MTC_t_m_per_cm"
    ws.cell(row=14, column=4).value = "t·m/cm"
    ws.cell(row=14, column=5).value = 33.99  # Stability Booklet @ Δ=1183.85t (기존: 41.47, 이전: 33.95)
    ws.cell(row=14, column=5).fill = styles["input_fill"]
    ws.cell(row=14, column=6).value = "MTC. Calc!D13"
    
    ws.cell(row=15, column=3).value = "LCF_m_from_midship"
    ws.cell(row=15, column=4).value = "m"
    ws.cell(row=15, column=5).value = 30.91  # 원본 파일과 동일 (기존: 30.91)
    ws.cell(row=15, column=5).fill = styles["input_fill"]
    ws.cell(row=15, column=6).value = "LCF. Calc!D14"
    
    ws.cell(row=16, column=3).value = "TPC_t_per_cm"
    ws.cell(row=16, column=4).value = "t/cm"
    ws.cell(row=16, column=5).value = 7.95  # Stability Booklet @ Δ=1183.85t (기존: 7.5)
    ws.cell(row=16, column=5).fill = styles["input_fill"]
    ws.cell(row=16, column=6).value = "TPC. Calc!D15"
    
    ws.cell(row=17, column=3).value = "Lpp_m"
    ws.cell(row=17, column=4).value = "m"
    ws.cell(row=17, column=5).value = 60.302  # Stability Booklet: Length Between Perpendiculars (기존: 60.302, 이전 잘못: 64.00)
    ws.cell(row=17, column=5).fill = styles["input_fill"]
    ws.cell(row=17, column=6).value = "Length between perpendiculars. Calc!D16"
    
    ws.cell(row=18, column=2).value = "OPERATIONS"
    ws.cell(row=18, column=3).value = "max_fwd_draft_ops_m"
    ws.cell(row=18, column=4).value = "m"
    ws.cell(row=18, column=5).value = 2.70
    ws.cell(row=18, column=5).fill = styles["input_fill"]
    ws.cell(row=18, column=6).value = "Max forward draft for operations. Calc!D9"
    
    ws.cell(row=19, column=3).value = "ramp_door_offset_m"
    ws.cell(row=19, column=4).value = "m"
    ws.cell(row=19, column=5).value = 0.15
    ws.cell(row=19, column=5).fill = styles["input_fill"]
    ws.cell(row=19, column=6).value = "Ramp door offset. Calc!D11"
    
    ws.cell(row=20, column=3).value = "linkspan_freeboard_target_m"
    ws.cell(row=20, column=4).value = "m"
    ws.cell(row=20, column=5).value = 0.28
    ws.cell(row=20, column=5).fill = styles["input_fill"]
    ws.cell(row=20, column=6).value = "Linkspan freeboard target. Calc!D12"
    
    ws.cell(row=21, column=3).value = "gm_target_m"
    ws.cell(row=21, column=4).value = "m"
    ws.cell(row=21, column=5).value = 1.50
    ws.cell(row=21, column=5).fill = styles["input_fill"]
    ws.cell(row=21, column=6).value = "GM target. Calc!D13"
    
    # 폰트 적용 (확장된 범위)
    for row in range(5, 22):
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = styles["normal_font"]

    
    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    
    print("  [OK] Calc sheet created")


# ============================================================================
# Tide Sheet Creation
# ============================================================================

def create_tide_sheet(wb):
    """December_Tide_2025 시트 생성 - 조수 데이터 (수식 없음)"""
    ws = wb.create_sheet("December_Tide_2025")
    styles = get_styles()
    
    # 헤더 설정 (원본과 정확히 일치)
    headers = ["datetime_gst", "tide_m               (Chart Datum)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"]
        )
    
    # JSON 파일에서 조수 데이터 로드 및 삽입
    json_path = os.path.join(os.path.dirname(__file__), "../../data/gateab_v3_tide_data.json")
    json_path = os.path.normpath(json_path)
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            tide_data = json.load(f)
        
        # 744개 행에 데이터 삽입 (12월 1일 00:00 ~ 12월 31일 23:00)
        for idx, entry in enumerate(tide_data, start=0):
            row = 2 + idx  # Excel row 2부터 시작
            if row > 745:  # 744개 행만 처리
                break
            
            # Column A: datetime (문자열로 저장)
            cell_a = ws.cell(row=row, column=1)
            cell_a.value = entry.get("datetime", "")
            cell_a.font = styles["normal_font"]
            
            # Column B: tide_m (숫자로 저장)
            cell_b = ws.cell(row=row, column=2)
            cell_b.value = entry.get("tide_m", 0.0)
            cell_b.font = styles["normal_font"]
            cell_b.number_format = "0.00"
        
        print(f"  [OK] December_Tide_2025 sheet created with {len(tide_data)} rows of tide data")
        
    except FileNotFoundError:
        # JSON 파일이 없으면 빈 행만 생성 (graceful degradation)
        print(f"  [WARNING] JSON file not found: {json_path}")
        print("  [INFO] Creating empty December_Tide_2025 sheet (user can input data manually)")
        for row in range(2, 746):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_a.font = styles["normal_font"]
            cell_b.font = styles["normal_font"]
            cell_b.number_format = "0.00"
        print("  [OK] December_Tide_2025 sheet created (empty)")
        
    except (json.JSONDecodeError, KeyError, ValueError) as e:
        # JSON 파싱 오류 또는 데이터 형식 오류
        print(f"  [WARNING] Error loading JSON data: {e}")
        print("  [INFO] Creating empty December_Tide_2025 sheet (user can input data manually)")
        for row in range(2, 746):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_a.font = styles["normal_font"]
            cell_b.font = styles["normal_font"]
            cell_b.number_format = "0.00"
        print("  [OK] December_Tide_2025 sheet created (empty)")
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


# ============================================================================
# Hourly Sheet Creation
# ============================================================================

def create_hourly_sheet(wb):
    """Hourly_FWD_AFT_Heights 시트 생성 - 시간별 계산 (7,440개 수식)"""
    ws = wb.create_sheet("Hourly_FWD_AFT_Heights")
    styles = get_styles()
    
    # 헤더 설정 (원본과 정확히 일치 - 14열)
    headers = [
        "DateTime (GST)", "Tide_m", "Dfwd_req_m (even)", "Trim_m (optional)",
        "Dfwd_adj_m", "Daft_adj_m", "Ramp_Angle_deg", 
        "Status", "FWD_Height_m", "AFT_Height_m", "Notes",
        "", "Trim_m (optional)", ""
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header if header else ""  # 빈 헤더도 빈 문자열로 설정
        if header:  # 빈 헤더가 아닌 경우에만 스타일 설정
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"]
            )
    
    # 744개 행에 수식 생성 (행 2부터 745까지)
    for row in range(2, 746):
        row_str = str(row)
        
        # A열: DateTime from December_Tide_2025
        ws.cell(row=row, column=1).value = f'=IF(December_Tide_2025!A{row_str}="","",December_Tide_2025!A{row_str})'
        
        # B열: Tide from December_Tide_2025
        ws.cell(row=row, column=2).value = f'=IF(December_Tide_2025!B{row_str}="","",December_Tide_2025!B{row_str})'
        
        # C열: Mean Draft 계산
        # =IF($A2="","", INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) + $B2 - INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)) * TAN(RADIANS(INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0)))))
        ws.cell(row=row, column=3).value = (
            f'=IF($A{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) + $B{row_str} - '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)) * '
            f'TAN(RADIANS(INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0)))))'
        )
        
        # D열: Trim (입력 셀, 수식 없음)
        # 사용자가 입력
        
        # E열: FWD Draft
        # =IF($C2="","", IF($D2="", $C2, $C2 - $D2/2))
        ws.cell(row=row, column=5).value = f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} - $D{row_str}/2))'
        
        # F열: AFT Draft
        # =IF($C2="","", IF($D2="", $C2, $C2 + $D2/2))
        ws.cell(row=row, column=6).value = f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} + $D{row_str}/2))'
        
        # G열: Ramp Angle
        # =IF($E2="","", DEGREES(ATAN((INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) - $E2 + $B2) / INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)))))
        ws.cell(row=row, column=7).value = (
            f'=IF($E{row_str}="","", '
            f'DEGREES(ATAN((INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str}) / '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)))))'
        )
        
        # H열: Status
        # =IF($E2="","", IF(AND($E2>=INDEX(Calc!$E:$E, MATCH("min_fwd_draft_m", Calc!$C:$C, 0)), $E2<=INDEX(Calc!$E:$E, MATCH("max_fwd_draft_m", Calc!$C:$C, 0)), $G2<=INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0))), "OK", "CHECK"))
        ws.cell(row=row, column=8).value = (
            f'=IF($E{row_str}="","", '
            f'IF(AND($E{row_str}>=INDEX(Calc!$E:$E, MATCH("min_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$E{row_str}<=INDEX(Calc!$E:$E, MATCH("max_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$G{row_str}<=INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0))), "OK", "CHECK"))'
        )
        
        # I열: FWD Height
        # =IF($E2="","", INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $E2 + $B2)
        ws.cell(row=row, column=9).value = (
            f'=IF($E{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str})'
        )
        
        # J열: AFT Height
        # =IF($F2="","", INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $F2 + $B2)
        ws.cell(row=row, column=10).value = (
            f'=IF($F{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $F{row_str} + $B{row_str})'
        )
        
        # K열: Even Keel
        # =IF(D2=0, "Even Keel", "")
        ws.cell(row=row, column=11).value = f'=IF(D{row_str}=0, "Even Keel", "")'
        
        # L열: 빈 컬럼 (헤더 없음)
        
        # M열: Trim_m (optional) - 중복 헤더 (D열과 동일)
        # 사용자가 입력할 수 있는 셀 (수식 없음)
        
        # N열: 설명 텍스트 (행 2에만)
        if row == 2:
            ws.cell(row=row, column=14).value = "← Defaults to 0.00 (Even-Keel). To apply the actual trim, manually enter the value in this cell."
            ws.cell(row=row, column=14).font = styles["normal_font"]
        
        # 숫자 포맷 적용
        ws.cell(row=row, column=2).number_format = "0.00"  # Tide
        ws.cell(row=row, column=3).number_format = "0.00"  # Mean Draft
        ws.cell(row=row, column=4).number_format = "0.00"  # Trim
        ws.cell(row=row, column=5).number_format = "0.00"  # FWD Draft
        ws.cell(row=row, column=6).number_format = "0.00"  # AFT Draft
        ws.cell(row=row, column=7).number_format = "0.00"  # Ramp Angle
        ws.cell(row=row, column=9).number_format = "0.00"  # FWD Height
        ws.cell(row=row, column=10).number_format = "0.00"  # AFT Height
        
        # 폰트 적용
        for col in range(1, 12):
            ws.cell(row=row, column=col).font = styles["normal_font"]
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 80  # 설명 텍스트용 넓은 컬럼
    
    print(f"  [OK] Hourly_FWD_AFT_Heights sheet created with {744} rows of formulas")


# ============================================================================
# RORO Sheet Creation
# ============================================================================

def create_roro_sheet(wb: Workbook) -> None:
    """
    Create the RORO_Stage_Scenarios sheet.

    This version aligns the Excel formulas with the lever-arm ballast model
    discussed in the AGI RORO trim notes.

    Key relationships (Excel side):

    - TM (t·m)           = W_stage_t * (x_stage_m - LCF)
    - Trim_cm (cm)       = TM / MTC
    - Trim_m  (m)        = Trim_cm / 100
    - ΔTM_cm_tm (t·m)    = MTC * (ABS(TRIM5_CM) - ABS(Trim_target_cm))
    - Lever_arm_m (m)    = X_Ballast - LCF
    - Ballast_t_calc (t) = ΔTM_cm_tm / Lever_arm_m
    - Ballast_time_h_calc (h) = Ballast_t_calc / PumpRate

    It also keeps the original "rule-of-thumb" Ballast_t and Ballast_time_h
    columns based on Trim_m and TPC for comparison.
    """
    ws = wb.create_sheet("RORO_Stage_Scenarios")

    styles = get_styles()

    # Title
    ws["A1"] = "RORO Stage Scenarios – LCT BUSHRA / AGI TR"
    ws["A1"].font = styles["title_font"]

    # BASELINE INPUTS (Row 5) - 원본 파일과 동일
    ws["B5"] = "Tmean_baseline (m)"
    ws["D5"] = 2.33  # 기본값 (입력 셀)
    ws["D5"].fill = styles["input_fill"]
    ws["F5"] = "Tide_ref (m)"
    ws["G5"] = 2.0  # 기본값 (입력 셀)
    ws["G5"].fill = styles["input_fill"]

    # Hydrostatic references from Calc sheet
    ws["B8"] = "MTC (t·m/cm):"
    ws["C8"] = '=INDEX(Calc!$E:$E, MATCH("MTC_t_m_per_cm", Calc!$C:$C, 0))'
    ws["B9"] = "LCF (m):"
    ws["C9"] = '=INDEX(Calc!$E:$E, MATCH("LCF_m_from_midship", Calc!$C:$C, 0))'

    ws["E8"] = "Lpp (m):"
    ws["F8"] = '=INDEX(Calc!$E:$E, MATCH("Lpp_m", Calc!$C:$C, 0))'
    ws["E9"] = "D_vessel (m):"
    ws["F9"] = 3.65  # 원본 파일과 동일한 직접 값

    ws["B10"] = "TPC (t/cm):"
    ws["C10"] = '=INDEX(Calc!$E:$E, MATCH("TPC_t_per_cm", Calc!$C:$C, 0))'

    ws["B11"] = "Pump rate (t/h):"
    ws["C11"] = 5.0
    ws["B12"] = "X_Ballast (m):"
    ws["C12"] = 52.53  # 원본 파일과 동일 (기존: 52.5)

    # Stage table header
    header_row = 14
    stage_headers = [
        "Stage",
        "W_stage_t",
        "x_stage_m",
        "TM (t·m)",
        "Trim_cm",
        "Trim_m",
        "Trim_target_cm",
        "ΔTM_cm_tm",
        "Lever_arm_m",
        "Ballast_t_calc",
        "Ballast_time_h_calc",
        "Ballast_t",
        "Ballast_time_h",
        "Trim_Check",
        "Dfwd_m",
        "Daft_m",
        "FWD_Height_m",
        "AFT_Height_m",
        "Notes",
        "",  # T열: 빈 헤더 (원본과 동일)
    ]
    for col, header in enumerate(stage_headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header if header else ""  # 빈 헤더도 빈 문자열로 설정
        if header:  # 빈 헤더가 아닌 경우에만 스타일 설정
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]

    # Stage names (rows 15–24)
    stages = [
        "Stage 1",
        "Stage 2",
        "Stage 3",
        "Stage 4",
        "Stage 5",
        "Stage 5A-1 (At-Limit)",
        "Stage 5A-2 (Optimized)",
        "Stage 5A-3 (Max-Safety)",
        "Stage 6",
        "Stage 7",
    ]
    target_trim_by_stage = {
        "Stage 1": 0.0,
        "Stage 2": -64.12225705329153,
        "Stage 3": -100.15915119363395,  # 정확한 값 유지
        "Stage 4": -181.8885941644562,
        "Stage 5": -163.67880395466602,  # 정확한 값 유지
        "Stage 5A-1 (At-Limit)": -121.0,
        "Stage 5A-2 (Optimized)": -96.5,
        "Stage 5A-3 (Max-Safety)": -84.34,
        "Stage 6": -96.5,
        "Stage 7": -96.5,  # Same as 5A-2 (Optimized) - Recommended for operations
    }
    
    # Stage Notes descriptions
    stage_notes = {
        "Stage 1": "Empty condition (Baseline) – Lightship + constant consumables only. Establishes reference drafts and trim.",
        "Stage 2": "SPMT 1st entry on ramp (~30% reaction) – TR1 begins ramp entry; approx. 30% of unit load transferred to vessel.",
        "Stage 3": "~50% on ramp (Half reaction) – TR1 midpoint on ramp; bow/stern reaction partially balanced.",
        "Stage 4": "Full on ramp / Break-even (1 unit full weight) – TR1 fully on ramp; full transformer weight acting on vessel through ramp system.",
        "Stage 5": "Deck full load (217t × 2) – Combined CG. TR1 & TR2 fully stowed at final deck locations with no ballast correction. Represents initial full-load condition (Trim = excessive).",
        "Stage 5A-1 (At-Limit)": "TR1 & TR2 at final deck positions. Ballast adjusted only to achieve trim at the maximum allowable limit. Used for boundary-case evaluation.",
        "Stage 5A-2 (Optimized)": "Recommended trim optimization case (legacy). Aft ballast ≈ 146t @ x≈50m used to minimize trim and increase freeboard.",
        "Stage 5A-3 (Max-Safety)": "Conservative variation of 5A-2, applying additional ballast margin for enhanced safety. Used for sensitivity and worst-case trimming review.",
        "Stage 6": "TR1 @ final stowage + TR2 full on bow ramp (No ballast adjustment) – Critical intermediate condition. Represents maximum forward loading moment. Used to verify: worst-case forward draft, linkspan connector deck-wetting risk, actual trim without correction.",
        "Stage 7": "TR1 @ final stowage + TR2 full on bow ramp + AFT ballast (Within limits) – Operational correction stage. TR1 final + TR2 full on ramp; apply aft ballast (≈ Option A-2 level) to achieve: Trim within allowable limits, acceptable forward draft margin, minimum deck-wetting risk. Final recommended configuration for safe operation during TR2 transition.",
    }

    first_data_row = header_row + 1
    trim5_row = None

    # Stage별 기본 W, x 값 (Stage 6, 7 계산값 포함)
    # Stage 6 계산: TR1 최종 위치 + TR2 램프 위 (Stage 4와 동일한 좌표계 사용)
    # - TR1: W₁ = 217t, x₁ = +15.27 m (Stage 5의 TR1 최종 위치)
    # - TR2: W₂ = 217t, x₂ = -3.85 m (Stage 4의 ramp 위치, 기존 좌표계와 일치)
    # - x_stage_6 = (217×15.27 + 217×(-3.85)) / 434 = 5.71 m
    # Stage 7 계산: Stage 6 화물 + AFT Ballast (하지만 W_stage_t는 화물만)
    # - 화물: W_cargo = 434 t, x_cargo = 5.71 m (Stage 6와 동일)
    # - AFT Ballast: W_ballast = 146 t, x_ballast = 50.0 m (Stage 5A-2 기준)
    # - W_stage_7 = 434 t (화물만, Ballast 포함 안 함)
    # - x_stage_7 = (434×5.71 + 146×50.0) / (434+146) = 13.06 m (합성 중심, 하지만 W는 화물만)
    stage_defaults = {
        "Stage 1": {"W": 0.0, "x": 0.0},
        "Stage 2": {"W": 65.0, "x": -10.0},  # 원본 파일과 동일
        "Stage 3": {"W": 110.0, "x": -6.85},  # 원본 파일과 동일
        "Stage 4": {"W": 217.0, "x": -3.85},  # 원본 파일과 동일
        "Stage 5": {"W": 434.0, "x": 15.27},  # 원본 파일과 동일
        "Stage 5A-1 (At-Limit)": {"W": 527.57, "x": 24.43},  # 원본 파일과 동일
        "Stage 5A-2 (Optimized)": {"W": 579.97, "x": 24.01},  # 원본 파일과 동일
        "Stage 5A-3 (Max-Safety)": {"W": 606.17, "x": 25.14},  # 원본 파일과 동일
        "Stage 6": {"W": 434.0, "x": 0.63},  # 원본 파일과 동일
        "Stage 7": {"W": 434.0, "x": 0.63},  # 원본 파일과 동일
    }

    for idx, stage_name in enumerate(stages, start=0):
        row = first_data_row + idx
        row_str = str(row)

        # Column A – Stage label
        ws.cell(row=row, column=1, value=stage_name)

        # Columns B (W_stage_t) and C (x_stage_m) are user inputs
        # 기본값이 있으면 입력, 없으면 빈 셀로 두기
        if stage_name in stage_defaults:
            defaults = stage_defaults[stage_name]
            ws.cell(row=row, column=2, value=defaults["W"])  # W_stage_t
            ws.cell(row=row, column=3, value=defaults["x"])  # x_stage_m
        ws.cell(row=row, column=2).fill = styles["input_fill"]  # W_stage_t input cell
        ws.cell(row=row, column=3).fill = styles["input_fill"]  # x_stage_m input cell

        # Column D – TM (t·m) = W * (x - LCF)
        ws.cell(row=row, column=4).value = (
            f'=IF(OR(B{row_str}="", C{row_str}="", $C$9=""), "", '
            f'B{row_str} * (C{row_str} - $C$9))'
        )

        # Column E – Trim_cm = TM / MTC
        ws.cell(row=row, column=5).value = (
            f'=IF(OR(D{row_str}="", OR($C$8="", $C$8=0)), "", D{row_str} / $C$8)'
        )

        # Column F – Trim_m = Trim_cm / 100
        ws.cell(row=row, column=6).value = (
            f'=IF(E{row_str}="", "", E{row_str} / 100)'
        )

        # Remember Stage 5 row for TRIM5_CM named range
        if stage_name == "Stage 5":
            trim5_row = row

        # Column G – Trim_target_cm (모든 Stage에 값 설정)
        target_trim = target_trim_by_stage.get(stage_name)
        if target_trim is not None:
            ws.cell(row=row, column=7).value = target_trim
            ws.cell(row=row, column=7).fill = styles["input_fill"]  # Input cell style

        # Column H – ΔTM_cm_tm = MTC * (|TRIM5_CM| - |Trim_target_cm|)
        ws.cell(row=row, column=8).value = (
            f'=IF(OR(ISERROR(MTC), ISERROR(TRIM5_CM), G{row_str}=""), "", '
            f'ROUND(MTC * (ABS(TRIM5_CM) - ABS(G{row_str})), 2))'
        )

        # Column I – Lever_arm_m = X_Ballast - LCF (same for all rows)
        # 직접 셀 참조 사용: $C$12 (X_Ballast), $C$9 (LCF)
        ws.cell(row=row, column=9).value = (
            f'=IF(OR(ISBLANK($C$12), ISBLANK($C$9), ISERROR($C$9)), "", '
            f'ROUND($C$12 - $C$9, 2))'
        )

        # Column J – Ballast_t_calc = ΔTM / Lever_arm
        ws.cell(row=row, column=10).value = (
            f'=IF(OR(H{row_str}="", I{row_str}=""), "", '
            f'ROUND(H{row_str} / I{row_str}, 2))'
        )

        # Column K – Ballast_time_h_calc = Ballast_t_calc / PumpRate
        # 직접 셀 참조 사용: $C$11 (PumpRate)
        ws.cell(row=row, column=11).value = (
            f'=IF(OR(J{row_str}="", $C$11="", $C$11=0, ISERROR($C$11)), "", '
            f'ROUND(J{row_str} / $C$11, 2))'
        )

        # Column L – Ballast_t (rule-of-thumb using Trim_m & TPC)
        ws.cell(row=row, column=12).value = (
            f'=IF(OR(F{row_str}="", OR($C$10="", $C$10=0)), "", '
            f'ROUND(ABS(F{row_str}) * 50 * $C$10, 2))'
        )

        # Column M – Ballast_time_h = Ballast_t / PumpRate
        # 직접 셀 참조 사용: $C$11 (PumpRate)
        ws.cell(row=row, column=13).value = (
            f'=IF(OR(L{row_str}="", $C$11="", $C$11=0, ISERROR($C$11)), "", '
            f'ROUND(L{row_str} / $C$11, 2))'
        )

        # Column N – Trim_Check based on Lpp/50 criterion
        ws.cell(row=row, column=14).value = (
            f'=IF(F{row_str}="", "", '
            f'IF(ABS(F{row_str}) <= ($F$8/50), "OK", "EXCESSIVE"))'
        )

        # Columns O/P – FWD/AFT drafts from mean draft & Trim_m
        # 원본: $D$5 - G15/2 ($D$5는 Tmean_baseline)
        # 새로운 구조: $D$5는 Tmean_baseline (Row 5, Col D)
        ws.cell(row=row, column=15).value = (
            f'=IF(OR($D$5="", F{row_str}=""), "", $D$5 - F{row_str}/2)'
        )
        ws.cell(row=row, column=16).value = (
            f'=IF(OR($D$5="", F{row_str}=""), "", $D$5 + F{row_str}/2)'
        )

        # Columns Q/R – Heights from keel using D_vessel, tide & drafts
        # 원본: $F$9 - H15 + $G$5 (H15는 Dfwd_m, $G$5는 Tide_ref)
        # 새로운 구조: O15는 Dfwd_m (Col 15), $G$5는 Tide_ref (Row 5, Col G)
        ws.cell(row=row, column=17).value = (
            f'=IF(O{row_str}="", "", $F$9 - O{row_str} + $G$5)'
        )
        ws.cell(row=row, column=18).value = (
            f'=IF(P{row_str}="", "", $F$9 - P{row_str} + $G$5)'
        )
        # Column S – Notes (with default descriptions)
        if stage_name in stage_notes:
            ws.cell(row=row, column=19).value = stage_notes[stage_name]
            ws.cell(row=row, column=19).fill = styles["input_fill"]  # Input cell style for manual editing

    # Define helpful named ranges (MTC, LCF, PumpRate, X_Ballast, TRIM5_CM)
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names['MTC'] = DefinedName('MTC', attr_text="'RORO_Stage_Scenarios'!$C$8")
    wb.defined_names['LCF'] = DefinedName('LCF', attr_text="'RORO_Stage_Scenarios'!$C$9")
    wb.defined_names['PumpRate'] = DefinedName('PumpRate', attr_text="'RORO_Stage_Scenarios'!$C$11")
    wb.defined_names['X_Ballast'] = DefinedName('X_Ballast', attr_text="'RORO_Stage_Scenarios'!$C$12")

    if trim5_row is not None:
        wb.defined_names['TRIM5_CM'] = DefinedName(
            'TRIM5_CM',
            attr_text=f"'RORO_Stage_Scenarios'!$E${trim5_row}",
        )

    # Create Excel Table for structured references (optional but recommended)
    from openpyxl.worksheet.table import Table, TableStyleInfo
    try:
        # Table range: A14:AD24 (header row + 10 data rows, 30 columns - T~AD 추가)
        table = Table(displayName="Stages", ref=f"A{header_row}:AD{first_data_row + len(stages) - 1}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        print("  [OK] Excel Table 'Stages' created")
    except Exception as e:
        print(f"  [WARNING] Could not create Excel Table: {e}")

    # Basic column widths for readability
    widths = {
        "A": 18,
        "B": 11,
        "C": 11,
        "D": 13,
        "E": 10,
        "F": 9,
        "G": 13,
        "H": 12,
        "I": 11,
        "J": 13,
        "K": 15,
        "L": 11,
        "M": 15,
        "N": 13,
        "O": 10,
        "P": 10,
        "Q": 13,
        "R": 13,
        "S": 40,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "B15"

    print("  [OK] RORO_Stage_Scenarios sheet created")


# ============================================================================
# v10 Additional Sheet Creation Functions
# ============================================================================

def extend_roro_captain_req(ws):
    """RORO 시트에 Captain Req 컬럼 추가 (Col T부터) - v10 확장"""
    styles = get_styles()
    
    captain_cols = ["GM(m)", "Fwd Draft(m)", "vs 2.70m", "De-ballast Qty(t)", "Timing", 
                    "Linkspan_Freeboard_m", "Clearance_Check", "GM_calc", "GM_Check", 
                    "Prop Imm(%)", "Vent_Time_h"]
    start_col = 20  # Col T
    
    # Row 14에 헤더 설정 (기존 헤더 행과 동일)
    for i, h in enumerate(captain_cols):
        col = start_col + i
        cell = ws.cell(row=14, column=col)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"]
        )
    
    # Row 15부터 각 Stage 행에 수식 생성 (10개 Stage)
    for row in range(15, 25):
        row_str = str(row)
        
        # Col T (20): GM(m) - VLOOKUP from Hydro_Table
        ws.cell(row=row, column=20).value = f'=IF(D{row_str}="", "", VLOOKUP(D{row_str},Hydro_Table!$A:$D,4,1))'
        ws.cell(row=row, column=20).number_format = "0.00"
        ws.cell(row=row, column=20).font = styles["normal_font"]
        
        # Col U (21): Fwd Draft(m) - Col O의 Dfwd_m 참조
        ws.cell(row=row, column=21).value = f'=IF(O{row_str}="", "", O{row_str})'
        ws.cell(row=row, column=21).number_format = "0.00"
        ws.cell(row=row, column=21).font = styles["normal_font"]
        
        # Col V (22): vs 2.70m - Col U와 2.70 비교
        ws.cell(row=row, column=22).value = f'=IF(U{row_str}="", "", IF(U{row_str}<=Calc!$E$18,"OK","NG"))'
        ws.cell(row=row, column=22).font = styles["normal_font"]
        
        # Col W (23): De-ballast Qty(t) - Col J의 Ballast_t_calc 참조
        ws.cell(row=row, column=23).value = f'=IF(J{row_str}="", "", J{row_str})'
        ws.cell(row=row, column=23).number_format = "0.00"
        ws.cell(row=row, column=23).font = styles["normal_font"]
        
        # Col X (24): Timing - 빈 셀 (사용자 입력)
        # ws.cell(row=row, column=24).value = ""  # 빈 셀
        
        # Col Y (25): Linkspan_Freeboard_m
        ws.cell(row=row, column=25).value = f'=IF(Q{row_str}="", "", Q{row_str} - Calc!$E$19)'
        ws.cell(row=row, column=25).number_format = "0.00"
        ws.cell(row=row, column=25).font = styles["normal_font"]
        
        # Col Z (26): Clearance_Check
        ws.cell(row=row, column=26).value = f'=IF(Y{row_str}="", "", IF(Y{row_str}>=Calc!$E$20,"OK","<0.28m CHECK"))'
        ws.cell(row=row, column=26).font = styles["normal_font"]
        
        # Col AA (27): GM_calc - Col T와 동일 (VLOOKUP 결과)
        ws.cell(row=row, column=27).value = f'=T{row_str}'
        ws.cell(row=row, column=27).number_format = "0.00"
        ws.cell(row=row, column=27).font = styles["normal_font"]
        
        # Col AB (28): GM_Check
        ws.cell(row=row, column=28).value = f'=IF(AA{row_str}="", "", IF(AA{row_str}>=Calc!$E$21,"OK","NG"))'
        ws.cell(row=row, column=28).font = styles["normal_font"]
        
        # Col AC (29): Prop Imm(%) - ($P{row}-2.10)/1.25*100
        ws.cell(row=row, column=29).value = f'=IF(P{row_str}="", "", ($P{row_str}-2.10)/1.25*100)'
        ws.cell(row=row, column=29).number_format = "0.00"
        ws.cell(row=row, column=29).font = styles["normal_font"]
        
        # Col AD (30): Vent_Time_h - De-ballast Qty / 45
        ws.cell(row=row, column=30).value = f'=IF(W{row_str}>0, W{row_str}/45, "-")'
        ws.cell(row=row, column=30).number_format = "0.00"
        ws.cell(row=row, column=30).font = styles["normal_font"]
    
    # 컬럼 너비 설정
    ws.column_dimensions['T'].width = 12
    ws.column_dimensions['U'].width = 12
    ws.column_dimensions['V'].width = 12
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 12
    ws.column_dimensions['Y'].width = 18
    ws.column_dimensions['Z'].width = 15
    ws.column_dimensions['AA'].width = 12
    ws.column_dimensions['AB'].width = 12
    ws.column_dimensions['AC'].width = 12
    ws.column_dimensions['AD'].width = 12
    
    print("  [OK] Captain Req columns added to RORO_Stage_Scenarios sheet (v10 extended)")


def create_ballast_tanks_sheet(wb):
    """Ballast_Tanks 시트 생성"""
    ws = wb.create_sheet("Ballast_Tanks")
    styles = get_styles()
    
    headers = ["TankName", "x_from_mid_m", "max_t", "SG", "use_flag"]
    data = [
        ["FWB1.P", -25.93, 80.00, 1.025, "Y"],
        ["FWB1.S", -25.93, 80.00, 1.025, "Y"],
        ["FWB2.P", -18.45, 110.00, 1.025, "Y"],
        ["FWB2.S", -18.45, 110.00, 1.025, "Y"],
        ["FWCARGO1.P", -3.66, 148.40, 1.000, "Y"],
        ["FWCARGO1.S", -3.66, 148.40, 1.000, "Y"],
    ]
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"]
        )
    
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = styles["normal_font"]
            if c in [3, 4]:  # 숫자 열
                cell.number_format = '0.00'
    
    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    
    print("  [OK] Ballast_Tanks sheet created")


def create_hydro_table_sheet(wb):
    """Hydro_Table 시트 생성"""
    ws = wb.create_sheet("Hydro_Table")
    styles = get_styles()
    
    headers = ["Disp_t", "Tmean_m", "Trim_m", "GM_m", "Draft_FWD", "Draft_AFT"]
    data = [
        [2991.25, 2.20, 0.20, 2.85, 2.10, 2.30],
        [3208.25, 3.18, -0.53, 1.68, 2.92, 3.45],
        [3265.25, 3.00, 0.60, 1.88, 2.68, 3.28],
        [3425.25, 3.00, 0.70, 1.85, 2.65, 3.35],
    ]
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"]
        )
    
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = styles["normal_font"]
            cell.number_format = '0.00'
    
    # 컬럼 너비 설정
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 12
    
    print("  [OK] Hydro_Table sheet created")


def create_captain_req_sheet(wb):
    """Captain_Req 시트 생성"""
    ws = wb.create_sheet("Captain_Req")
    styles = get_styles()
    
    # Row 1: 헤더
    headers = ["Stage", "GM(m)", "Fwd Draft(m)", "vs 2.70m", "De-ballast Qty(t)", 
               "Timing", "Critical Trim(m aft)", "GM", "Freeboard(m)", "vs Prev 0.46m", 
               "Prop Imm(%)", "Emergency OK", "Vent Rate(t/h)", "Time(h)", "Notes"]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"]
        )
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Row 2-5: 샘플 데이터
    data = [
        ("0", 2.85, 2.10, "OK", 0, "-", 0.20, 2.85, 1.55, "-", 100, "OK", "-", "-", "Empty baseline"),
        ("1", 1.68, 2.92, "NG", 160, "After 1st TR", -0.53, 1.68, 0.73, "-", 92, "OK", 45, 3.56, "De-ballast start"),
        ("Critical", 1.88, 2.68, "OK", 0, "-", 0.60, 1.88, 0.42, -0.04, 96, "OK", "-", "-", "TR1 stow + TR2 ramp"),
        ("Final", 1.85, 2.65, "OK", 50, "Fine tune", 0.70, 1.85, 1.00, "-", 97, "OK", "-", "-", "Optimized")
    ]
    
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = styles["normal_font"]
            if c in [2, 3, 8, 9, 11]:  # 숫자 열
                cell.number_format = '0.00'
    
    # Row 2-10: Live 수식 연동 (RORO 시트 참조)
    for r in range(2, 11):
        row_str = str(r)
        roro_row = r + 13  # RORO 시트의 Row 15-24에 매핑
        # Col B (2): GM(m) - RORO 시트 Col T 참조
        ws.cell(row=r, column=2).value = f'=RORO_Stage_Scenarios!T{roro_row}'
        # Col C (3): Fwd Draft(m) - RORO 시트 Col U 참조
        ws.cell(row=r, column=3).value = f'=RORO_Stage_Scenarios!U{roro_row}'
        # Col D (4): vs 2.70m - RORO 시트 Col V 참조
        ws.cell(row=r, column=4).value = f'=RORO_Stage_Scenarios!V{roro_row}'
        # Col E (5): De-ballast Qty(t) - RORO 시트 Col W 참조
        ws.cell(row=r, column=5).value = f'=RORO_Stage_Scenarios!W{roro_row}'
        # Col H (8): Freeboard(m) - RORO 시트 Col Y 참조
        ws.cell(row=r, column=8).value = f'=RORO_Stage_Scenarios!Y{roro_row}'
        # Col I (9): vs Prev 0.46m
        ws.cell(row=r, column=9).value = f'=IF(C{r}-0.46<0, C{r}-0.46, "-")'
        # Col K (11): Prop Imm(%) - RORO 시트 Col AC 참조
        ws.cell(row=r, column=11).value = f'=RORO_Stage_Scenarios!AC{roro_row}'
        # Col M (13): Vent Rate(t/h) - 기본값 45
        ws.cell(row=r, column=13).value = 45
        # Col N (14): Time(h) - RORO 시트 Col AD 참조
        ws.cell(row=r, column=14).value = f'=RORO_Stage_Scenarios!AD{roro_row}'
    
    print("  [OK] Captain_Req sheet created (v10 with live formulas)")


def create_captain_report_sheet(wb):
    """CAPTAIN_REPORT 시트 생성 - 캡틴/Harbour Master용 요약"""
    ws = wb.create_sheet("CAPTAIN_REPORT")
    styles = get_styles()

    # 1. 제목
    ws["A1"] = "LCT BUSHRA – Captain Summary (Draft / Trim / Freeboard)"
    ws.merge_cells("A1:I1")
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["center_align"]

    # 2. LIMIT / REF 영역 헤더 (Row 3)
    header_row = 3
    param_headers = ["Parameter", "Value", "Unit", "Remark"]
    for col_idx, header in enumerate(param_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 3. LIMIT / REF 값 (Row 4~7)
    # Row 4: Summer draft limit
    ws["A4"] = "Summer draft limit (max draft)"
    ws["B4"] = 2.70  # 캡틴 요구: 모든 stage에서 2.70m 초과 금지
    ws["C4"] = "m"
    ws["D4"] = "As per summer draft, any mark ≤ 2.70 m"
    ws["B4"].fill = styles["input_fill"]

    # Row 5: Linkspan freeboard limit
    ws["A5"] = "Linkspan freeboard limit"
    ws["B5"] = 0.28  # 캡틴 요구: linkspan freeboard 0.28m 이상
    ws["C5"] = "m"
    ws["D5"] = "Minimum freeboard at ramp connector"
    ws["B5"].fill = styles["input_fill"]

    # Row 6: Tmean_baseline ref (from RORO_Stage_Scenarios!D5)
    ws["A6"] = "Tmean_baseline (ref)"
    ws["B6"] = '=RORO_Stage_Scenarios!$D$5'
    ws["C6"] = "m"
    ws["D6"] = "Baseline mean draft used in RORO stages"

    # Row 7: Tide_ref ref (from RORO_Stage_Scenarios!G5)
    ws["A7"] = "Tide_ref (ref)"
    ws["B7"] = '=RORO_Stage_Scenarios!$G$5'
    ws["C7"] = "m"
    ws["D7"] = "Reference tide for RORO stages"

    # LIMIT / REF 영역 기본 폰트 적용
    for row in range(4, 8):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = styles["normal_font"]
                if col == 2 and row in (4, 5):
                    # 이미 input_fill 적용됨
                    cell.number_format = "0.00"

    # 4. Stage 요약 테이블 헤더 (Row 9)
    summary_header_row = 9
    summary_headers = [
        "Stage",
        "Dfwd_m",
        "Daft_m",
        "Trim_m",
        "Max_draft_m",
        "Draft_OK",
        "FWD_Height_m",
        "Freeboard_OK",
        "Notes",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_header_row, column=col_idx)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 5. RORO_Stage_Scenarios의 Stage별 row 매핑
    #   Stage 1  → 15
    #   Stage 2  → 16
    #   Stage 3  → 17
    #   Stage 4  → 18
    #   Stage 5  → 19
    #   Stage 5A-1 → 20
    #   Stage 5A-2 → 21
    #   Stage 5A-3 → 22
    #   Stage 6  → 23
    #   Stage 7  → 24
    roro_rows = [15, 16, 17, 18, 19, 20, 21, 22, 23, 24]
    data_start_row = summary_header_row + 1  # 10행부터

    for idx, roro_row in enumerate(roro_rows):
        row = data_start_row + idx
        row_str = str(row)
        roro_row_str = str(roro_row)

        # A: Stage 이름
        ws.cell(row=row, column=1).value = f'=RORO_Stage_Scenarios!A{roro_row_str}'
        # B: Dfwd_m (O열)
        ws.cell(row=row, column=2).value = f'=RORO_Stage_Scenarios!O{roro_row_str}'
        # C: Daft_m (P열)
        ws.cell(row=row, column=3).value = f'=RORO_Stage_Scenarios!P{roro_row_str}'
        # D: Trim_m (F열)
        ws.cell(row=row, column=4).value = f'=RORO_Stage_Scenarios!F{roro_row_str}'
        # E: Max_draft_m = MAX(B,C)
        ws.cell(row=row, column=5).value = f'=IF(OR(B{row_str}="",C{row_str}=""),"",MAX(B{row_str},C{row_str}))'
        # F: Draft_OK = E ≤ Summer draft limit?
        ws.cell(row=row, column=6).value = (
            f'=IF($B$4="","",IF(E{row_str}<=$B$4,"OK",">2.70m"))'
        )
        # G: FWD_Height_m = RORO Q열
        ws.cell(row=row, column=7).value = f'=RORO_Stage_Scenarios!Q{roro_row_str}'
        # H: Freeboard_OK = G ≥ 0.28?
        ws.cell(row=row, column=8).value = (
            f'=IF($B$5="","",IF(G{row_str}>=$B$5,"OK","<0.28m"))'
        )
        # I: Notes = RORO S열
        ws.cell(row=row, column=9).value = f'=RORO_Stage_Scenarios!S{roro_row_str}'

        # 숫자 포맷/스타일
        for col in range(2, 8):  # B~G
            cell = ws.cell(row=row, column=col)
            cell.font = styles["normal_font"]
            cell.number_format = "0.00"

        # Draft_OK / Freeboard_OK / Notes 스타일
        for col in (1, 6, 8, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = styles["normal_font"]
            cell.alignment = styles["left_align"]

    # 6. 컬럼 너비 조정
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 45

    # 7. 보기 좋게 그리드 보더 (요약 테이블 구간만)
    last_row = data_start_row + len(roro_rows) - 1
    for row in range(summary_header_row, last_row + 1):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # 8. Panes 고정 (Stage 헤더 기준)
    ws.freeze_panes = "A10"
    
    print("  [OK] CAPTAIN_REPORT sheet created")


# ============================================================================
# Main Orchestration Function
# ============================================================================

def create_workbook_from_scratch():
    """워크북을 처음부터 생성"""
    print("=" * 80)
    print("LCT_BUSHRA_AGI_TR.xlsx Creation from Scratch")
    print("=" * 80)
    
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(OUTPUT_FILE)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[OK] Created output directory: {output_dir}")
    
    # 파일이 열려있으면 타임스탬프 추가
    final_output_file = OUTPUT_FILE
    if os.path.exists(OUTPUT_FILE):
        try:
            # 파일이 쓰기 가능한지 확인
            with open(OUTPUT_FILE, 'r+b'):
                pass
        except PermissionError:
            # 파일이 열려있으면 새 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(OUTPUT_FILE)[0]
            final_output_file = f"{base_name}_{timestamp}.xlsx"
            print(f"[WARNING] Original file is open. Saving as: {final_output_file}")
    
    # 새 워크북 생성
    print(f"\n[1/7] Creating new workbook")
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 각 시트 생성
    print(f"\n[2/8] Creating sheets:")
    create_calc_sheet(wb)
    create_tide_sheet(wb)
    create_hourly_sheet(wb)
    create_roro_sheet(wb)
    # Ballast_Tanks 시트 생성
    create_ballast_tanks_sheet(wb)
    # Hydro_Table 시트 생성
    create_hydro_table_sheet(wb)
    # RORO 시트에 Captain Req 컬럼 추가
    roro_ws = wb["RORO_Stage_Scenarios"]
    extend_roro_captain_req(roro_ws)
    # Captain_Req 시트 생성
    create_captain_req_sheet(wb)
    # CAPTAIN_REPORT 시트 생성
    create_captain_report_sheet(wb)
    
    # 저장
    print(f"\n[7/8] Saving workbook: {final_output_file}")
    try:
        wb.save(final_output_file)
        print(f"  [OK] File saved successfully")
    except Exception as e:
        print(f"  [ERROR] Failed to save: {e}")
        sys.exit(1)
    
    wb.close()
    
    # 파일 크기 확인
    print(f"\n[8/8] Verification:")
    if os.path.exists(final_output_file):
        file_size = os.path.getsize(final_output_file) / 1024  # KB
        print(f"  [OK] File created: {final_output_file}")
        print(f"  [OK] File size: {file_size:.2f} KB")
        print(f"  [OK] Sheets: {len(wb.sheetnames)}")
    else:
        print(f"  [ERROR] Output file was not created")
        sys.exit(1)
    
    print("\n" + "=" * 80)
    print("[SUCCESS] Workbook creation from scratch complete!")
    print("=" * 80)

if __name__ == "__main__":
    create_workbook_from_scratch()

