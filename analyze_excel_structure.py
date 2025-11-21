# -*- coding: utf-8 -*-
"""
통합 엑셀 구조 분석 도구

기능:
- 엑셀 파일 구조 분석
- 시트별 상세 정보
- 헤더 위치 확인
- 수식 참조 분석
- 데이터 범위 확인

사용법:
    python analyze_excel_structure.py [--file PATH] [--sheet SHEET_NAME] [--formulas] [--headers]
"""

import argparse
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def analyze_workbook_structure(wb, file_path):
    """워크북 전체 구조 분석"""
    print("=" * 80)
    print(f"엑셀 파일 구조 분석: {file_path.name}")
    print("=" * 80)
    print(f"\n총 시트 수: {len(wb.sheetnames)}")
    print(f"시트 목록: {', '.join(wb.sheetnames)}")
    print()


def analyze_sheet_columns(ws, sheet_name, detailed=False):
    """시트별 컬럼 구조 분석"""
    print("-" * 80)
    print(f"시트: {sheet_name}")
    print("-" * 80)
    print(f"  행 수: {ws.max_row}")
    print(f"  열 수: {ws.max_column}")
    
    # 데이터 범위
    if ws.max_row > 0 and ws.max_column > 0:
        print(f"  데이터 범위: A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    
    # 수식 개수
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
    
    print(f"  수식이 있는 셀: {formula_count}개")
    
    if detailed:
        # 각 컬럼별 데이터 타입 분석
        print(f"\n  컬럼별 정보:")
        for col_idx in range(1, min(ws.max_column + 1, 21)):  # 최대 20열까지
            col_letter = get_column_letter(col_idx)
            has_data = False
            has_formula = False
            sample_value = None
            
            for row_idx in range(1, min(ws.max_row + 1, 100)):  # 최대 100행까지 확인
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    has_data = True
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formula = True
                    if sample_value is None:
                        sample_value = cell.value
                    break
            
            if has_data:
                type_info = "수식" if has_formula else "데이터"
                if sample_value and isinstance(sample_value, str) and len(sample_value) > 30:
                    sample_value = str(sample_value)[:30] + "..."
                print(f"    {col_letter}: {type_info} (샘플: {sample_value})")
    
    print()


def analyze_headers(ws, sheet_name, header_row=1):
    """헤더 분석"""
    print(f"\n[{sheet_name} 헤더 분석 (행 {header_row})]")
    print("-" * 80)
    
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if cell.value:
            headers.append((col_letter, cell.value))
            print(f"  {col_letter}{header_row}: {cell.value}")
        else:
            # 빈 헤더도 표시
            print(f"  {col_letter}{header_row}: (빈 헤더)")
    
    return headers


def analyze_formula_references(ws, sheet_name, max_rows=10):
    """수식 참조 분석"""
    print(f"\n[{sheet_name} 수식 참조 분석]")
    print("-" * 80)
    
    formula_refs = {}
    
    # 각 셀의 수식에서 참조 추출
    for row_idx in range(1, min(ws.max_row + 1, max_rows + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 21)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = str(cell.value)
                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_idx}"
                
                # 절대 참조 추출 ($C$8, $C$9 등)
                abs_refs = re.findall(r'\$[A-Z]+\$\d+', formula)
                # 시트 참조 추출 (Calc!$E:$E 등)
                sheet_refs = re.findall(r'([A-Za-z_][A-Za-z0-9_]*!)?\$?[A-Z]+\$?\d*:?\$?[A-Z]+\$?\d*', formula)
                
                if abs_refs or sheet_refs:
                    formula_refs[cell_ref] = {
                        'formula': formula[:100] + "..." if len(formula) > 100 else formula,
                        'abs_refs': abs_refs,
                        'sheet_refs': [r for r in sheet_refs if r]  # 빈 문자열 제거
                    }
    
    if formula_refs:
        print(f"수식이 있는 셀: {len(formula_refs)}개")
        for cell_ref, refs in list(formula_refs.items())[:10]:  # 최대 10개만 표시
            print(f"\n  {cell_ref}:")
            print(f"    수식: {refs['formula']}")
            if refs['abs_refs']:
                print(f"    절대 참조: {', '.join(set(refs['abs_refs']))}")
            if refs['sheet_refs']:
                print(f"    시트 참조: {', '.join(set(refs['sheet_refs'][:5]))}")
    else:
        print("수식이 있는 셀이 없습니다.")


def analyze_data_ranges(ws, sheet_name):
    """데이터 범위 분석"""
    print(f"\n[{sheet_name} 데이터 범위 분석]")
    print("-" * 80)
    
    # 각 컬럼의 데이터 범위
    for col_idx in range(1, min(ws.max_column + 1, 21)):
        col_letter = get_column_letter(col_idx)
        first_row = None
        last_row = None
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                if first_row is None:
                    first_row = row_idx
                last_row = row_idx
        
        if first_row is not None:
            print(f"  {col_letter}열: 행 {first_row} ~ {last_row}")


def analyze_specific_sheet(wb, sheet_name, options):
    """특정 시트 상세 분석"""
    if sheet_name not in wb.sheetnames:
        print(f"❌ 시트를 찾을 수 없습니다: {sheet_name}")
        return
    
    ws = wb[sheet_name]
    
    analyze_sheet_columns(ws, sheet_name, detailed=options.detailed)
    
    if options.headers:
        # 시트별 헤더 행 찾기
        if sheet_name == "RORO_Stage_Scenarios":
            analyze_headers(ws, sheet_name, header_row=14)
        else:
            analyze_headers(ws, sheet_name, header_row=1)
    
    if options.formulas:
        analyze_formula_references(ws, sheet_name)
    
    if options.ranges:
        analyze_data_ranges(ws, sheet_name)


def main():
    parser = argparse.ArgumentParser(description="엑셀 파일 구조 분석")
    parser.add_argument("--file", type=str, default="LCT_BUSHRA_AGI_TR.xlsx", help="분석할 엑셀 파일 경로")
    parser.add_argument("--sheet", type=str, default=None, help="특정 시트만 분석")
    parser.add_argument("--formulas", action="store_true", help="수식 참조 분석")
    parser.add_argument("--headers", action="store_true", help="헤더 분석")
    parser.add_argument("--ranges", action="store_true", help="데이터 범위 분석")
    parser.add_argument("--detailed", action="store_true", help="상세 정보 출력")
    parser.add_argument("--all", action="store_true", help="모든 분석 실행")
    
    args = parser.parse_args()
    
    # 파일 경로 설정
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return 1
    
    # 워크북 로드
    wb = load_workbook(file_path, data_only=False)
    
    # 전체 구조 분석
    analyze_workbook_structure(wb, file_path)
    
    # all 옵션이면 모든 옵션 활성화
    if args.all:
        args.formulas = True
        args.headers = True
        args.ranges = True
        args.detailed = True
    
    # 특정 시트만 분석
    if args.sheet:
        analyze_specific_sheet(wb, args.sheet, args)
    else:
        # 모든 시트 분석
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            analyze_sheet_columns(ws, sheet_name, detailed=args.detailed)
            
            if args.headers:
                if sheet_name == "RORO_Stage_Scenarios":
                    analyze_headers(ws, sheet_name, header_row=14)
                else:
                    analyze_headers(ws, sheet_name, header_row=1)
            
            if args.formulas:
                analyze_formula_references(ws, sheet_name, max_rows=5)
            
            if args.ranges:
                analyze_data_ranges(ws, sheet_name)
            
            print()
    
    wb.close()
    
    print("=" * 80)
    print("분석 완료")
    print("=" * 80)
    
    return 0


if __name__ == "__main__":
    exit(main())

