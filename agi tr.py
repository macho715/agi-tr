# -*- coding: utf-8 -*-
# build_bushra_agi_tr_integrated.py
# LCT_BUSHRA_AGI_TR.xlsx 파일을 프로그래밍 방식으로 Excel 함수를 생성하여 만드는 스크립트
# Integrated version: Includes Structural Limits, Option 1 Ballast Fix Check, and all features
# Updated with CAPTAIN_REPORT sheet, Structural Strength columns, and Option 1 Fix Check

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
import sys
import json
from datetime import datetime
import numpy as np
from scipy.interpolate import RectBivariateSpline
from bisect import bisect_left
from typing import Dict, Tuple, Any, List
import math
from enum import Enum, auto
import logging
import shutil

# 출력 파일 경로를 스크립트 위치 기준 루트 폴더로 설정
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "LCT_BUSHRA_AGI_TR_Final_v3.xlsx")

# ============================================================================
# Helper Functions
# ============================================================================


# 스타일 정의
def get_styles():
    """공통 스타일 정의"""
    return {
        "title_font": Font(name="Calibri", size=18, bold=True),
        "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "normal_font": Font(name="Calibri", size=11),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "ok_fill": PatternFill("solid", fgColor="C6E0B4"),
        "ng_fill": PatternFill("solid", fgColor="F8CBAD"),
        "structure_fill": PatternFill(
            "solid", fgColor="C65911"
        ),  # Orange for Structure
        "opt1_fill": PatternFill(
            "solid", fgColor="7030A0"
        ),  # Purple for Option 1 (Ballast Fix)
        "thin_border": Side(border_style="thin", color="C0C0C0"),
        "center_align": Alignment(
            horizontal="center", vertical="center", wrap_text=True
        ),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


# ============================================================================
# 공통 유틸: JSON 로더 + Frame↔x 변환 + Tank Lookup
# ============================================================================


def _load_json(filename):
    """
    JSON 파일 로더 with backup strategy.
    - 우선: 스크립트 위치 기준
    - 다음: 현재 작업 디렉토리
    - 마지막: /mnt/data (Notebook 환경용)

    BACKUP: Returns None if file not found or parsing fails
    """
    base_dirs = [
        os.path.dirname(os.path.abspath(__file__)),
        os.getcwd(),
        r"/mnt/data",
    ]
    for base_dir in base_dirs:
        path = os.path.join(base_dir, filename)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    print(f"[OK] Loaded: {filename}")
                    return data
            except json.JSONDecodeError as e:
                print(f"[BACKUP] JSON parse error in {filename}: {e}")
                continue  # Try next directory
            except Exception as e:
                print(f"[BACKUP] Error reading {filename}: {e}")
                continue
    print(f"[BACKUP] {filename} not found → using fallback")
    return None


def gm_2d_bilinear(disp, trim_m):
    """
    2D bilinear interpolation for GM calculation.
    Uses Hydro_Table_2D.json with 7x7 grid (7 displacement × 7 trim values).

    BACKUP: Returns safe minimum GM (1.50m) if calculation fails.

    Args:
        disp: Displacement (tons)
        trim_m: Trim in meters

    Returns:
        GM value (meters) or 1.50m fallback if data not available
    """
    data = _load_json("data/Hydro_Table_2D.json")
    if not data or "disp" not in data or "trim" not in data or "gm_grid" not in data:
        print(f"[BACKUP] GM data unavailable → using fallback GM=1.50m")
        return 1.50  # Safe minimum GM requirement

    try:
        d = np.array(data["disp"])
        t = np.array(data["trim"])
        g = np.array(data["gm_grid"])

        # Create 2D bilinear interpolator
        interp = RectBivariateSpline(d, t, g, kx=1, ky=1)
        gm_value = float(interp(disp, trim_m)[0][0])

        # Sanity check for unrealistic GM values
        if gm_value < 0 or gm_value > 5.0:
            print(f"[BACKUP] GM={gm_value:.2f}m unrealistic → fallback GM=1.50m")
            return 1.50

        return gm_value
    except Exception as e:
        print(f"[BACKUP] GM calculation error: {e} → fallback GM=1.50m")
        return 1.50


def calc_draft_with_lcf(
    tmean_m: float, trim_cm: float, lcf_m: float, lbp_m: float
) -> tuple[float, float]:
    """
    LCF 기반 정밀 Dfwd/Daft 계산.

    Parameters
    ----------
    tmean_m : 평균 흘수 (m)
    trim_cm : Trim (cm)  # +면 선미침(AFT deeper), -면 선수침(FWD deeper)
    lcf_m   : LCF (m), F.P. 기준 길이
    lbp_m   : LBP (m)

    Returns
    -------
    (Dfwd_m, Daft_m) : 선수/선미 흘수 (m)
    """
    trim_m = trim_cm / 100.0

    if lbp_m <= 0:
        raise ValueError("LBP must be > 0")

    r = lcf_m / lbp_m  # 무차원 비율 (0~1 근처)

    dfwd_m = tmean_m - trim_m * (1.0 - r)
    daft_m = tmean_m + trim_m * r

    return dfwd_m, daft_m


GMGrid = Dict[float, Dict[float, float]]


def _nearest_two(sorted_vals: list[float], target: float) -> tuple[float, float]:
    """
    정렬된 축(sorted_vals)에서 target을 끼우는 양쪽 점 2개를 반환.
    경계에서는 (첫2개) 또는 (마지막2개).
    """
    if not sorted_vals:
        raise ValueError("empty axis")

    n = len(sorted_vals)
    if n == 1:
        return sorted_vals[0], sorted_vals[0]

    pos = bisect_left(sorted_vals, target)

    if pos <= 0:
        return sorted_vals[0], sorted_vals[1]
    if pos >= n:
        return sorted_vals[-2], sorted_vals[-1]

    return sorted_vals[pos - 1], sorted_vals[pos]


def get_gm_bilinear(disp_t: float, trim_m: float, gm_grid: GMGrid) -> float:
    """
    Δ(ton), Trim(m)에 대해 GM(m) 2D 보간 (bilinear).

    gm_grid 형식:
        {disp: {trim: GM, ...}, ...}
    """
    if not gm_grid:
        raise ValueError("gm_grid is empty")

    disp_axis = sorted(gm_grid.keys())
    disp1, disp2 = _nearest_two(disp_axis, disp_t)

    trim_axis = sorted(next(iter(gm_grid.values())).keys())
    trim1, trim2 = _nearest_two(trim_axis, trim_m)

    # 경계선(한 방향 값만 존재)에서는 단순 선형 보간/직접 값 사용
    def gm_at(d: float, tr: float) -> float:
        return gm_grid[d][tr]

    # 네 모서리 값
    q11 = gm_at(disp1, trim1)
    q21 = gm_at(disp2, trim1)
    q12 = gm_at(disp1, trim2)
    q22 = gm_at(disp2, trim2)

    # 축이 collapse 된 경우 처리
    if disp1 == disp2 and trim1 == trim2:
        return q11
    if disp1 == disp2:  # Trim 방향만 보간
        t = (trim_m - trim1) / (trim2 - trim1) if trim2 != trim1 else 0.0
        return q11 + t * (q12 - q11)
    if trim1 == trim2:  # Δ 방향만 보간
        t = (disp_t - disp1) / (disp2 - disp1) if disp2 != disp1 else 0.0
        return q11 + t * (q21 - q11)

    # Bilinear interpolation
    # 참조: 표준 bilinear 공식
    xd = (disp_t - disp1) / (disp2 - disp1)
    yd = (trim_m - trim1) / (trim2 - trim1)

    gm_interp = (
        q11 * (1 - xd) * (1 - yd)
        + q21 * xd * (1 - yd)
        + q12 * (1 - xd) * yd
        + q22 * xd * yd
    )
    return gm_interp


def calc_heel_from_offset(
    weight_t: float, y_offset_m: float, disp_t: float, gm_m: float
) -> float:
    """
    횡 방향 편심 하중에 의한 Heel 각도 (deg) 계산.
    Small-angle 가정: tan(φ) ≈ φ(rad) 사용.
    """
    if disp_t <= 0 or gm_m <= 0 or weight_t == 0 or y_offset_m == 0:
        return 0.0

    m_heeling = weight_t * y_offset_m  # t·m
    m_restoring = disp_t * gm_m  # t·m

    phi_rad = m_heeling / m_restoring
    heel_deg = math.degrees(phi_rad)
    return heel_deg


def calc_gm_effective(disp_t: float, gm_m: float, fse_t_m: float) -> float:
    """
    자유수면효과(FSE)를 반영한 유효 GM.
    GM_eff = GM - FSE/Δ
    """
    if disp_t <= 0:
        return gm_m

    gm_eff = gm_m - (fse_t_m / disp_t)
    return gm_eff


def heel_and_gm_check(
    weight_t: float,
    y_offset_m: float,
    disp_t: float,
    gm_m: float,
    fse_t_m: float,
    heel_limit_deg: float = 3.0,
    gm_min_m: float = 1.50,
) -> tuple[float, float, bool, bool]:
    """
    Heel 각도 + GM_eff 계산 및 체크 결과.

    Returns
    -------
    heel_deg  : 계산된 heel (deg)
    gm_eff    : FSE 반영 후 GM (m)
    heel_ok   : |heel| ≤ heel_limit_deg 여부
    gm_ok     : gm_eff ≥ gm_min_m 여부
    """
    heel_deg = calc_heel_from_offset(weight_t, y_offset_m, disp_t, gm_m)
    gm_eff = calc_gm_effective(disp_t, gm_m, fse_t_m)

    heel_ok = abs(heel_deg) <= heel_limit_deg
    gm_ok = gm_eff >= gm_min_m

    return heel_deg, gm_eff, heel_ok, gm_ok


class LoadCase(Enum):
    STATIC = auto()  # A: 정적
    DYNAMIC = auto()  # B: 동적계수만
    BRAKING = auto()  # C: 동적 + 제동/편심


def apply_dynamic_loads(
    share_load_t: float,
    pin_stress_mpa: float,
    load_case: LoadCase,
) -> tuple[float, float]:
    """
    LoadCase에 따라 동적·제동 하중 계수 적용.

    Returns
    -------
    (share_dyn_t, pin_dyn_mpa)
    """
    if load_case == LoadCase.STATIC:
        f_vert = 1.00
        f_pin = 1.00
    elif load_case == LoadCase.DYNAMIC:
        f_vert = 1.10  # 예: 동적계수 1.10
        f_pin = 1.10
    elif load_case == LoadCase.BRAKING:
        f_vert = 1.20  # 예: 동적+제동 영향
        f_pin = 1.30
    else:
        f_vert = 1.00
        f_pin = 1.00

    return share_load_t * f_vert, pin_stress_mpa * f_pin


# ============================================================================
# FRAME ↔ x_from_mid_m Mapping (BUSHRA 757 TCP aligned)
# ============================================================================

# BUSHRA Tank Plan 757 TCP 기준:
# - Frame 번호는 FWD 방향으로 증가
# - Midship ≈ Fr_mid = Lpp / 2 ≈ 30.151
# - 좌표계:  x_from_mid_m < 0.0  → FWD
#           x_from_mid_m > 0.0  → AFT
#
# ⇒ x = _FRAME_OFFSET + _FRAME_SLOPE * Fr
#    Fr_mid = 30.151  →  x = 0.0
#    큰 Frame(예: 48–65, FWB1/2 실제 위치) → x < 0.0 (FWD 쪽)
#    작은 Frame(예: 5–10)                → x > 0.0 (AFT 쪽)

_FRAME_SLOPE = -1.0  # x 는 Frame 이 커질수록 감소 (Frame 증가 = FWD)
_FRAME_OFFSET = 30.151  # Midship Frame → x = 0.0m


def _init_frame_mapping():
    """
    Frame ↔ x_from_mid_m 매핑 초기화.
    data/Frame_x_from_mid_m.json 이 있으면 거기서 SLOPE/OFFSET 자동 추정,
    없으면 757 TCP 기준 기본값을 사용한다.
    """
    global _FRAME_SLOPE, _FRAME_OFFSET
    data = _load_json("data/Frame_x_from_mid_m.json")
    if not data or not isinstance(data, list) or len(data) < 2:
        # Fallback: BUSHRA 757 TCP default (Fr 증가 = FWD, Midship = 30.151)
        _FRAME_SLOPE = -1.0
        _FRAME_OFFSET = 30.151
        print(
            f"[INFO] Frame mapping: SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f} (default)"
        )
        return

    try:
        fr1 = float(data[0]["Fr"])
        x1 = float(data[0]["x_from_mid_m"])
        fr2 = float(data[1]["Fr"])
        x2 = float(data[1]["x_from_mid_m"])
        if fr2 != fr1:
            _FRAME_SLOPE = (x2 - x1) / (fr2 - fr1)
            _FRAME_OFFSET = x1 - _FRAME_SLOPE * fr1
        print(
            f"[INFO] Frame mapping: SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f}"
        )
    except Exception as e:
        # Fallback: BUSHRA 757 TCP default (Fr 증가 = FWD, Midship = 30.151)
        print(f"[ERROR] Frame JSON parse fail → default: {e}")
        _FRAME_SLOPE = -1.0
        _FRAME_OFFSET = 30.151
        print(
            f"[INFO] Frame mapping: SLOPE={_FRAME_SLOPE:.6f}, OFFSET={_FRAME_OFFSET:.3f} (default)"
        )


def fr_to_x(fr: float) -> float:
    """
    Frame 번호 → x_from_mid_m (m) 변환.

    - 입력: Frame 번호 (Tank Plan 757 TCP 기준, FWD 방향으로 증가)
    - 출력: Midship 기준 x 좌표 (m)
        * x < 0.0  → FWD side
        * x > 0.0  → AFT side

    공식:
        x = _FRAME_OFFSET + _FRAME_SLOPE * Fr

    예시 (BUSHRA, Lpp ≈ 60.30m, Fr_mid ≈ 30.151):
        Fr = 30.151 → x ≈ 0.00 m  (Midship)
        Fr = 60.30  → x ≈ -30.15 m (Forward extremity)
        Fr = 0.00   → x ≈ +30.15 m (Aft extremity)
    """
    return _FRAME_OFFSET + _FRAME_SLOPE * float(fr)


def x_to_fr(x: float) -> float:
    """
    x_from_mid_m (m) → Frame 번호 역변환.

    fr_to_x() 의 정확한 역함수:
        Fr = (x - _FRAME_OFFSET) / _FRAME_SLOPE
    """
    return (float(x) - _FRAME_OFFSET) / _FRAME_SLOPE


def debug_frame_mapping():
    """
    Frame_x_from_mid_m.json 기반으로 현재 SLOPE/OFFSET과
    주요 기준 Frame들의 x_from_mid_m를 출력하는 디버그 함수.
    """
    print("=" * 60)
    print("LCT BUSHRA Frame ↔ x Debug (757 TCP aligned)")
    print("=" * 60)
    print(f"_FRAME_SLOPE  = {_FRAME_SLOPE:.6f}")
    print(f"_FRAME_OFFSET = {_FRAME_OFFSET:.3f}")
    test = {
        "AP approx": 0.0,
        "Midship (Lpp/2)": 30.151,
        "FP approx": 60.30,
        "FWB1/2 center 55": 55.0,
    }
    for label, fr in test.items():
        x = fr_to_x(fr)
        print(
            f"{label:25}  Fr={fr:6.2f}  →  x={x:8.3f} m  ({'FWD' if x < 0 else 'AFT'})"
        )
    print("=" * 60)
    sys.exit(0)


# ============================================================================
# BACKUP PLAN Support Functions
# ============================================================================


def setup_logging(output_file):
    """
    BACKUP PLAN: 실행 로그 설정
    logs/ 폴더에 타임스탬프 로그 파일 생성
    """
    log_dir = os.path.join(os.path.dirname(output_file), "logs")
    os.makedirs(log_dir, exist_ok=True)

    log_file = os.path.join(
        log_dir, f"agi_tr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
        force=True,  # Override any existing configuration
    )

    logging.info("=" * 60)
    logging.info("LCT BUSHRA AGI TR Excel Generation")
    logging.info(f"Output: {output_file}")
    logging.info(f"Log: {log_file}")
    logging.info("=" * 60)

    return log_file


def create_backup_file(original_path):
    """
    BACKUP PLAN: 생성된 파일의 백업 자동 생성
    backups/ 폴더에 타임스탬프 백업 저장 (최근 5개 유지)
    """
    if not os.path.exists(original_path):
        logging.warning(f"Backup failed: {original_path} not found")
        return None

    # 백업 디렉토리 생성
    backup_dir = os.path.join(os.path.dirname(original_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    # 타임스탬프 백업
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.basename(original_path)
    backup_path = os.path.join(backup_dir, f"{timestamp}_{filename}")

    try:
        shutil.copy2(original_path, backup_path)
        logging.info(f"[BACKUP] Created: {backup_path}")
        print(f"  [BACKUP] Created: {os.path.basename(backup_path)}")

        # 오래된 백업 정리 (최근 5개만 유지)
        cleanup_old_backups(backup_dir, keep=5)

        return backup_path
    except Exception as e:
        logging.error(f"[BACKUP] Failed to create backup: {e}")
        print(f"  [BACKUP] Warning: Backup creation failed: {e}")
        return None


def cleanup_old_backups(backup_dir, keep=5):
    """
    BACKUP PLAN: 오래된 백업 파일 정리
    """
    try:
        backups = [
            os.path.join(backup_dir, f)
            for f in os.listdir(backup_dir)
            if f.endswith(".xlsx")
        ]
        backups.sort(key=os.path.getmtime, reverse=True)

        if len(backups) > keep:
            for old_backup in backups[keep:]:
                os.remove(old_backup)
                logging.info(
                    f"[BACKUP] Removed old backup: {os.path.basename(old_backup)}"
                )
    except Exception as e:
        logging.warning(f"[BACKUP] Cleanup failed: {e}")


def preflight_check():
    """
    BACKUP PLAN: 생성 전 환경 검증
    Returns: List of issue messages
    """
    issues = []

    # 1. 필수 디렉토리 확인
    if not os.path.exists("data"):
        issues.append("WARNING: data/ directory not found (fallback will be used)")

    # 2. 필수 JSON 파일 확인
    optional_jsons = [
        "data/Hydro_Table_2D.json",
        "data/gateab_v3_tide_data.json",
        "data/tank_coordinates.json",
        "data/tank_data.json",
        "data/Frame_x_from_mid_m.json",
        "data/hydro_table.json",
    ]
    for json_file in optional_jsons:
        if not os.path.exists(json_file):
            issues.append(f"INFO: {json_file} missing (fallback will be used)")

    # 3. 디스크 공간 확인 (Windows용)
    try:
        import ctypes

        free_bytes = ctypes.c_ulonglong(0)
        ctypes.windll.kernel32.GetDiskFreeSpaceExW(
            ctypes.c_wchar_p("."), None, None, ctypes.pointer(free_bytes)
        )
        free_mb = free_bytes.value / (1024 * 1024)
        if free_mb < 10:
            issues.append(f"ERROR: Low disk space ({free_mb:.1f}MB)")
    except Exception:
        # Unix-like systems
        try:
            statvfs = os.statvfs(".")
            free_mb = (statvfs.f_frsize * statvfs.f_bavail) / (1024 * 1024)
            if free_mb < 10:
                issues.append(f"ERROR: Low disk space ({free_mb:.1f}MB)")
        except Exception:
            pass  # Skip disk check if not supported

    return issues


class BackupRecoveryError(Exception):
    """BACKUP PLAN: 백업 복구 가능한 에러"""

    pass


def safe_sheet_creation(wb, sheet_func, sheet_name, *args, **kwargs):
    """
    BACKUP PLAN: 시트 생성 wrapper with error recovery
    시트 생성 실패 시에도 계속 진행
    """
    try:
        logging.info(f"Creating sheet: {sheet_name}")
        result = sheet_func(wb, *args, **kwargs)
        logging.info(f"✓ {sheet_name} created successfully")
        return result
    except Exception as e:
        logging.error(f"✗ {sheet_name} creation failed: {e}")
        logging.warning(f"[BACKUP] Skipping {sheet_name}, continuing...")
        print(f"  [BACKUP] Warning: {sheet_name} creation failed, continuing...")
        return None


def get_fixed_tank_data():
    """
    Forward fresh water ballast tanks (FWB1/2, Fr 48–65) - 757 TCP Tank Plan 기준.
    JSON tank_coordinates.json의 실제 Frame 값 사용:
    - FWB1: Fr 56-65, Mid_Fr=60.5
    - FWB2: Fr 48-53, Mid_Fr=50.5
    """
    # Frame to X 변환: x = 30.151 - Frame (757 TCP 기준, Midship=30.151)
    return {
        # Forward fresh water ballast tanks (FWB1/2, Fr 48–65) - 757 TCP Tank Plan
        "FWB1.P": {
            "x": fr_to_x(60.5),  # Fr 56-65, Mid_Fr=60.5 (757 TCP Tank Plan)
            "max_t": 50.57,
            "SG": 1.025,
            "note": "Forward Port (Fr 56-65, Mid_Fr=60.5)",
        },
        "FWB1.S": {
            "x": fr_to_x(60.5),  # Fr 56-65, Mid_Fr=60.5 (757 TCP Tank Plan)
            "max_t": 50.57,
            "SG": 1.025,
            "note": "Forward Stbd (Fr 56-65, Mid_Fr=60.5)",
        },
        "FWB2.P": {
            "x": fr_to_x(50.5),  # Fr 48-53, Mid_Fr=50.5 (757 TCP Tank Plan)
            "max_t": 109.98,
            "SG": 1.025,
            "note": "Forward-Mid Port (Fr 48-53, Mid_Fr=50.5)",
        },
        "FWB2.S": {
            "x": fr_to_x(50.5),  # Fr 48-53, Mid_Fr=50.5 (757 TCP Tank Plan)
            "max_t": 109.98,
            "SG": 1.025,
            "note": "Forward-Mid Stbd (Fr 48-53, Mid_Fr=50.5)",
        },
        # 선수/중앙 탱크 (참고용, 실제 사용 시나리오에서는 비워둠)
        "FWCARGO1.P": {
            "x": fr_to_x(42.0),
            "max_t": 148.35,
            "SG": 1.000,
            "note": "Fwd Cargo (Empty)",
        },
        "FWCARGO1.S": {
            "x": fr_to_x(42.0),
            "max_t": 148.35,
            "SG": 1.000,
            "note": "Fwd Cargo (Empty)",
        },
    }


def build_tank_lookup():
    """
    Tank 좌표/용량 JSON에서 Ballast 탱크 정보 취합.

    사용 키:
    - tank_coordinates.json: data[].Tank_Name, Mid_Fr, Weight_MT, Volume_m3
    - tank_data.json       : data[].Tank_Name, Weight_MT (실측 100% kg/ton 기준)

    [CRITICAL FIX] FWB1, FWB2는 get_fixed_tank_data()의 수정된 좌표 사용
    """
    # 먼저 고정된 탱크 데이터 가져오기
    fixed_data = get_fixed_tank_data()

    coords = _load_json("data/tank_coordinates.json")
    tdata = _load_json("data/tank_data.json")

    if not coords or not tdata:
        # JSON이 없으면 고정 데이터만 반환
        lookup = {}
        for tank_name, tank_info in fixed_data.items():
            lookup[tank_name] = {
                "x_from_mid_m": round(float(tank_info["x"]), 2),
                "max_t": tank_info["max_t"],
                "SG": tank_info["SG"],
                "air_vent_mm": 80 if tank_name.startswith("FWB") else 125,
            }
        return lookup

    coord_map = {row["Tank_Name"]: row for row in coords.get("data", [])}
    data_map = {row["Tank_Name"]: row for row in tdata.get("data", [])}

    lookup = {}

    for tank_name, c in coord_map.items():
        # FWB1, FWB2는 고정 데이터 사용 (CRITICAL FIX)
        if tank_name in fixed_data:
            fixed_info = fixed_data[tank_name]
            lookup[tank_name] = {
                "x_from_mid_m": round(float(fixed_info["x"]), 2),
                "max_t": fixed_info["max_t"],
                "SG": fixed_info["SG"],
                "air_vent_mm": 80 if tank_name.startswith("FWB") else 125,
            }
            continue

        # 다른 탱크는 기존 로직 사용
        mid_fr = c.get("Mid_Fr")
        x_from_mid = fr_to_x(mid_fr) if mid_fr is not None else None

        d = data_map.get(tank_name, {})
        weight_mt = d.get("Weight_MT") or c.get("Weight_MT")

        # SG 및 air vent는 기본 규칙 (필요하면 여기서 조정)
        if tank_name.startswith("FWB"):
            sg = 1.025
            air_vent = 80
        elif tank_name.startswith("FWCARGO"):
            sg = 1.000
            air_vent = 125
        else:
            sg = 1.000
            air_vent = ""

        lookup[tank_name] = {
            "x_from_mid_m": (
                round(float(x_from_mid), 2) if x_from_mid is not None else None
            ),
            "max_t": round(float(weight_mt), 2) if weight_mt is not None else None,
            "SG": sg,
            "air_vent_mm": air_vent,
        }

    return lookup


# 함수 생성 헬퍼 함수
def create_index_match_formula(lookup_value, lookup_range, return_range):
    """INDEX/MATCH 조합 수식 생성"""
    return f'=INDEX({return_range}, MATCH("{lookup_value}", {lookup_range}, 0))'


# ============================================================================
# Calc Sheet Creation
# ============================================================================


def create_calc_sheet(wb):
    """Calc 시트 생성 - 원본 파일 구조와 동일하게 생성"""
    ws = wb.create_sheet("Calc")
    styles = get_styles()

    # Row 2: 제목
    ws.cell(row=2, column=2).value = "LCT BUSHRA — RORO Calculator & Limits"
    ws.cell(row=2, column=2).font = styles["title_font"]

    # Row 3: 헤더
    headers_row3 = ["", "SECTION", "PARAMETER", "UNIT", "VALUE", "NOTES"]
    for col_idx, header in enumerate(headers_row3, start=1):
        if header:
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # INPUT CONSTANTS 섹션 (Row 5-8)
    ws.cell(row=5, column=2).value = "INPUT CONSTANTS"
    ws.cell(row=5, column=3).value = "L_ramp_m"
    ws.cell(row=5, column=4).value = "m"
    ws.cell(row=5, column=5).value = 12.0
    ws.cell(row=5, column=5).fill = styles["input_fill"]
    ws.cell(row=5, column=6).value = "Linkspan length. Calc!D4"

    ws.cell(row=6, column=3).value = "theta_max_deg"
    ws.cell(row=6, column=4).value = "deg"
    ws.cell(row=6, column=5).value = 6.0
    ws.cell(row=6, column=5).fill = styles["input_fill"]
    ws.cell(row=6, column=6).value = "Max ramp angle. Calc!D5"

    ws.cell(row=7, column=3).value = "KminusZ_m"
    ws.cell(row=7, column=4).value = "m"
    ws.cell(row=7, column=5).value = 3.0
    ws.cell(row=7, column=5).fill = styles["input_fill"]
    ws.cell(row=7, column=6).value = "K - Z (UPDATE!). Calc!D6"

    ws.cell(row=8, column=3).value = "D_vessel_m"
    ws.cell(row=8, column=4).value = "m"
    ws.cell(row=8, column=5).value = 3.65
    ws.cell(row=8, column=5).fill = styles["input_fill"]
    ws.cell(row=8, column=6).value = "Molded depth. Calc!D7"

    # LIMITS & OPS 섹션 (Row 10-12)
    ws.cell(row=10, column=2).value = "LIMITS & OPS"
    ws.cell(row=10, column=3).value = "min_fwd_draft_m"
    ws.cell(row=10, column=4).value = "m"
    ws.cell(row=10, column=5).value = 1.5
    ws.cell(row=10, column=5).fill = styles["input_fill"]
    ws.cell(row=10, column=6).value = "Min draft. Calc!D9"

    ws.cell(row=11, column=3).value = "max_fwd_draft_m"
    ws.cell(row=11, column=4).value = "m"
    ws.cell(row=11, column=5).value = 3.5
    ws.cell(row=11, column=5).fill = styles["input_fill"]
    ws.cell(row=11, column=6).value = "Max draft. Calc!D10"

    ws.cell(row=12, column=3).value = "pump_rate_tph"
    ws.cell(row=12, column=4).value = "t/h"
    ws.cell(row=12, column=5).value = 10.0
    ws.cell(row=12, column=5).fill = styles["input_fill"]
    ws.cell(row=12, column=6).value = "Pump rate. Calc!D11"

    # STABILITY 섹션 (Row 14-17)
    ws.cell(row=14, column=2).value = "STABILITY"
    ws.cell(row=14, column=3).value = "MTC_t_m_per_cm"
    ws.cell(row=14, column=4).value = "t·m/cm"
    ws.cell(row=14, column=5).value = (
        34.00  # BUSHRA verified: Reverse-eng from ΔTM 26035 t·m / Δtrim 765cm + booklet
    )
    ws.cell(row=14, column=5).fill = styles["input_fill"]
    ws.cell(row=14, column=6).value = "MTC. Calc!D13"

    ws.cell(row=15, column=3).value = "LCF_m_from_midship"
    ws.cell(row=15, column=4).value = "m"
    ws.cell(row=15, column=5).value = 0.76
    ws.cell(row=15, column=5).fill = styles["input_fill"]
    ws.cell(row=15, column=6).value = "LCF from Midship (Corrected). Calc!D14"

    ws.cell(row=16, column=3).value = "TPC_t_per_cm"
    ws.cell(row=16, column=4).value = "t/cm"
    ws.cell(row=16, column=5).value = (
        8.00  # BUSHRA verified: Approx waterplane 14×60.3×0.85×1.025 ≈680 m² → TPC≈8.00
    )
    ws.cell(row=16, column=5).fill = styles["input_fill"]
    ws.cell(row=16, column=6).value = "TPC. Calc!D15"

    ws.cell(row=17, column=3).value = "Lpp_m"
    ws.cell(row=17, column=4).value = "m"
    ws.cell(row=17, column=5).value = 60.302
    ws.cell(row=17, column=5).fill = styles["input_fill"]
    ws.cell(row=17, column=6).value = "Length between perpendiculars. Calc!D16"

    ws.cell(row=18, column=2).value = "OPERATIONS"
    ws.cell(row=18, column=3).value = "max_fwd_draft_ops_m"
    ws.cell(row=18, column=4).value = "m"
    ws.cell(row=18, column=5).value = 2.70
    ws.cell(row=18, column=5).fill = styles["input_fill"]
    ws.cell(row=18, column=6).value = "Max forward draft for operations. Calc!D9"

    ws.cell(row=19, column=3).value = "ramp_door_offset_m"
    ws.cell(row=19, column=4).value = "m"
    ws.cell(row=19, column=5).value = 0.15
    ws.cell(row=19, column=5).fill = styles["input_fill"]
    ws.cell(row=19, column=6).value = "Ramp door offset. Calc!D11"

    ws.cell(row=20, column=3).value = "linkspan_freeboard_target_m"
    ws.cell(row=20, column=4).value = "m"
    ws.cell(row=20, column=5).value = 0.28
    ws.cell(row=20, column=5).fill = styles["input_fill"]
    ws.cell(row=20, column=6).value = "Linkspan freeboard target. Calc!D12"

    ws.cell(row=21, column=3).value = "gm_target_m"
    ws.cell(row=21, column=4).value = "m"
    ws.cell(row=21, column=5).value = 1.50
    ws.cell(row=21, column=5).fill = styles["input_fill"]
    ws.cell(row=21, column=6).value = "GM target. Calc!D13"

    # STRUCTURAL LIMITS 섹션 (Row 23-26)
    ws.cell(row=23, column=2).value = "STRUCTURAL LIMITS"
    ws.cell(row=23, column=3).value = "limit_reaction_t"
    ws.cell(row=23, column=4).value = "t"
    ws.cell(row=23, column=5).value = 201.60
    ws.cell(row=23, column=5).fill = styles["input_fill"]
    ws.cell(row=23, column=6).value = (
        "Aries Ramp hinge limit 201.60 t (share ratio 0.545, 2025-11-18)"
    )

    ws.cell(row=24, column=3).value = "limit_share_load_t"
    ws.cell(row=24, column=4).value = "t"
    ws.cell(row=24, column=5).value = 118.80
    ws.cell(row=24, column=5).fill = styles["input_fill"]
    ws.cell(row=24, column=6).value = "Max Share Load on LCT (Mammoet)"

    ws.cell(row=25, column=3).value = "limit_deck_press_tpm2"
    ws.cell(row=25, column=4).value = "t/m²"
    ws.cell(row=25, column=5).value = 10.00
    ws.cell(row=25, column=5).fill = styles["input_fill"]
    ws.cell(row=25, column=6).value = "Max Deck Pressure (Spec)"

    ws.cell(row=26, column=3).value = "linkspan_area_m2"
    ws.cell(row=26, column=4).value = "m²"
    ws.cell(row=26, column=5).value = 12.00
    ws.cell(row=26, column=5).fill = styles["input_fill"]
    ws.cell(row=26, column=6).value = (
        "Linkspan 실제 접지 12.00 m² (Ramp 1 TR only 규정)"
    )

    # BALLAST FIX CHECK 섹션 (Row 27-28)
    ws.cell(row=27, column=2).value = "BALLAST FIX CHECK"
    ws.cell(row=27, column=3).value = "max_aft_ballast_cap_t"
    ws.cell(row=27, column=4).value = "t"
    ws.cell(row=27, column=5).value = 1200.00
    ws.cell(row=27, column=5).fill = styles["input_fill"]
    ws.cell(row=27, column=6).value = "Max Forward Ballast Capacity (FWB1/2, Fr 48–65)"

    ws.cell(row=28, column=3).value = "max_pump_time_h"
    ws.cell(row=28, column=4).value = "h"
    ws.cell(row=28, column=5).value = 6.00
    ws.cell(row=28, column=5).fill = styles["input_fill"]
    ws.cell(row=28, column=6).value = "Max Allowed Pump Time for Fix"

    # VENT & PUMP 섹션 (Row 29-31)
    ws.cell(row=29, column=2).value = "VENT & PUMP"
    ws.cell(row=29, column=3).value = "vent_flow_coeff"
    ws.cell(row=29, column=4).value = "t/h per mm"
    ws.cell(row=29, column=5).value = 0.86
    ws.cell(row=29, column=5).fill = styles["input_fill"]
    ws.cell(row=29, column=6).value = "실측 보정 0.86 (2025-11-18, MAPE 0.30%)"

    ws.cell(row=30, column=3).value = "pump_rate_tph"
    ws.cell(row=30, column=4).value = "t/h"
    ws.cell(row=30, column=5).value = 100.00
    ws.cell(row=30, column=5).fill = styles["input_fill"]
    ws.cell(row=30, column=6).value = "Hired pump rate"

    ws.cell(row=31, column=3).value = "pump_rate_effective_tph"
    ws.cell(row=31, column=4).value = "t/h"
    ws.cell(row=31, column=5).value = (
        '=MIN(E30, SUMPRODUCT((Ballast_Tanks!E$2:E$100="Y")*(Ballast_Tanks!F$2:F$100)*E29))'
    )
    ws.cell(row=31, column=5).fill = styles["ok_fill"]
    ws.cell(row=31, column=6).value = "실효 펌프 속도 (vent bottleneck, 68.80 t/h)"

    # RAMP GEOMETRY 섹션 (Row 32-35)
    ws.cell(row=32, column=2).value = "RAMP GEOMETRY"
    ws.cell(row=32, column=3).value = "ramp_hinge_x_mid_m"
    ws.cell(row=32, column=4).value = "m"
    ws.cell(row=32, column=5).value = -30.151
    ws.cell(row=32, column=5).fill = styles["input_fill"]
    ws.cell(row=32, column=6).value = "LBP 60.302 m 기준"

    ws.cell(row=33, column=3).value = "ramp_length_m"
    ws.cell(row=33, column=4).value = "m"
    ws.cell(row=33, column=5).value = 8.30
    ws.cell(row=33, column=5).fill = styles["input_fill"]
    ws.cell(row=33, column=6).value = "TRE Cert 2020-08-04"

    ws.cell(row=34, column=3).value = "linkspan_height_m"
    ws.cell(row=34, column=4).value = "m"
    ws.cell(row=34, column=5).value = 2.00
    ws.cell(row=34, column=5).fill = styles["input_fill"]

    ws.cell(row=35, column=3).value = "ramp_end_clearance_min_m"
    ws.cell(row=35, column=4).value = "m"
    ws.cell(row=35, column=5).value = 0.40
    ws.cell(row=35, column=5).fill = styles["input_fill"]

    # HINGE STRESS 섹션 (Row 36-37)
    ws.cell(row=36, column=2).value = "HINGE STRESS"
    ws.cell(row=36, column=3).value = "hinge_pin_area_m2"
    ws.cell(row=36, column=4).value = "m²"
    ws.cell(row=36, column=5).value = 0.117
    ws.cell(row=36, column=5).fill = styles["input_fill"]
    ws.cell(row=36, column=6).value = "Doubler 390x300 mm (Aries)"

    ws.cell(row=37, column=3).value = "hinge_limit_rx_t"
    ws.cell(row=37, column=4).value = "t"
    ws.cell(row=37, column=5).value = 201.60
    ws.cell(row=37, column=5).fill = styles["input_fill"]
    ws.cell(row=37, column=6).value = (
        "Max Hinge Reaction (duplicate of E23 for clarity)"
    )

    # PRECISION PARAMETERS 섹션 (Row 39-43)
    ws.cell(row=39, column=2).value = "PRECISION PARAMETERS"
    ws.cell(row=39, column=2).font = styles["normal_font"]

    ws.cell(row=40, column=3).value = "LBP_m"
    ws.cell(row=40, column=4).value = "m"
    ws.cell(row=40, column=5).value = 60.302  # LBP_m
    ws.cell(row=40, column=5).comment = Comment("LBP (m) - Calc!$E$40", "System")
    ws.cell(row=40, column=5).fill = styles["input_fill"]
    ws.cell(row=40, column=6).value = (
        "Length Between Perpendiculars (for precise draft calculation)"
    )

    ws.cell(row=41, column=3).value = "LCF_from_mid_m"
    ws.cell(row=41, column=4).value = "m"
    ws.cell(row=41, column=5).value = (
        0.76  # BUSHRA verified: LCF_from_AP (30.91) - AP_to_midship (30.151) = 0.759 m
    )
    ws.cell(row=41, column=5).comment = Comment(
        "LCF from mid (m) - Calc!$E$41", "System"
    )
    ws.cell(row=41, column=5).fill = styles["input_fill"]
    ws.cell(row=41, column=6).value = (
        "LCF from midship (BUSHRA verified, for precise draft calculation)"
    )

    ws.cell(row=42, column=3).value = "dynamic_factor"
    ws.cell(row=42, column=4).value = "-"
    ws.cell(row=42, column=5).value = 1.15
    ws.cell(row=42, column=5).fill = styles["input_fill"]
    ws.cell(row=42, column=6).value = (
        "Dynamic load amplification factor (for Load Case B)"
    )

    ws.cell(row=43, column=3).value = "heel_y_offset_m"
    ws.cell(row=43, column=4).value = "m"
    ws.cell(row=43, column=5).value = 1.50
    ws.cell(row=43, column=5).fill = styles["input_fill"]
    ws.cell(row=43, column=6).value = "Heel y-offset (for heel angle calculation)"

    # 폰트 적용
    for row in range(5, 44):
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = styles["normal_font"]

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35

    print("  [OK] Calc sheet created with VENT&PUMP 실측 0.86")


# ============================================================================
# Tide Sheet Creation
# ============================================================================


def create_tide_sheet(wb):
    """December_Tide_2025 시트 생성"""
    ws = wb.create_sheet("December_Tide_2025")
    styles = get_styles()

    headers = ["datetime_gst", "tide_m               (Chart Datum)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # JSON 파일 로드 (상대 경로 사용)
    tide_data = _load_json("data/gateab_v3_tide_data.json")

    if tide_data:
        try:
            for idx, entry in enumerate(tide_data, start=0):
                row = 2 + idx
                if row > 745:
                    break

                cell_a = ws.cell(row=row, column=1)
                cell_a.value = entry.get("datetime", "")
                cell_a.font = styles["normal_font"]

                cell_b = ws.cell(row=row, column=2)
                cell_b.value = entry.get("tide_m", 0.0)
                cell_b.font = styles["normal_font"]
                cell_b.number_format = "0.00"
            print(f"  [OK] December_Tide_2025 sheet created with {len(tide_data)} rows")
        except Exception as e:
            print(f"  [WARNING] Error processing tide data: {e}. Creating empty sheet.")
            tide_data = None

    if not tide_data:
        print(
            f"  [WARNING] JSON file not found. Creating empty December_Tide_2025 sheet."
        )
        for row in range(2, 746):
            ws.cell(row=row, column=1).font = styles["normal_font"]
            ws.cell(row=row, column=2).font = styles["normal_font"]
            ws.cell(row=row, column=2).number_format = "0.00"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


# ============================================================================
# Hourly Sheet Creation
# ============================================================================


def create_hourly_sheet(wb):
    """Hourly_FWD_AFT_Heights 시트 생성"""
    ws = wb.create_sheet("Hourly_FWD_AFT_Heights")
    styles = get_styles()

    headers = [
        "DateTime (GST)",
        "Tide_m",
        "Dfwd_req_m (even)",
        "Trim_m (optional)",
        "Dfwd_adj_m",
        "Daft_adj_m",
        "Ramp_Angle_deg",
        "Status",
        "FWD_Height_m",
        "AFT_Height_m",
        "Notes",
        "",
        "Trim_m (optional)",
        "",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header if header else ""
        if header:
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    for row in range(2, 746):
        row_str = str(row)
        ws.cell(row=row, column=1).value = (
            f'=IF(December_Tide_2025!A{row_str}="","",December_Tide_2025!A{row_str})'
        )
        ws.cell(row=row, column=2).value = (
            f'=IF(December_Tide_2025!B{row_str}="","",December_Tide_2025!B{row_str})'
        )
        ws.cell(row=row, column=3).value = (
            f'=IF($A{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) + $B{row_str} - '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)) * '
            f'TAN(RADIANS(INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0)))))'
        )
        ws.cell(row=row, column=5).value = (
            f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} - $D{row_str}/2))'
        )
        ws.cell(row=row, column=6).value = (
            f'=IF($C{row_str}="","", IF($D{row_str}="", $C{row_str}, $C{row_str} + $D{row_str}/2))'
        )
        ws.cell(row=row, column=7).value = (
            f'=IF($E{row_str}="","", '
            f'DEGREES(ATAN((INDEX(Calc!$E:$E, MATCH("KminusZ_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str}) / '
            f'INDEX(Calc!$E:$E, MATCH("L_ramp_m", Calc!$C:$C, 0)))))'
        )
        ws.cell(row=row, column=8).value = (
            f'=IF($E{row_str}="","", '
            f'IF(AND($E{row_str}>=INDEX(Calc!$E:$E, MATCH("min_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$E{row_str}<=INDEX(Calc!$E:$E, MATCH("max_fwd_draft_m", Calc!$C:$C, 0)), '
            f'$G{row_str}<=INDEX(Calc!$E:$E, MATCH("theta_max_deg", Calc!$C:$C, 0))), "OK", "CHECK"))'
        )
        ws.cell(row=row, column=9).value = (
            f'=IF($E{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $E{row_str} + $B{row_str})'
        )
        ws.cell(row=row, column=10).value = (
            f'=IF($F{row_str}="","", '
            f'INDEX(Calc!$E:$E, MATCH("D_vessel_m", Calc!$C:$C, 0)) - $F{row_str} + $B{row_str})'
        )
        ws.cell(row=row, column=11).value = f'=IF(D{row_str}=0, "Even Keel", "")'

        if row == 2:
            ws.cell(row=row, column=14).value = (
                "← Defaults to 0.00 (Even-Keel). To apply the actual trim, manually enter the value in this cell."
            )
            ws.cell(row=row, column=14).font = styles["normal_font"]

        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=3).number_format = "0.00"
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.00"
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"
        ws.cell(row=row, column=10).number_format = "0.00"

        for col in range(1, 12):
            ws.cell(row=row, column=col).font = styles["normal_font"]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 80

    print(f"  [OK] Hourly_FWD_AFT_Heights sheet created")


# ============================================================================
# RORO Sheet Creation
# ============================================================================


def build_opt_c_stage():
    """
    Stage 6A_Critical (Opt C) - Ballast Integrated (DAS Method v4.3 Final Optimized)
    - TR1: final deck position (Fr ≈ 42.0)
    - TR2: ramp tip position (Fr ≈ -5.0)
    - 화물 중량: 280t (TR 217t + SPMT/Acc 63t)
    - Pre-Ballast: 250t at forward fresh water ballast tanks (FWB1/2, Fr 48–65) center
      - FWB1: Fr 56-65, Mid_Fr=60.5
      - FWB2: Fr 48-53, Mid_Fr=50.5
      - Center: Fr 55.5 → x = fr_to_x(55.5) ≈ -25.35m
    - Total: 810t (Cargo 560t + Ballast 250t)
    - Combined LCG: 재계산 (Even Keel에 근접)
    - target_trim_cm 은 Even Keel 목표로 설정
    """
    # 1. 화물 (Cargo)
    fr_tr1 = 42.0  # TR1 final stowage frame
    fr_tr2 = -5.0  # TR2 ramp tip frame (estimated)
    w_tr = 280.0  # TR 217t + SPMT/Acc 63t = 280t
    tr1_weight = w_tr
    tr1_pos_x = fr_to_x(fr_tr1)  # ≈ 11.85 m
    tr2_weight = w_tr
    tr2_pos_x = fr_to_x(fr_tr2)  # ≈ -35.15 m

    # 2. 밸러스트 (Pre-Ballast for Opt C - DAS Method v4.3 Final)
    # DAS Method: 250t Shore Water to Forward Ballast Tanks
    # 위치: forward fresh water ballast tanks (FWB1/2, Fr 48–65) 중심
    # FWB1: Fr 56-65, Mid_Fr=60.5 → x = fr_to_x(60.5) ≈ -30.35m
    # FWB2: Fr 48-53, Mid_Fr=50.5 → x = fr_to_x(50.5) ≈ -20.35m
    # 평균 중심: (60.5 + 50.5) / 2 = 55.5 → x = fr_to_x(55.5) ≈ -25.35m
    # 무게: 250t (최적화된 값)
    w_bal = 250.0
    # Forward ballast tanks center (FWB1/2, Fr 48–65) - 757 TCP Tank Plan 기준
    ballast_fr_center = 55.5  # (FWB1 Mid_Fr 60.5 + FWB2 Mid_Fr 50.5) / 2
    ballast_pos_x = fr_to_x(ballast_fr_center)

    # 3. Total Combined (Cargo + Ballast)
    total_weight_opt_c = tr1_weight + tr2_weight + w_bal

    # 모멘트 합산 / 총중량 = 새로운 LCG
    # (Cargo Moment + Ballast Moment) / Total Weight
    moment_cargo = (tr1_weight * tr1_pos_x) + (tr2_weight * tr2_pos_x)
    moment_ballast = w_bal * ballast_pos_x
    combined_lcg_opt_c = (moment_cargo + moment_ballast) / total_weight_opt_c

    return {
        "name": "Stage 6A_Critical (Opt C)",
        "weight_t": total_weight_opt_c,  # 810.00 t
        "x_from_mid_m": combined_lcg_opt_c,  # 재계산된 LCG (Ballast 포함) ≈ -0.34m
        "target_trim_cm": 0.0,  # Even Keel 목표
    }


def create_roro_sheet(wb: Workbook):
    """RORO_Stage_Scenarios 생성 (최적 안전값 240cm 적용 완료)"""
    ws = wb.create_sheet("RORO_Stage_Scenarios")
    styles = get_styles()

    # Title
    ws["A1"] = "RORO Stage Scenarios – Option C (Target 240cm Safe Margin)"
    ws["A1"].font = styles["title_font"]

    # Row 2: Input parameter 안내
    ws["C2"] = "← Input parameter(yellow cellls only)"
    ws["C2"].font = styles["normal_font"]

    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # Row 4: 섹션 제목 추가
    ws["A4"] = "1.Critical Stage Verification: Trim & Draft Status Sequence"
    ws["A4"].font = styles["normal_font"]
    ws["F4"] = "2. Stage Sequence Table"
    ws["F4"].font = styles["normal_font"]
    ws["N4"] = "3.Ballast Range Table"
    ws["N4"].font = styles["normal_font"]

    # Row 5: 헤더 추가
    ws["A5"] = "Parameter"
    ws["A5"].font = styles["header_font"]
    ws["A5"].fill = styles["header_fill"]
    ws["A5"].alignment = styles["center_align"]

    ws["B5"] = "Value"
    ws["B5"].font = styles["header_font"]
    ws["B5"].fill = styles["header_fill"]
    ws["B5"].alignment = styles["center_align"]

    ws["C5"] = "Unit"
    ws["C5"].font = styles["header_font"]
    ws["C5"].fill = styles["header_fill"]
    ws["C5"].alignment = styles["center_align"]

    ws["D5"] = "REMARK"
    ws["D5"].font = styles["header_font"]
    ws["D5"].fill = styles["header_fill"]
    ws["D5"].alignment = styles["center_align"]

    ws["F5"] = "Stage"
    ws["F5"].font = styles["header_font"]
    ws["F5"].fill = styles["header_fill"]
    ws["F5"].alignment = styles["center_align"]

    ws["G5"] = "EXPLANATION"
    ws["G5"].font = styles["header_font"]
    ws["G5"].fill = styles["header_fill"]
    ws["G5"].alignment = styles["center_align"]

    ws["N5"] = "Ballast (ton)"
    ws["N5"].font = styles["header_font"]
    ws["N5"].fill = styles["header_fill"]
    ws["N5"].alignment = styles["center_align"]

    ws["O5"] = "Fwd Draft (m)"
    ws["O5"].font = styles["header_font"]
    ws["O5"].fill = styles["header_fill"]
    ws["O5"].alignment = styles["center_align"]

    ws["P5"] = "Freeboard (m)"
    ws["P5"].font = styles["header_font"]
    ws["P5"].fill = styles["header_fill"]
    ws["P5"].alignment = styles["center_align"]

    ws["Q5"] = "Aft Draft (m)"
    ws["Q5"].font = styles["header_font"]
    ws["Q5"].fill = styles["header_fill"]
    ws["Q5"].alignment = styles["center_align"]

    ws["R5"] = "Status Assessment"
    ws["R5"].font = styles["header_font"]
    ws["R5"].fill = styles["header_fill"]
    ws["R5"].alignment = styles["center_align"]

    # 파라미터 세로 배치 (A6-A15: Parameter, B6-B15: Value, C6-C15: Unit, D6-D15: REMARK, F6-F15: Stage)
    # A6: Tmean_baseline
    ws["A6"] = "Tmean_baseline"
    ws["B6"] = 2.00  # BUSHRA verified: Sim Stage 1/7 even keel
    ws["B6"].fill = styles["input_fill"]
    ws["B6"].font = styles["normal_font"]
    ws["B6"].number_format = number_format
    ws["C6"] = "m"
    ws["C6"].font = styles["normal_font"]
    ws["D6"] = "Baseline mean draft"
    ws["D6"].font = styles["normal_font"]
    ws["F6"] = "Stage 1"
    ws["F6"].font = styles["normal_font"]

    # A7: Tide_ref
    ws["A7"] = "Tide_ref"
    ws["B7"] = 2.00  # BUSHRA verified: Mina Zayed high tide avg 1.80-2.20m
    ws["B7"].fill = styles["input_fill"]
    ws["B7"].font = styles["normal_font"]
    ws["B7"].number_format = number_format
    ws["C7"] = "m"
    ws["C7"].font = styles["normal_font"]
    ws["D7"] = "Reference tide level"
    ws["D7"].font = styles["normal_font"]
    ws["F7"] = "Stage 2"
    ws["F7"].font = styles["normal_font"]

    # A8: Trim_target_cm
    ws["A8"] = "Trim_target_cm"
    ws["B8"] = 10.00  # BUSHRA verified: Ops safe limit (by stern max)
    ws["B8"].fill = styles["input_fill"]
    ws["B8"].font = styles["normal_font"]
    ws["B8"].number_format = number_format
    ws["C8"] = "cm"
    ws["C8"].font = styles["normal_font"]
    ws["D8"] = "Target trim"
    ws["D8"].font = styles["normal_font"]
    ws["F8"] = "Stage 3"
    ws["F8"].font = styles["normal_font"]

    # A9: MTC
    ws["A9"] = "MTC"
    ws["B9"] = '=INDEX(Calc!$E:$E, MATCH("MTC_t_m_per_cm", Calc!$C:$C, 0))'
    ws["B9"].fill = styles["input_fill"]
    ws["B9"].font = styles["normal_font"]
    ws["B9"].number_format = number_format
    ws["C9"] = "t·m/cm"
    ws["C9"].font = styles["normal_font"]
    ws["D9"] = "Moment to change trim"
    ws["D9"].font = styles["normal_font"]
    ws["F9"] = "Stage 4"
    ws["F9"].font = styles["normal_font"]

    # A10: LCF (midship 기준, TM 계산용)
    ws["A10"] = "LCF"
    ws["B10"] = (
        '=INDEX(Calc!$E:$E, MATCH("LCF_m_from_midship", Calc!$C:$C, 0))'  # BUSHRA verified: 0.76 m (midship 기준)
    )
    ws["B10"].font = styles["normal_font"]
    ws["B10"].number_format = number_format
    ws["B10"].fill = styles["input_fill"]
    ws["C10"] = "m"
    ws["C10"].font = styles["normal_font"]
    ws["D10"] = "Longitudinal center of flotation (from midship, for TM calculation)"
    ws["D10"].font = styles["normal_font"]
    ws["F10"] = "Stage 5"
    ws["F10"].font = styles["normal_font"]

    # A11: D_vessel
    ws["A11"] = "D_vessel"
    ws["B11"] = 3.65  # BUSHRA verified: Booklet + TCP confirmed
    ws["B11"].font = styles["normal_font"]
    ws["B11"].number_format = number_format
    ws["B11"].fill = styles["input_fill"]
    ws["C11"] = "m"
    ws["C11"].font = styles["normal_font"]
    ws["D11"] = "Vessel depth"
    ws["D11"].font = styles["normal_font"]
    ws["F11"] = "Stage 5_PreBallast"
    ws["F11"].font = styles["normal_font"]

    # A12: TPC
    ws["A12"] = "TPC"
    ws["B12"] = '=INDEX(Calc!$E:$E, MATCH("TPC_t_per_cm", Calc!$C:$C, 0))'
    ws["B12"].font = styles["normal_font"]
    ws["B12"].number_format = number_format
    ws["B12"].fill = styles["input_fill"]
    ws["C12"] = "t/cm"
    ws["C12"].font = styles["normal_font"]
    ws["D12"] = "Tons per centimeter"
    ws["D12"].font = styles["normal_font"]
    ws["F12"] = "Stage 6A_Critical (Opt C)"
    ws["F12"].font = styles["normal_font"]

    # A13: pump_rate_effective_tph
    ws["A13"] = "pump_rate_effective_tph"
    ws["B13"] = (
        '=INDEX(Calc!$E:$E, MATCH("pump_rate_effective_tph", Calc!$C:$C, 0))'  # BUSHRA verified: 100.00 t/h (2×50 t/h pumps)
    )
    ws["B13"].fill = styles["input_fill"]
    ws["B13"].font = styles["normal_font"]
    ws["B13"].number_format = number_format
    ws["C13"] = "t/h"
    ws["C13"].font = styles["normal_font"]
    ws["D13"] = "Effective pump rate"
    ws["D13"].font = styles["normal_font"]
    ws["F13"] = "Stage 6C"
    ws["F13"].font = styles["normal_font"]

    # A14: X_Ballast (AP 기준)
    ws["A14"] = "X_Ballast"
    ws["B14"] = (
        52.50  # BUSHRA verified: Forward ballast avg from AP (FWB2 Mid_Fr 50.5 + FWB1 Mid_Fr 60.5)/2 ≈55.5, ops 52.50m
    )
    ws["B14"].font = styles["normal_font"]
    ws["B14"].number_format = number_format
    ws["B14"].fill = styles["input_fill"]
    ws["C14"] = "m"
    ws["C14"].font = styles["normal_font"]
    ws["D14"] = "Ballast center position (from AP)"
    ws["D14"].font = styles["normal_font"]
    ws["F14"] = "Stage 7"
    ws["F14"].font = styles["normal_font"]

    # A15: Lpp
    ws["A15"] = "Lpp"
    ws["B15"] = 60.302  # BUSHRA verified: Booklet confirmed
    ws["B15"].font = styles["normal_font"]
    ws["B15"].number_format = number_format
    ws["B15"].fill = styles["input_fill"]
    ws["C15"] = "m"
    ws["C15"].font = styles["normal_font"]
    ws["D15"] = "Length between perpendiculars"
    ws["D15"].font = styles["normal_font"]

    # Ballast Range Table 데이터 (N6-R12)
    # Row 6: 0 ~ 170 ton
    ws["N6"] = "0 ~ 170"
    ws["N6"].font = styles["normal_font"]
    ws["O6"] = "> 3.65"
    ws["O6"].font = styles["normal_font"]
    ws["P6"] = "(-) Submerged"
    ws["P6"].font = styles["normal_font"]
    ws["Q6"] = "< 2.80"
    ws["Q6"].font = styles["normal_font"]
    ws["R6"] = "DANGER (Deck Submerged)"
    ws["R6"].font = styles["normal_font"]

    # Row 7: 180 ton
    ws["N7"] = 180
    ws["N7"].font = styles["normal_font"]
    ws["N7"].number_format = number_format
    ws["O7"] = 3.64
    ws["O7"].font = styles["normal_font"]
    ws["O7"].number_format = number_format
    ws["P7"] = 0.01
    ws["P7"].font = styles["normal_font"]
    ws["P7"].number_format = number_format
    ws["Q7"] = 2.88
    ws["Q7"].font = styles["normal_font"]
    ws["Q7"].number_format = number_format
    ws["R7"] = "Minimum Limit (Dry Deck)"
    ws["R7"].font = styles["normal_font"]

    # Row 8: 200 ton
    ws["N8"] = 200
    ws["N8"].font = styles["normal_font"]
    ws["N8"].number_format = number_format
    ws["O8"] = 3.6
    ws["O8"].font = styles["normal_font"]
    ws["O8"].number_format = number_format
    ws["P8"] = 0.05
    ws["P8"].font = styles["normal_font"]
    ws["P8"].number_format = number_format
    ws["Q8"] = 2.98
    ws["Q8"].font = styles["normal_font"]
    ws["Q8"].number_format = number_format
    ws["R8"] = "Baseline (Draft Optimized)"
    ws["R8"].font = styles["normal_font"]

    # Row 9: 250 ton (Safe - 빨간색 강조)
    red_font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    yellow_fill = PatternFill(
        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
    )
    ws["N9"] = 250
    ws["N9"].font = red_font
    ws["N9"].number_format = number_format
    ws["O9"] = 3.48
    ws["O9"].font = red_font
    ws["O9"].number_format = number_format
    ws["P9"] = 0.17
    ws["P9"].font = red_font
    ws["P9"].number_format = number_format
    ws["Q9"] = 3.22
    ws["Q9"].font = red_font
    ws["Q9"].number_format = number_format
    ws["R9"] = "Safe"
    ws["R9"].font = red_font
    ws["R9"].fill = yellow_fill

    # Row 10: 300 ton
    ws["N10"] = 300
    ws["N10"].font = styles["normal_font"]
    ws["N10"].number_format = number_format
    ws["O10"] = 3.36
    ws["O10"].font = styles["normal_font"]
    ws["O10"].number_format = number_format
    ws["P10"] = 0.29
    ws["P10"].font = styles["normal_font"]
    ws["P10"].number_format = number_format
    ws["Q10"] = 3.46
    ws["Q10"].font = styles["normal_font"]
    ws["Q10"].number_format = number_format
    ws["R10"] = "Recommended (Rec. Min)"
    ws["R10"].font = styles["normal_font"]

    # Row 11: 350 ~ 500 ton
    ws["N11"] = "350 ~ 500"
    ws["N11"].font = styles["normal_font"]
    ws["O11"] = "< 3.25"
    ws["O11"].font = styles["normal_font"]
    ws["P11"] = "> 0.40"
    ws["P11"].font = styles["normal_font"]
    ws["Q11"] = "3.7 ~ 4.4"
    ws["Q11"].font = styles["normal_font"]
    ws["R11"] = "Very Safe"
    ws["R11"].font = styles["normal_font"]

    # Row 12: 550 ~ 620 ton
    ws["N12"] = "550 ~ 620"
    ws["N12"].font = styles["normal_font"]
    ws["O12"] = "< 2.80"
    ws["O12"].font = styles["normal_font"]
    ws["P12"] = "> 0.86"
    ws["P12"].font = styles["normal_font"]
    ws["Q12"] = "> 4.50"
    ws["Q12"].font = styles["normal_font"]
    ws["R12"] = "Caution (Deep Aft Draft)"
    ws["R12"].font = styles["normal_font"]

    # Row 17: 섹션 제목 추가
    ws["A17"] = (
        "4. Ballast Water Optimization Matrix & Safety Margins(Ballast Setting 250ton)"
    )
    ws["A17"].font = styles["normal_font"]

    # Stage 1-9 Notes를 G6-G15로 배치 (DAS Method Operation)
    # stages 리스트 순서대로 G6부터 시작
    stages_list = [
        "Stage 1",  # Arrival
        "Stage 2",  # TR1 Ramp Start
        "Stage 3",  # TR1 Mid-Ramp
        "Stage 4",  # TR1 On Deck
        "Stage 5",  # TR1 Final Position
        "Stage 5_PreBallast",  # [D-1 Night] Water Supply Complete
        "Stage 6A_Critical (Opt C)",  # [D-Day] TR2 Ramp Entry
        "Stage 6C",  # Final Stowage
        "Stage 7",  # Departure
    ]

    # DAS / AGI 공통 – Pre-ballast + Tug 보조 컨셉 설명
    explanations = {
        "Stage 1": "Arrival condition (lightship). Check initial drafts and trim.",
        "Stage 2": "TR1 roll-on start: 1st axle on ramp. Initial trim change (bow down).",
        "Stage 3": "TR1 mid-ramp: cargo COG on ramp. Progressive bow down, within limits.",
        "Stage 4": "TR1 on deck: full weight transferred to vessel deck. Verify deck & ramp condition.",
        "Stage 5": "TR1 secured at Fr.42 (aft). Drafts back to baseline; ready for D-1 pre-ballasting.",
        "Stage 5_PreBallast": "[D-1] Pre-ballast using shore water. Set intentional trim for TR2 bow moment. No major dynamic ballasting during RORO.",
        "Stage 6A_Critical (Opt C)": "[D-day] TR2 ramp entry under fixed pre-ballast condition. Check critical drafts/trim without additional pumping.",
        "Stage 6C": "Final stowage with both TRs secured. Confirm departure stability based on pre-ballast distribution.",
        "Stage 7": "Departure / reference condition after completion of RORO sequence.",
    }

    for idx, stage_name in enumerate(stages_list, start=0):
        g_row = 6 + idx  # G6부터 시작 (G6-G15)
        if stage_name in explanations:
            ws.cell(row=g_row, column=7).value = explanations[
                stage_name
            ]  # G = column 7
            ws.cell(row=g_row, column=7).font = styles["normal_font"]
            ws.cell(row=g_row, column=7).fill = styles["input_fill"]  # G6-G15 색상 추가

    # Stage table header
    header_row = 18  # 기존 파일과 동일하게 Row 18
    stage_headers = [
        "Stage",
        "W_stage_t",
        "Fr_stage",
        "x_stage_m",
        "TM (t·m)",
        "Trim_cm",
        "FWD_precise_m",  # v4.0: Changed from Trim_m to FWD_precise_m
        "AFT_precise_m",  # v4.0: Changed from Trim_target_cm to AFT_precise_m
        "ΔTM_cm_tm",
        "Lever_arm_m",
        "Ballast_t_calc",
        "Ballast_time_h_calc",
        "Ballast_t",
        "Ballast_time_h",
        "Trim_Check",
        "Dfwd_m",
        "Daft_m",
        "Trim_target_stage_cm",  # sdsdds.md: Q열 추가 (Stage별 타깃, 없으면 B6 사용)
        "FWD_Height_m",
        "AFT_Height_m",
        "Difference",  # T(20) 컬럼: 기존 파일과 동일하게
    ]
    for col, header in enumerate(stage_headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header if header else ""
        if header:
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center_align"]

    # Stage names (DAS Method Operation - 9 stages)
    stages = [
        "Stage 1",  # Arrival
        "Stage 2",  # TR1 Ramp Start
        "Stage 3",  # TR1 Mid-Ramp
        "Stage 4",  # TR1 On Deck
        "Stage 5",  # TR1 Final Position
        "Stage 5_PreBallast",  # [D-1 Night] Water Supply Complete
        "Stage 6A_Critical (Opt C)",  # [D-Day] TR2 Ramp Entry
        "Stage 6C",  # Final Stowage
        "Stage 7",  # Departure
    ]
    # Trim target by stage (cm) – Matching current Excel (screenshot 기준)
    # Q열: Trim_target_stage_cm
    target_trim_by_stage = {
        "Stage 1": 0.0,
        "Stage 2": -96.5,
        "Stage 3": -96.5,
        "Stage 4": -96.5,
        "Stage 5": -89.58,
        "Stage 5_PreBallast": 240.0,
        "Stage 6A_Critical (Opt C)": 0.0,
        "Stage 6C": -96.5,
        "Stage 7": 0.0,
    }

    first_data_row = 19  # header_row가 18이므로 first_data_row는 19
    trim5_row = None

    # ------------------------------------------------------------------
    # [PL 기반] Stage별 W_stage_t & x_stage_m 정의
    #  - W_stage_t: 기존 Aries v4.3 Stage 패턴(배수)은 유지
    #               단, TR 217t → TR+SPMT 271.20t(PL 기준)로 스케일 변경
    #  - 이 값은 "시뮬레이션/설명용"이며, 최종 안전 판단은
    #    Aries Stability Booklet 공식 Stage 값 우선.
    # ------------------------------------------------------------------

    # TR+SPMT 기준 단위 중량 (PL 기준)
    EFFECTIVE_TR_UNIT_T = 271.20

    # Stage 6C Total Mass Opt용 상수
    TOTAL_CARGO_WEIGHT_T = 568.83  # 총 화물 중량 (예시값, 필요시 조정)
    PREBALLAST_T_TARGET = 300.00  # Pre-ballast 목표 중량 (예시값, 필요시 조정)

    # Stage별 배수 (Aries v4.3 패턴 유지, TR+SPMT 기준으로 스케일)
    STAGE_MUL = {
        "Stage 2": 0.2994,  # ≈ 81.24t / 271.20t
        "Stage 3": 0.5068,  # ≈ 137.47t / 271.20t
        "Stage 4": 1.0,  # = 271.20t
        "Stage 5_PreBallast": 2.0,  # = 542.40t
        "Stage 6A_Critical (Opt C)": 4.0,  # = 1,084.80t
    }

    # TR+SPMT 기준으로 환산된 Stage별 W (ton)
    w_stage2 = round(EFFECTIVE_TR_UNIT_T * STAGE_MUL["Stage 2"], 2)  # ≈ 81.24t
    w_stage3 = round(EFFECTIVE_TR_UNIT_T * STAGE_MUL["Stage 3"], 2)  # ≈ 137.47t
    w_stage4 = round(EFFECTIVE_TR_UNIT_T * STAGE_MUL["Stage 4"], 2)  # = 271.20t
    w_stage5_pre = round(
        EFFECTIVE_TR_UNIT_T * STAGE_MUL["Stage 5_PreBallast"], 2
    )  # = 542.40t
    w_stage6 = round(
        EFFECTIVE_TR_UNIT_T * STAGE_MUL["Stage 6A_Critical (Opt C)"], 2
    )  # = 1,084.80t

    # Stage별 Fr_stage 값 정의 (757 TCP Tank Plan 기준)
    # x_stage_m은 fr_to_x(Fr_stage)로 자동 계산
    stage_fr_data = {
        # Stage 1: Lightship – 기준점, 추가 중량 없음
        "Stage 1": {"W": 0.0, "Fr": None, "note": "Empty"},
        # Stage 2/3/4: TR1 + SPMT가 ramp→deck으로 이동 (PL 기반 중량 반영)
        # 가정: 기타 강재(steel mat, stool 등)는 Trim 영향이 상대적으로 작다고 보고
        #       Stage 6C에서만 총중량(TOTAL_CARGO_WEIGHT_T)로 반영.
        # 현재 x 값으로 역산: x_to_fr(-10.0) ≈ 40.15, x_to_fr(-6.85) ≈ 37.00, x_to_fr(-3.85) ≈ 34.00
        # 실제 물리적 위치 확인 필요 (ramp entry/mid/end)
        "Stage 2": {
            "W": w_stage2,
            "Fr": 40.15,
            "note": "SPMT 1st Entry (TR1+SPMT per PL)",
        },
        "Stage 3": {
            "W": w_stage3,
            "Fr": 37.00,
            "note": "SPMT Mid-Ramp (TR1+SPMT per PL)",
        },
        "Stage 4": {
            "W": w_stage4,
            "Fr": 34.00,
            "note": "Break-even (TR1+SPMT full on deck)",
        },
        # Stage 5: TR1만 최종 위치(Fr.42) – 여기서는 Trim용 이벤트 중량 0으로 유지
        # FWB1/2 Center: (Fr 60.5 + Fr 50.5) / 2 = Fr 55.5
        "Stage 5": {"W": 0.0, "Fr": 55.5, "note": "Ballast Only (FWB1+2, Fr 48–65)"},
        # Stage 5_PreBallast: Pre-ballast 완료 상태 (TR1+TR2 등가 × TR+SPMT)
        #  → 542.40t = 2 × 271.20t, 실제 ballast 분배는 별도 시트에서 계산
        # 현재 x = fr_to_x(5.0) → Fr = 5.0 (AFT, 확인 필요)
        "Stage 5_PreBallast": {
            "W": w_stage5_pre,
            "Fr": 5.0,
            "note": "Opt C: TR1 Aft + Forward Ballast (FWB1/2, Fr 48–65, PL-based)",
        },
        # Stage 6A: TR2 Ramp Entry – 두 대 TR+SPMT 기준 + Ballast effect를 등가 중량으로 반영
        #  → 1,084.80t = 4 × 271.20t (기존 868t 패턴을 TR+SPMT 기준으로 확대)
        #  ⚠ 가정: 실제 ΔDisp와 정확히 일치하지 않을 수 있으며, Trim 경향 확인용.
        # Forward ballast tanks center (FWB1/2, Fr 48–65): Fr 55.5
        "Stage 6A_Critical (Opt C)": {
            "W": w_stage6,
            "Fr": 55.5,
            "note": "Opt C: TR2 on Ramp + Forward Ballast (FWB1/2, Fr 48–65, PL-based)",
        },
        # Stage 6C: Final Stowage – 동일 등가중량 유지, 위치만 최종 Frame 기준
        "Stage 6C": {
            "W": w_stage6,
            "Fr": 40.0,
            "note": "Symmetric Final (Both TRs+SPMT, PL-based)",
        },
        # Stage 7: Cargo Off – 기준 조건 복귀
        "Stage 7": {"W": 0.0, "Fr": 30.15, "note": "Cargo Off"},
    }

    # stage_defaults 딕셔너리로 변환 (Fr_stage 기반으로 x_stage_m 자동 계산)
    stage_defaults = {}
    for stage_name, data in stage_fr_data.items():
        fr_val = data["Fr"]
        x_val = fr_to_x(fr_val) if fr_val is not None else 0.0
        stage_defaults[stage_name] = {"W": data["W"], "Fr": fr_val, "x": x_val}

    for idx, stage_name in enumerate(stages, start=0):
        row = first_data_row + idx
        row_str = str(row)

        ws.cell(row=row, column=1, value=stage_name)

        if stage_name in stage_defaults:
            defaults = stage_defaults[stage_name]
            ws.cell(row=row, column=2, value=defaults["W"])  # W_stage_t
            # Fr_stage 컬럼 (column 3)
            if defaults["Fr"] is not None:
                ws.cell(row=row, column=3, value=defaults["Fr"])
            else:
                ws.cell(row=row, column=3, value="")  # Stage 1은 Fr 없음
            # x_stage_m 컬럼 (column 4) - fr_to_x(Fr_stage)로 자동 계산
            # Excel 수식: fr_to_x는 Frame_x_from_mid_m.json 기반 VLOOKUP 또는 직접 계산
            # 직접 계산: x = 30.151 - Fr (757 TCP 기준)
            if defaults["Fr"] is not None:
                # Excel 수식으로 자동 계산 (Fr_stage 컬럼 참조)
                ws.cell(
                    row=row,
                    column=4,
                    value=f'=IF(C{row_str}="", "", 30.151 - C{row_str})',
                )
            else:
                ws.cell(row=row, column=4, value=0.0)  # Stage 1은 x=0
        ws.cell(row=row, column=2).fill = styles["input_fill"]
        ws.cell(row=row, column=2).font = styles["normal_font"]
        ws.cell(row=row, column=2).number_format = number_format
        ws.cell(row=row, column=3).fill = styles["input_fill"]
        ws.cell(row=row, column=3).font = styles["normal_font"]
        ws.cell(row=row, column=3).number_format = number_format
        ws.cell(row=row, column=4).fill = styles["input_fill"]
        ws.cell(row=row, column=4).font = styles["normal_font"]
        ws.cell(row=row, column=4).number_format = number_format

        # TM (t·m) 계산: B * (D - $B$10) (x_stage_m이 이제 column 4)
        ws.cell(row=row, column=5).value = (
            f'=IF(OR(B{row_str}="", D{row_str}="", $B$10=""), "", B{row_str} * (D{row_str} - $B$10))'
        )
        ws.cell(row=row, column=5).font = styles["normal_font"]
        ws.cell(row=row, column=5).number_format = number_format
        # Trim_cm 계산: E / $B$9 (TM이 이제 column 5)
        ws.cell(row=row, column=6).value = (
            f'=IF(OR(E{row_str}="", OR($B$9="", $B$9=0)), "", E{row_str} / $B$9)'
        )
        ws.cell(row=row, column=6).font = styles["normal_font"]
        ws.cell(row=row, column=6).number_format = number_format
        # FWD_precise_m (column 7) - v4.0: Trim_m은 F/100 사용 (Trim_cm이 이제 column 6)
        # 임시 수식, extend_precision_columns에서 LCF 기반 정밀 계산으로 덮어씀
        ws.cell(row=row, column=7).value = f'=IF(F{row_str}="", "", F{row_str} / 100)'
        ws.cell(row=row, column=7).font = styles["normal_font"]
        ws.cell(row=row, column=7).number_format = number_format
        # AFT_precise_m (column 8) - v4.0: 임시 수식, extend_precision_columns에서 덮어씀
        ws.cell(row=row, column=8).value = f'=IF(F{row_str}="", "", F{row_str} / 100)'
        ws.cell(row=row, column=8).font = styles["normal_font"]
        ws.cell(row=row, column=8).number_format = number_format

        if stage_name == "Stage 5":
            trim5_row = row

        # v4.0: G 컬럼이 AFT_precise_m로 변경되므로 target_trim은 제거
        # target_trim은 이제 사용되지 않음 (G가 AFT_precise_m로 계산됨)
        # target_trim = target_trim_by_stage.get(stage_name)
        # if target_trim is not None:
        #     ws.cell(row=row, column=8).value = target_trim
        #     ws.cell(row=row, column=8).fill = styles["input_fill"]

        # H (8): ΔTM_cm_tm - sdsdds.md 가이드: Stage별 타깃(R열) 우선, 없으면 전역 타깃(B8) 사용
        # 의미: (현재 Trim_cm − Trim_target) × MTC = 필요한 Trim 모멘트(cm·t·m)
        # Trim_target = IF(R="", $B$8, R) - R열이 비어있으면 전역 타깃(B8) 사용, 있으면 Stage별 타깃 사용
        ws.cell(row=row, column=9).value = (
            f'=IF($A{row_str}="","",'
            f'(F{row_str} - IF($R{row_str}="",$B$8,$R{row_str})) * $B$9)'
        )
        ws.cell(row=row, column=9).number_format = number_format
        ws.cell(row=row, column=9).font = styles["normal_font"]
        # I (9): Lever_arm_m - BUSHRA verified: X_Ballast_from_AP (52.50) - LCF_from_AP (30.91) = 21.59 m
        # LCF_from_AP = LCF_from_mid (B10) + Lpp/2 (B15/2)
        ws.cell(row=row, column=10).value = (
            f'=IF(OR(ISBLANK($B$14), ISBLANK($B$10), ISBLANK($B$15), ISERROR($B$10)), "", ROUND($B$14 - ($B$10 + $B$15/2), 2))'
        )
        ws.cell(row=row, column=10).font = styles["normal_font"]
        ws.cell(row=row, column=10).number_format = number_format
        # J (10): Ballast_t_calc - zzzzz.md 가이드: ΔTM(I) / Lever_arm(J) = 이론상 필요한 Ballast_t
        # I = column 9 (ΔTM_cm_tm), J = column 10 (Lever_arm_m)
        ws.cell(row=row, column=11).value = (
            f'=IF(OR($A{row_str}="",$J{row_str}="", $J{row_str}=0),"",ROUND(I{row_str} / $J{row_str}, 2))'
        )
        ws.cell(row=row, column=11).number_format = number_format
        ws.cell(row=row, column=11).font = styles["normal_font"]
        # K (11): Ballast_time_h_calc - zzzzz.md 가이드: K(Ballast_t_calc)를 펌프 레이트(B13)로 나눈 시간
        # K = column 11 (Ballast_t_calc)
        ws.cell(row=row, column=12).value = (
            f'=IF(OR(K{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(K{row_str} / $B$13, 2))'
        )
        ws.cell(row=row, column=12).font = styles["normal_font"]
        ws.cell(row=row, column=12).number_format = number_format
        # L (12): Ballast_t - v4.0: G가 FWD_precise_m로 변경되었으므로 Trim_m은 F/100 사용
        ws.cell(row=row, column=13).value = (
            f'=IF(OR(F{row_str}="", OR($B$12="", $B$12=0)), "", ROUND(ABS(F{row_str}/100) * 50 * $B$12, 2))'
        )
        ws.cell(row=row, column=13).font = styles["normal_font"]
        ws.cell(row=row, column=13).number_format = number_format
        # M (13): Ballast_time_h - L(Ballast_t)를 펌프 레이트(B13)로 나눈 시간
        # L = column 13 (Ballast_t)
        ws.cell(row=row, column=14).value = (
            f'=IF(OR(L{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(L{row_str} / $B$13, 2))'
        )
        ws.cell(row=row, column=14).font = styles["normal_font"]
        ws.cell(row=row, column=14).number_format = number_format
        # N (14): Trim_Check - v4.0: G가 FWD_precise_m로 변경되었으므로 Trim_m은 F/100 사용
        ws.cell(row=row, column=15).value = (
            f'=IF(F{row_str}="", "", IF(ABS(F{row_str}/100) <= ($B$15/50), "OK", "EXCESSIVE"))'
        )
        # O (15): Dfwd_m - FWD_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        # extend_precision_columns에서 계산된 FWD_precise_m (Column 7 = G) 참조
        ws.cell(row=row, column=16).value = f'=IF(G{row_str}="", "", G{row_str})'
        ws.cell(row=row, column=16).font = styles["normal_font"]
        ws.cell(row=row, column=16).number_format = number_format
        # P (16): Daft_m - AFT_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        # extend_precision_columns에서 계산된 AFT_precise_m (Column 8 = H) 참조
        ws.cell(row=row, column=17).value = f'=IF(H{row_str}="", "", H{row_str})'
        ws.cell(row=row, column=17).font = styles["normal_font"]
        ws.cell(row=row, column=17).number_format = number_format
        # Q (17): Trim_target_stage_cm - Stage별 타깃, 없으면 전역(B8) 사용
        target_trim = target_trim_by_stage.get(stage_name)
        ws.cell(row=row, column=18).value = target_trim
        ws.cell(row=row, column=18).number_format = number_format
        ws.cell(row=row, column=18).fill = styles["input_fill"]
        ws.cell(row=row, column=18).font = styles["normal_font"]
        # R (18): FWD_Height_m - Fr_stage 컬럼 추가로 한 칸 밀림
        # P = column 16 (Dfwd_m)
        ws.cell(row=row, column=19).value = (
            f'=IF(P{row_str}="", "", $B$11 - P{row_str} + $B$7)'
        )
        ws.cell(row=row, column=19).font = styles["normal_font"]
        ws.cell(row=row, column=19).number_format = number_format
        # S (19): AFT_Height_m - Fr_stage 컬럼 추가로 한 칸 밀림
        # Q = column 17 (Daft_m)
        ws.cell(row=row, column=20).value = (
            f'=IF(Q{row_str}="", "", $B$11 - Q{row_str} + $B$7)'
        )
        ws.cell(row=row, column=20).font = styles["normal_font"]
        ws.cell(row=row, column=20).number_format = number_format

        # T (20): Notes - 제거됨 (G4-G15로 이동)
        # Notes는 더 이상 T(20) 컬럼에 배치하지 않음

    # ===============================================================
    # Stage 6 전용 옵션: "Stage 6C_TotalMassOpt"
    #  - W_stage_t = TOTAL_CARGO_WEIGHT_T + PREBALLAST_T_TARGET
    #  - x_stage_m = 기존 Stage 6C와 동일 (LCG 공유)
    #  - Trim_target_stage_cm = 기존 Stage 6C와 동일(-96.5 cm)
    #  - 목적: "실제 출항 시점 총 하중(화물+Pre-ballast)을 기준으로 한
    #          Stage 6C 시나리오"를 별도 행으로 계산/비교
    # ===============================================================
    # 1) Stage 6C 원래 행(row) 찾기
    try:
        stage6c_index = stages.index("Stage 6C")
    except ValueError:
        stage6c_index = None

    # Stage 6C_TotalMassOpt 행 생성 여부 플래그
    mass_opt_created = False

    if stage6c_index is not None:
        row_6c = first_data_row + stage6c_index  # 기존 Stage 6C 행
        row_6c_str = str(row_6c)

        # 2) 옵션 행을 Stage 7 바로 아래(= Stage 블록 끝 다음 행)에 생성
        mass_opt_row = first_data_row + len(stages)  # Stage 1~7 바로 아래
        mass_opt_row_str = str(mass_opt_row)

        # 3) Stage 6C Total Mass (= 화물 + Pre-ballast) 계산
        stage6c_total_mass_t = round(TOTAL_CARGO_WEIGHT_T + PREBALLAST_T_TARGET, 2)
        # 예시: 568.83 + 300.00 = 868.83 t

        # A: Stage 이름
        ws.cell(row=mass_opt_row, column=1, value="Stage 6C_TotalMassOpt")

        # B: W_stage_t – 총 하중(화물+Pre-ballast)
        c = ws.cell(row=mass_opt_row, column=2)
        c.value = stage6c_total_mass_t
        c.font = styles["normal_font"]
        c.number_format = number_format
        c.fill = styles["input_fill"]  # 사용자가 인지하기 쉽도록 입력색 유지

        # C: Fr_stage – 기존 Stage 6C와 동일한 Fr 참조 (Fr_stage 컬럼)
        c = ws.cell(row=mass_opt_row, column=3)
        c.value = f"=C{row_6c_str}"  # Stage 6C의 Fr_stage 값 참조
        c.font = styles["normal_font"]
        c.number_format = number_format
        c.fill = styles["input_fill"]

        # D: x_stage_m – Fr_stage 기반으로 자동 계산 (일반 Stage와 동일)
        c = ws.cell(row=mass_opt_row, column=4)
        c.value = f"=D{row_6c_str}"  # Stage 6C의 x_stage_m 값 참조 (fr_to_x(Fr_stage))
        c.font = styles["normal_font"]
        c.number_format = number_format
        c.fill = styles["input_fill"]

        # E: TM (t·m) = W * (x - LCF) (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=5)
        c.value = f'=IF(OR(B{mass_opt_row_str}="", D{mass_opt_row_str}="", $B$10=""), "", B{mass_opt_row_str} * (D{mass_opt_row_str} - $B$10))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # F: Trim_cm = TM / MTC (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=6)
        c.value = f'=IF(OR(E{mass_opt_row_str}="", OR($B$9="", $B$9=0)), "", E{mass_opt_row_str} / $B$9)'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # G: FWD_precise_m - LCF 기반 정밀 Forward Draft (일반 Stage와 동일한 수식)
        # MD 파일 공식: Dfwd = Tmean - Trim_m * (1 - LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        c = ws.cell(row=mass_opt_row, column=7)
        c.value = (
            f'=IF($A{mass_opt_row_str}="", "", '
            f"$B$6 - (F{mass_opt_row_str}/100) * (0.5 - Calc!$E$41 / Calc!$E$40))"
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # H: AFT_precise_m - LCF 기반 정밀 Aft Draft (일반 Stage와 동일한 수식)
        # MD 파일 공식: Daft = Tmean + Trim_m * (LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        c = ws.cell(row=mass_opt_row, column=8)
        c.value = (
            f'=IF($A{mass_opt_row_str}="", "", '
            f"$B$6 + (F{mass_opt_row_str}/100) * (Calc!$E$41 / Calc!$E$40 + 0.5))"
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # I: ΔTM_cm_tm = (Trim_cm - Trim_target) * MTC (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=9)
        c.value = (
            f'=IF($A{mass_opt_row_str}="","",'
            f'(F{mass_opt_row_str} - IF($R{mass_opt_row_str}="",$B$8,$R{mass_opt_row_str})) * $B$9)'
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # J: Lever_arm_m - BUSHRA verified: X_Ballast_from_AP (52.50) - LCF_from_AP (30.91) = 21.59 m
        # LCF_from_AP = LCF_from_mid (B10) + Lpp/2 (B15/2)
        c = ws.cell(row=mass_opt_row, column=10)
        c.value = f'=IF(OR(ISBLANK($B$14), ISBLANK($B$10), ISBLANK($B$15), ISERROR($B$10)), "", ROUND($B$14 - ($B$10 + $B$15/2), 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # K: Ballast_t_calc = ΔTM / Lever_arm (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=11)
        c.value = f'=IF(OR($A{mass_opt_row_str}="",$J{mass_opt_row_str}="", $J{mass_opt_row_str}=0),"",ROUND(I{mass_opt_row_str} / $J{mass_opt_row_str}, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # L: Ballast_time_h_calc = Ballast_t_calc / pump_rate (B13) (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=12)
        c.value = f'=IF(OR(K{mass_opt_row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(K{mass_opt_row_str} / $B$13, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # M: Ballast_t (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=13)
        c.value = f'=IF(OR(F{mass_opt_row_str}="", OR($B$12="", $B$12=0)), "", ROUND(ABS(F{mass_opt_row_str}/100) * 50 * $B$12, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # N: Ballast_time_h (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=14)
        c.value = f'=IF(OR(M{mass_opt_row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(M{mass_opt_row_str} / $B$13, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # O: Trim_Check – Stage 6C와 동일한 로직 사용 (Trim_m 절대값이 $B$15/50 이하인지 체크)
        c = ws.cell(row=mass_opt_row, column=15)
        c.value = f'=IF(F{mass_opt_row_str}="", "", IF(ABS(F{mass_opt_row_str}/100) <= ($B$15/50), "OK", "EXCESSIVE"))'
        c.font = styles["normal_font"]

        # P: Dfwd_m - FWD_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        c = ws.cell(row=mass_opt_row, column=16)
        c.value = f'=IF(G{mass_opt_row_str}="", "", G{mass_opt_row_str})'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # Q: Daft_m - AFT_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        c = ws.cell(row=mass_opt_row, column=17)
        c.value = f'=IF(H{mass_opt_row_str}="", "", H{mass_opt_row_str})'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # R: Trim_target_stage_cm – Stage 6C와 동일(-96.5 cm)로 고정
        c = ws.cell(row=mass_opt_row, column=18)
        c.value = -96.5
        c.font = styles["normal_font"]
        c.number_format = number_format
        c.fill = styles["input_fill"]

        # S: FWD_Height_m (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=19)
        c.value = f'=IF(P{mass_opt_row_str}="", "", $B$11 - P{mass_opt_row_str} + $B$7)'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # T: AFT_Height_m (일반 Stage와 동일한 수식)
        c = ws.cell(row=mass_opt_row, column=20)
        c.value = f'=IF(Q{mass_opt_row_str}="", "", $B$11 - Q{mass_opt_row_str} + $B$7)'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # Captain Req 컬럼 추가 (U-AA) - extend_roro_captain_req에서 처리하지 않으므로 직접 추가
        # U (21): GM - Tmean (O와 P의 평균) 기반으로 Hydro_Table에서 조회
        c = ws.cell(row=mass_opt_row, column=21)
        c.value = f'=IF(O{mass_opt_row_str}="", "", VLOOKUP(AVERAGE(O{mass_opt_row_str},P{mass_opt_row_str}), Hydro_Table!$B:$D, 3, 1))'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # V (22): Fwd Draft copy
        c = ws.cell(row=mass_opt_row, column=22)
        c.value = f"=O{mass_opt_row_str}"
        c.number_format = number_format
        c.font = styles["normal_font"]

        # W (23): Check vs 2.70m
        c = ws.cell(row=mass_opt_row, column=23)
        c.value = f'=IF(V{mass_opt_row_str}="", "", IF(V{mass_opt_row_str}<=Calc!$E$18, "OK", "NG"))'
        c.font = styles["normal_font"]

        # X (24): Ballast Qty copy
        c = ws.cell(row=mass_opt_row, column=24)
        c.value = f"=K{mass_opt_row_str}"
        c.number_format = number_format
        c.font = styles["normal_font"]

        # Y (25): Timing - 비어있음 (사용자 입력)

        # Z (26): Physical Freeboard
        c = ws.cell(row=mass_opt_row, column=26)
        c.value = f'=IF(P{mass_opt_row_str}="", "", $B$11 - P{mass_opt_row_str})'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AA (27): Clearance Check vs 0.28m
        c = ws.cell(row=mass_opt_row, column=27)
        c.value = f'=IF(Z{mass_opt_row_str}="", "", IF(Z{mass_opt_row_str}>=Calc!$E$20, "OK", "<0.28m CHECK"))'
        c.font = styles["normal_font"]

        # AB (28): GM_calc - U(21) GM 값 복사
        c = ws.cell(row=mass_opt_row, column=28)
        c.value = f"=U{mass_opt_row_str}"
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AC (29): GM Check vs 1.50m
        c = ws.cell(row=mass_opt_row, column=29)
        c.value = f'=IF(AB{mass_opt_row_str}="", "", IF(AB{mass_opt_row_str}>=Calc!$E$21, "OK", "NG"))'
        c.font = styles["normal_font"]

        # AD (30): Total Displacement - 패치 v20251122 최소: Lightship 950.00t + Ballast(L) + Cargo(B)
        c = ws.cell(row=mass_opt_row, column=30)
        c.value = f'=IF(A{mass_opt_row_str}="", "", 950.00 + L{mass_opt_row_str} + B{mass_opt_row_str})'
        c.number_format = number_format
        c.font = styles["normal_font"]
        c.fill = styles["input_fill"]

        # AE (31): Vent_Time_h
        c = ws.cell(row=mass_opt_row, column=31)
        c.value = f'=IF(X{mass_opt_row_str}="", "", IF(X{mass_opt_row_str}=0, "", X{mass_opt_row_str} / Calc!$E$29))'
        c.number_format = number_format
        c.font = styles["normal_font"]

        # Structural 컬럼 추가 (AR-AS) - extend_roro_structural_opt1에서 처리하지 않으므로 직접 추가
        # AR (44): Heel_deg - 패치 v20251122 최소: 전체 Δ(AD) 사용
        c = ws.cell(row=mass_opt_row, column=44)
        c.value = (
            f'=IF(OR($A{mass_opt_row_str}="", U{mass_opt_row_str}="", U{mass_opt_row_str}=0, AD{mass_opt_row_str}="", AD{mass_opt_row_str}=0), "", '
            f"DEGREES((B{mass_opt_row_str} * Calc!$E$43) / (AD{mass_opt_row_str} * U{mass_opt_row_str})))"
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

        # AS (45): GM_eff_m - 패치 v20251122 최소: 전체 Δ(AD) 사용
        fse_value = 0  # Simplified
        c = ws.cell(row=mass_opt_row, column=45)
        c.value = (
            f'=IF(OR($A{mass_opt_row_str}="", U{mass_opt_row_str}="", AD{mass_opt_row_str}="", AD{mass_opt_row_str}=0), "", '
            f"U{mass_opt_row_str} - {fse_value} / AD{mass_opt_row_str})"
        )
        c.number_format = number_format
        c.font = styles["normal_font"]

        mass_opt_created = True

    # =====================================================================
    # 5. Optional Tuning Stages (Scenario Options for Pre-Ballast / 6A / 6C)
    #    - Stage 1~7 아래에 옵션 시나리오 행을 자동 생성
    #    - W_stage_t, x_stage_m, Trim_target_stage_cm(Q열)은 사용자가 입력
    # =====================================================================

    option_stage_names = [
        "Stage 5_PreBallast_Opt1",
        "Stage 5_PreBallast_Opt2",
        "Stage 6A_Critical_Opt1",
        "Stage 6A_Critical_Opt2",
        "Stage 6C_Opt1",
        "Stage 6C_Opt2",
    ]

    # 기존 Stage 데이터가 시작하는 행 + Stage 개수 + Stage 6C_TotalMassOpt(1개) 바로 아래부터 옵션 행 시작
    option_start_row = first_data_row + len(stages) + (1 if mass_opt_created else 0)

    for idx, opt_name in enumerate(option_stage_names):
        row = option_start_row + idx
        row_str = str(row)

        # A: Stage 이름
        ws.cell(row=row, column=1, value=opt_name)

        # B, C, D: W_stage_t / Fr_stage / x_stage_m  → 사용자가 직접 입력 (노랑 인풋셀)
        # Fr_stage는 선택 입력 (비워두면 x_stage_m만 사용)
        for col in (2, 3, 4):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = styles["input_fill"]
            cell.font = styles["normal_font"]
            cell.number_format = number_format

        # E: TM (t·m) = W_stage_t * (x_stage_m - LCF) (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=5)
        c.value = f'=IF(OR(B{row_str}="", D{row_str}="", $B$10=""), "", B{row_str} * (D{row_str} - $B$10))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # F: Trim_cm = TM / MTC (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=6)
        c.value = f'=IF(OR(E{row_str}="", OR($B$9="", $B$9=0)), "", E{row_str} / $B$9)'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # G: FWD_precise_m - LCF 기반 정밀 Forward Draft (일반 Stage와 동일한 수식)
        # MD 파일 공식: Dfwd = Tmean - Trim_m * (1 - LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        c = ws.cell(row=row, column=7)
        c.value = (
            f'=IF($A{row_str}="", "", '
            f"$B$6 - (F{row_str}/100) * (0.5 - Calc!$E$41 / Calc!$E$40))"
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # H: AFT_precise_m - LCF 기반 정밀 Aft Draft (일반 Stage와 동일한 수식)
        # MD 파일 공식: Daft = Tmean + Trim_m * (LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        c = ws.cell(row=row, column=8)
        c.value = (
            f'=IF($A{row_str}="", "", '
            f"$B$6 + (F{row_str}/100) * (Calc!$E$41 / Calc!$E$40 + 0.5))"
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # I: ΔTM_cm_tm = (Trim_cm - Trim_target) * MTC (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=9)
        c.value = (
            f'=IF($A{row_str}="","",'
            f'(F{row_str} - IF($R{row_str}="",$B$8,$R{row_str})) * $B$9)'
        )
        c.font = styles["normal_font"]
        c.number_format = number_format

        # J: Lever_arm_m - BUSHRA verified: X_Ballast_from_AP (52.50) - LCF_from_AP (30.91) = 21.59 m
        # LCF_from_AP = LCF_from_mid (B10) + Lpp/2 (B15/2)
        c = ws.cell(row=row, column=10)
        c.value = f'=IF(OR(ISBLANK($B$14), ISBLANK($B$10), ISBLANK($B$15), ISERROR($B$10)), "", ROUND($B$14 - ($B$10 + $B$15/2), 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # K: Ballast_t_calc = ΔTM / Lever_arm (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=11)
        c.value = f'=IF(OR($A{row_str}="",$J{row_str}="", $J{row_str}=0),"",ROUND(I{row_str} / $J{row_str}, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # L: Ballast_time_h_calc = Ballast_t_calc / pump_rate (B13) (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=12)
        c.value = f'=IF(OR(K{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(K{row_str} / $B$13, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # M: Ballast_t (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=13)
        c.value = f'=IF(OR(F{row_str}="", OR($B$12="", $B$12=0)), "", ROUND(ABS(F{row_str}/100) * 50 * $B$12, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # N: Ballast_time_h (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=14)
        c.value = f'=IF(OR(M{row_str}="", $B$13="", $B$13=0, ISERROR($B$13)), "", ROUND(M{row_str} / $B$13, 2))'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # O: Trim_Check – 일반 Stage와 동일한 로직 사용 (Trim_m 절대값이 $B$15/50 이하인지 체크)
        c = ws.cell(row=row, column=15)
        c.value = f'=IF(F{row_str}="", "", IF(ABS(F{row_str}/100) <= ($B$15/50), "OK", "EXCESSIVE"))'
        c.font = styles["normal_font"]

        # P: Dfwd_m - FWD_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        c = ws.cell(row=row, column=16)
        c.value = f'=IF(G{row_str}="", "", G{row_str})'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # Q: Daft_m - AFT_precise_m 참조 (LCF 기반 정밀 계산 결과 사용)
        c = ws.cell(row=row, column=17)
        c.value = f'=IF(H{row_str}="", "", H{row_str})'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # R: Trim_target_stage_cm – 옵션별 목표 Trim (사용자 입력, 노랑 인풋셀)
        c = ws.cell(row=row, column=18)
        c.value = None
        c.fill = styles["input_fill"]
        c.font = styles["normal_font"]
        c.number_format = number_format

        # S: FWD_Height_m (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=19)
        c.value = f'=IF(P{row_str}="", "", $B$11 - P{row_str} + $B$7)'
        c.font = styles["normal_font"]
        c.number_format = number_format

        # T: AFT_Height_m (일반 Stage와 동일한 수식)
        c = ws.cell(row=row, column=20)
        c.value = f'=IF(Q{row_str}="", "", $B$11 - Q{row_str} + $B$7)'
        c.font = styles["normal_font"]
        c.number_format = number_format

    # F6:F15에 Stage 이름 복사 (A19:A27에서) - 기존 파일과 동일하게
    # F 컬럼은 이미 Row 6-15에 설정되어 있으므로 복사 불필요 (이미 설정됨)

    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names["MTC"] = DefinedName(
        "MTC", attr_text="'RORO_Stage_Scenarios'!$B$9"
    )
    wb.defined_names["LCF"] = DefinedName(
        "LCF", attr_text="'RORO_Stage_Scenarios'!$B$10"
    )
    wb.defined_names["PumpRate"] = DefinedName(
        "PumpRate", attr_text="'RORO_Stage_Scenarios'!$B$13"
    )
    wb.defined_names["X_Ballast"] = DefinedName(
        "X_Ballast", attr_text="'RORO_Stage_Scenarios'!$B$14"
    )

    if trim5_row is not None:
        wb.defined_names["TRIM5_CM"] = DefinedName(
            "TRIM5_CM", attr_text=f"'RORO_Stage_Scenarios'!$E${trim5_row}"
        )

    widths = {
        "A": 18,
        "B": 11,
        "C": 11,
        "D": 13,
        "E": 10,
        "F": 9,
        "G": 13,
        "H": 12,
        "I": 11,
        "J": 13,
        "K": 15,
        "L": 11,
        "M": 15,
        "N": 13,
        "O": 10,
        "P": 10,
        "Q": 13,
        "R": 13,
        "S": 40,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "G2"  # 기존 파일과 동일하게 G2

    # Opt C Stage 정보 출력 (DAS Method v4.3 Final Optimized)
    opt_c_data = build_opt_c_stage()
    print(
        f"  [INFO] Opt C Updated (DAS Method v4.3 Final): W={opt_c_data['weight_t']:.1f}t (Cargo 560t + Ballast 250t), LCG={opt_c_data['x_from_mid_m']:.2f}m (Even Keel 근접)"
    )

    print("  [OK] RORO_Stage_Scenarios sheet created")

    # Optional Tuning Stages의 행 범위 계산
    option_start_row = first_data_row + len(stages) + (1 if mass_opt_created else 0)
    option_end_row = option_start_row + len(option_stage_names)
    total_rows = (
        option_end_row - first_data_row
    )  # 전체 행 수 (일반 Stage + Stage 6C_TotalMassOpt + Optional Tuning Stages)

    return stages, first_data_row, total_rows


# ============================================================================
# CAPTAIN REPORT Sheet Creation (NEW)
# ============================================================================


def create_captain_report_sheet(wb, stages, first_data_row):
    """
    Creates the OPERATION SUMMARY sheet summary.
    Updated for v4.3: DAS Method & Critical Checkpoints Highlight.
    """
    if "OPERATION SUMMARY" in wb.sheetnames:
        wb.remove(wb["OPERATION SUMMARY"])

    ws = wb.create_sheet("OPERATION SUMMARY")
    styles = get_styles()

    # --- 1. Report Title & Header ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "LCT BUSHRA – OPERATION SUMMARY (DAS METHOD)"
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["center_align"]
    ws["A1"].fill = styles["header_fill"]

    # --- 2. Critical Limits (기준값) ---
    headers_limit = ["PARAMETER", "LIMIT", "UNIT", "REMARK"]
    limits_data = [
        ("Summer Draft Max", 2.70, "m", "Operational Limit (Harbour: Check Depth)"),
        ("Min Freeboard", 0.28, "m", "Linkspan Connection Safety"),
        ("Min GM", 1.50, "m", "Stability Requirement"),
        ("Max Ramp Angle", 6.0, "deg", "SPMT Climbing Limit"),
    ]

    # Write Limit Table
    ws.merge_cells("A3:D3")
    ws["A3"] = "1. OPERATIONAL LIMITS"
    ws["A3"].font = styles["header_font"]
    ws["A3"].fill = styles["structure_fill"]  # Orange

    for col, text in enumerate(headers_limit, 1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    for i, (param, val, unit, rem) in enumerate(limits_data, 5):
        ws.cell(row=i, column=1, value=param).font = styles["normal_font"]
        ws.cell(row=i, column=2, value=val).font = styles["normal_font"]
        ws.cell(row=i, column=2).number_format = "0.00"
        ws.cell(row=i, column=2).fill = styles["input_fill"]
        ws.cell(row=i, column=3, value=unit).font = styles["normal_font"]
        ws.cell(row=i, column=4, value=rem).font = styles["normal_font"]

        # 테두리 적용
        for c in range(1, 5):
            ws.cell(row=i, column=c).border = Border(bottom=styles["thin_border"])

    # --- 3. Stage Summary Table ---
    table_start_row = 10
    ws.merge_cells(f"A{table_start_row-1}:J{table_start_row-1}")
    ws[f"A{table_start_row-1}"] = "2. STAGE-BY-STAGE SAFETY CHECK"
    ws[f"A{table_start_row-1}"].font = styles["header_font"]
    ws[f"A{table_start_row-1}"].fill = styles["structure_fill"]

    headers_table = [
        "Stage",
        "Condition",
        "Trim (m)",
        "Fwd Draft",
        "Aft Draft",
        "Draft Check",
        "Freeboard",
        "Deck Check",
        "Action / Note",
    ]

    # Write Headers
    for col, text in enumerate(headers_table, 1):
        cell = ws.cell(row=table_start_row, column=col, value=text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]

    # Write Data Rows (Link to RORO Sheet)
    # Mapping from RORO sheet columns:
    # Stage(A), Trim_m(F), Dfwd(O), Daft(P), Phys_Freeboard_m(Z)

    roro_sheet = "RORO_Stage_Scenarios"

    for i, stage_name in enumerate(stages):
        r_rept = table_start_row + 1 + i
        r_roro = first_data_row + i
        r_str = str(r_roro)

        # A: Stage Name
        ws.cell(row=r_rept, column=1, value=f"='{roro_sheet}'!A{r_str}")

        # B: Condition (Custom Logic based on Stage Name)
        ws.cell(
            row=r_rept,
            column=2,
            value=f'=IF(ISNUMBER(SEARCH("PreBallast",A{r_rept})),"PRE-BALLAST",IF(ISNUMBER(SEARCH("Critical",A{r_rept})),"CRITICAL","NORMAL"))',
        )

        # C: Trim
        ws.cell(row=r_rept, column=3, value=f"='{roro_sheet}'!F{r_str}")
        ws.cell(row=r_rept, column=3).number_format = "0.00"

        # D: Fwd Draft
        ws.cell(row=r_rept, column=4, value=f"='{roro_sheet}'!O{r_str}")
        ws.cell(row=r_rept, column=4).number_format = "0.00"

        # E: Aft Draft
        ws.cell(row=r_rept, column=5, value=f"='{roro_sheet}'!P{r_str}")
        ws.cell(row=r_rept, column=5).number_format = "0.00"

        # F: Draft Check (Max Draft vs Limit 2.70)
        # Logic: If PreBallast -> Check Depth (Warning), Else -> Check 2.70
        check_formula = (
            f'=IF(B{r_rept}="PRE-BALLAST", "CHECK DEPTH", '
            f'IF(MAX(D{r_rept},E{r_rept})<=$B$5, "OK", "OVER DRAFT"))'
        )
        ws.cell(row=r_rept, column=6, value=check_formula)

        # G: Freeboard (Phys_Freeboard_m from RORO sheet Z column)
        ws.cell(row=r_rept, column=7, value=f"='{roro_sheet}'!Z{r_str}")
        ws.cell(row=r_rept, column=7).number_format = "0.00"

        # H: Deck Check (Freeboard vs 0.28)
        ws.cell(
            row=r_rept, column=8, value=f'=IF(G{r_rept}>=$B$6, "OK", "SUBMERGED/LOW")'
        )

        # I: Action / Note (From RORO G column - Explanation)
        ws.cell(row=r_rept, column=9, value=f"='{roro_sheet}'!G{6 + i}")
        ws.cell(row=r_rept, column=9).alignment = styles["left_align"]

        # Formatting (Colors for Critical Stages)
        # Conditional Formatting logic is hard in openpyxl, so we use static logic if possible or just row styling
        if "PreBallast" in stage_name or "Critical" in stage_name:
            for c in range(1, 10):
                ws.cell(row=r_rept, column=c).fill = styles[
                    "input_fill"
                ]  # Yellow highlight

        # Borders
        for c in range(1, 10):
            ws.cell(row=r_rept, column=c).border = Border(bottom=styles["thin_border"])

    # Column Widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 50

    print("  [OK] OPERATION SUMMARY sheet updated (DAS Method Specialized)")


# ============================================================================
# Additional Sheet Creation Functions
# ============================================================================


def extend_roro_captain_req(ws, first_data_row, num_stages):
    """RORO 시트에 Captain Req 컬럼 추가 (Col U부터) - sdsdds.md: Q열 추가로 한 칸 밀림"""
    styles = get_styles()

    # 패치: 헤더 이름 변경 (Linkspan_Freeboard_m → Phys_Freeboard_m)
    captain_cols = [
        "GM(m)",
        "Fwd Draft(m)",
        "vs 2.70m",
        "De-ballast Qty(t)",
        "Timing",
        "Phys_Freeboard_m",
        "Clearance_Check",
        "GM_calc",
        "GM_Check",
        "Disp_total_t",  # 패치 v20251122 최소: Total Displacement
        "Vent_Time_h",
    ]
    start_col = 21  # sdsdds.md: Q열 추가로 T(20) → U(21)로 이동

    for i, h in enumerate(captain_cols):
        col = start_col + i
        header_row = first_data_row - 1  # Row 17 (first_data_row=18이므로)
        cell = ws.cell(row=header_row, column=col)
        cell.value = h
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    for row in range(first_data_row, first_data_row + num_stages):
        row_str = str(row)

        # U (21): GM - sdsdds.md: Q열 추가로 T(20) → U(21)로 이동
        # 패치: Tmean (O와 P의 평균) 기반으로 Hydro_Table에서 조회
        # Hydro Table: A=Disp, B=Tmean, C=Trim, D=GM
        # 숫자 포맷 통일
        number_format = "#,##0.00"

        ws.cell(row=row, column=21).value = (
            f'=IF(O{row_str}="", "", VLOOKUP(AVERAGE(O{row_str},P{row_str}), Hydro_Table!$B:$D, 3, 1))'
        )
        ws.cell(row=row, column=21).number_format = number_format
        ws.cell(row=row, column=21).font = styles["normal_font"]

        # V (22): Fwd Draft copy - sdsdds.md: Q열 추가로 U(21) → V(22)로 이동
        ws.cell(row=row, column=22).value = f"=O{row_str}"
        ws.cell(row=row, column=22).number_format = number_format
        ws.cell(row=row, column=22).font = styles["normal_font"]

        # W (23): Check vs 2.70m (Calc!E18) - sdsdds.md: Q열 추가로 V(22) → W(23)로 이동
        ws.cell(row=row, column=23).value = (
            f'=IF(V{row_str}="", "", IF(V{row_str}<=Calc!$E$18, "OK", "NG"))'
        )
        ws.cell(row=row, column=23).font = styles["normal_font"]

        # X (24): Ballast Qty copy - sdsdds.md: Q열 추가로 W(23) → X(24)로 이동
        # K = column 11 (Ballast_t_calc)
        ws.cell(row=row, column=24).value = f"=K{row_str}"
        ws.cell(row=row, column=24).number_format = number_format
        ws.cell(row=row, column=24).font = styles["normal_font"]

        # Y (25): Timing - 비어있음 (사용자 입력) - sdsdds.md: Q열 추가로 X(24) → Y(25)로 이동

        # Z (26): Physical Freeboard - sdsdds.md: Q열 추가로 Y(25) → Z(26)로 이동
        # 패치: 물리적 freeboard 계산 (Depth - Dfwd, tide 없이)
        # Fr_stage 컬럼 추가로 Dfwd_m이 Column 16 (P)로 이동
        ws.cell(row=row, column=26).value = (
            f'=IF(P{row_str}="", "", $B$11 - P{row_str})'
        )
        ws.cell(row=row, column=26).number_format = number_format
        ws.cell(row=row, column=26).font = styles["normal_font"]

        # AA (27): Clearance Check vs 0.28m (Calc!E20) - sdsdds.md: Q열 추가로 Z(26) → AA(27)로 이동
        ws.cell(row=row, column=27).value = (
            f'=IF(Z{row_str}="", "", IF(Z{row_str}>=Calc!$E$20, "OK", "<0.28m CHECK"))'
        )
        ws.cell(row=row, column=27).font = styles["normal_font"]

        # AB (28): GM copy - sdsdds.md: Q열 추가로 AA(27) → AB(28)로 이동
        ws.cell(row=row, column=28).value = f"=U{row_str}"
        ws.cell(row=row, column=28).number_format = number_format
        ws.cell(row=row, column=28).font = styles["normal_font"]

        # AC (29): GM Check vs 1.50m (Calc!E21) - sdsdds.md: Q열 추가로 AB(28) → AC(29)로 이동
        ws.cell(row=row, column=29).value = (
            f'=IF(AB{row_str}="", "", IF(AB{row_str}>=Calc!$E$21, "OK", "NG"))'
        )
        ws.cell(row=row, column=29).font = styles["normal_font"]

        # AD (30): Total Displacement - 패치 v20251122 최소: Lightship 950.00t + Ballast(L) + Cargo(B)
        ws.cell(row=row, column=30).value = (
            f'=IF(A{row_str}="", "", 950.00 + L{row_str} + B{row_str})'
        )
        ws.cell(row=row, column=30).number_format = number_format
        ws.cell(row=row, column=30).font = styles["normal_font"]
        ws.cell(row=row, column=30).fill = styles["input_fill"]

        # AE (31): Vent Time - sdsdds.md: Q열 추가로 AD(30) → AE(31)로 이동
        ws.cell(row=row, column=31).value = f'=IF(X{row_str}>0, X{row_str}/45, "-")'
        ws.cell(row=row, column=31).number_format = number_format
        ws.cell(row=row, column=31).font = styles["normal_font"]

    ws.column_dimensions["U"].width = 12
    ws.column_dimensions["V"].width = 12
    ws.column_dimensions["W"].width = 12
    ws.column_dimensions["X"].width = 15
    ws.column_dimensions["Y"].width = 12
    ws.column_dimensions["Z"].width = 18
    ws.column_dimensions["AA"].width = 15
    ws.column_dimensions["AB"].width = 12
    ws.column_dimensions["AC"].width = 12
    ws.column_dimensions["AD"].width = 12
    ws.column_dimensions["AE"].width = 12

    print("  [OK] Captain Req columns added to RORO_Stage_Scenarios sheet (Patched)")


def extend_roro_structural_opt1(ws, first_data_row, num_stages):
    """RORO 시트에 Structural Strength 및 Option 1 Ballast Fix Check 컬럼 추가 (Col AE-AO)"""
    styles = get_styles()
    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # Structural Strength 컬럼 (AE-AJ)
    structural_cols = [
        "Share_Load_t",
        "Share_Check",
        "Hinge_Rx_t",
        "Rx_Check",
        "Deck_Press_t/m²",
        "Press_Check",
    ]

    # Dynamic Load Case B/C 컬럼 (AK-AL) - v4.0: Replace Option 1 with Dynamic Load Case
    dynamic_load_cols = [
        "Load_Case_B_t",  # AK (37): Dynamic factor applied
        "Load_Case_C_t",  # AL (38): Case B + 0.2g braking
    ]

    # Option 1 Ballast Fix Check 컬럼 (AM-AP) - v4.0: Moved after Dynamic Load Case, Fix_Status removed (replaced by Heel/FSE)
    opt1_cols = [
        "ΔTM_needed_cm·tm",
        "Ballast_req_t",
        "Ballast_gap_t",
        "Time_Add_h",
    ]

    # Heel/FSE 컬럼 (AQ-AR) - v4.0: Replace Fix_Status with Heel/FSE
    # Note: AQ (43) was Fix_Status, now becomes Heel_deg
    # AR (44) is new GM_eff_m column
    heel_fse_cols = [
        "Heel_deg",  # AQ (43): Heel angle in degrees (replaces Fix_Status)
        "GM_eff_m",  # AR (44): Effective GM accounting for FSE
    ]

    # Ramp Angle & Pin Stress 컬럼 (AT-AW) - sdsdds.md: Q열 추가로 AS(45) → AT(46)로 이동
    ramp_stress_cols = [
        "Ramp_Angle_deg",  # AT (46)
        "Ramp_Angle_Check",  # AU (47)
        "Pin_Stress_N/mm²",  # AV (48)
        "Von_Mises_Check",  # AW (49)
    ]

    # Opt C / High Tide 관련 컬럼 (AX-AY)
    opt_c_tide_cols = [
        "Required_Tide_m",  # AX (50)
        "Tide_OK",  # AY (51)
    ]

    all_cols = (
        structural_cols
        + dynamic_load_cols
        + opt1_cols
        + heel_fse_cols
        + ramp_stress_cols
        + opt_c_tide_cols
    )
    start_col = 32  # sdsdds.md: Q열 추가로 AE(31) → AF(32)로 이동

    for i, h in enumerate(all_cols):
        col = start_col + i
        header_row = first_data_row - 1  # Row 17 (first_data_row=18이므로)
        cell = ws.cell(row=header_row, column=col)
        cell.value = h
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )
        # Structural 컬럼은 주황색, Dynamic Load는 주황색, Option 1 컬럼은 보라색, Heel/FSE는 주황색, Ramp/Stress 컬럼은 주황색, Opt C Tide 컬럼은 보라색
        if i < len(structural_cols):
            cell.fill = styles["structure_fill"]
        elif i < len(structural_cols) + len(dynamic_load_cols):
            cell.fill = styles[
                "structure_fill"
            ]  # Dynamic Load Case uses structure fill
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(opt1_cols):
            cell.fill = styles["opt1_fill"]
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(opt1_cols) + len(
            heel_fse_cols
        ):
            cell.fill = styles["structure_fill"]  # Heel/FSE uses structure fill
        elif i < len(structural_cols) + len(dynamic_load_cols) + len(opt1_cols) + len(
            heel_fse_cols
        ) + len(ramp_stress_cols):
            cell.fill = styles[
                "structure_fill"
            ]  # Ramp/Stress columns use structure fill
        else:
            cell.fill = styles["opt1_fill"]  # Opt C Tide columns use opt1 fill

    for row in range(first_data_row, first_data_row + num_stages):
        row_str = str(row)

        # Structural Strength 컬럼 (AF-AK) - sdsdds.md: Q열 추가로 AE(31) → AF(32)로 이동
        # AF (32): Share_Load_t - 입력
        ws.cell(row=row, column=32).fill = styles["input_fill"]
        ws.cell(row=row, column=32).font = styles["normal_font"]

        # AG (33): Share_Check - sdsdds.md: Q열 추가로 AF(32) → AG(33)로 이동
        ws.cell(row=row, column=33).value = (
            f'=IF(AF{row_str}="", "", IF(AF{row_str}<=Calc!$E$24, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=33).font = styles["normal_font"]

        # AH (34): Hinge_Rx_t - sdsdds.md: Q열 추가로 AG(33) → AH(34)로 이동
        # 자동 계산 (Ramp 자중 45 t + share 54.5%)
        ws.cell(row=row, column=34).value = (
            f'=IF(AF{row_str}="", 45, 45 + AF{row_str} * 0.545)'
        )
        ws.cell(row=row, column=34).number_format = number_format
        ws.cell(row=row, column=34).font = styles["normal_font"]

        # AI (35): Rx_Check - sdsdds.md: Q열 추가로 AH(34) → AI(35)로 이동
        ws.cell(row=row, column=35).value = (
            f'=IF(AH{row_str}="", "", IF(AH{row_str}<=Calc!$E$37, "OK", "NG"))'
        )
        ws.cell(row=row, column=35).font = styles["normal_font"]

        # AJ (36): Deck_Press_t/m² - sdsdds.md: Q열 추가로 AI(35) → AJ(36)로 이동
        ws.cell(row=row, column=36).value = (
            f'=IF(AF{row_str}="", "", AF{row_str}/Calc!$E$26)'
        )
        ws.cell(row=row, column=36).number_format = number_format
        ws.cell(row=row, column=36).font = styles["normal_font"]

        # AK (37): Press_Check - sdsdds.md: Q열 추가로 AJ(36) → AK(37)로 이동
        ws.cell(row=row, column=37).value = (
            f'=IF(AJ{row_str}="", "", IF(AJ{row_str}<=Calc!$E$25, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=37).font = styles["normal_font"]

        # Dynamic Load Case B/C 컬럼 (AL-AM) - sdsdds.md: Q열 추가로 AK(37) → AL(38)로 이동
        # AL (38): Load_Case_B_t - Dynamic factor applied to Share_Load
        ws.cell(row=row, column=38).value = (
            f'=IF(AF{row_str}="", "", AF{row_str} * Calc!$E$42)'
        )
        ws.cell(row=row, column=38).number_format = number_format
        ws.cell(row=row, column=38).font = styles["normal_font"]

        # AM (39): Load_Case_C_t - Case B + 0.2g braking (0.2 * weight * 9.81)
        # Note: D column is TM (t·m), using B (weight) for braking force calculation
        ws.cell(row=row, column=39).value = (
            f'=IF(AL{row_str}="", "", AL{row_str} + 0.2 * B{row_str} * 9.81 / 1000)'
        )
        ws.cell(row=row, column=39).number_format = number_format
        ws.cell(row=row, column=39).font = styles["normal_font"]

        # Option 1 Ballast Fix Check 컬럼 (AN-AQ) - sdsdds.md: Q열 추가로 AM(39) → AN(40)로 이동
        # AN (40): ΔTM_needed_cm·tm - zzzzz.md 가이드: H의 절대값
        ws.cell(row=row, column=40).value = f'=IF($A{row_str}="","",ABS(H{row_str}))'
        ws.cell(row=row, column=40).number_format = number_format
        ws.cell(row=row, column=40).font = styles["normal_font"]

        # AO (41): Ballast_req_t - zzzzz.md 가이드: J열과 같은 개념 (H/I)
        ws.cell(row=row, column=41).value = (
            f'=IF($A{row_str}="","",'
            f'IF(OR($I{row_str}="",$I{row_str}=0),0,ROUND(H{row_str}/$I{row_str},2)))'
        )
        ws.cell(row=row, column=41).number_format = number_format
        ws.cell(row=row, column=41).font = styles["normal_font"]

        # AP (42): Ballast_gap_t - zzzzz.md 가이드: 필요한 Ballast - 실제 Ballast(L열)
        ws.cell(row=row, column=42).value = (
            f'=IF($A{row_str}="","",AO{row_str} - $L{row_str})'
        )
        ws.cell(row=row, column=42).number_format = number_format
        ws.cell(row=row, column=42).font = styles["normal_font"]

        # AQ (43): Time_Add_h - zzzzz.md 가이드: AO를 펌프 레이트(B13)로 나눈 추가 시간
        ws.cell(row=row, column=43).value = (
            f'=IF($A{row_str}="","",' f"IF($B$13=0,0,AP{row_str}/$B$13))"
        )
        ws.cell(row=row, column=43).number_format = number_format
        ws.cell(row=row, column=43).font = styles["normal_font"]

        # Heel/FSE 컬럼 (AR-AS) - sdsdds.md: Q열 추가로 AQ(43) → AR(44)로 이동
        # AR (44): Heel_deg - 패치 v20251122 최소: 전체 Δ(AD) 사용
        # MD 파일 공식: heel_deg = DEGREES((weight_t * y_offset_m) / (disp_t * gm_m))
        # B{row}: weight_t (W_stage_t)
        # Calc!$E$43: y_offset_m (heel_y_offset_m)
        # AD{row}: disp_t (Total Displacement = Lightship + Ballast + Cargo)
        # U{row}: gm_m - sdsdds.md: Q열 추가로 T(20) → U(21)로 이동
        ws.cell(row=row, column=44).value = (
            f'=IF(OR($A{row_str}="", U{row_str}="", U{row_str}=0, AD{row_str}="", AD{row_str}=0), "", '
            f"DEGREES((B{row_str} * Calc!$E$43) / (AD{row_str} * U{row_str})))"
        )
        ws.cell(row=row, column=44).number_format = number_format
        ws.cell(row=row, column=44).font = styles["normal_font"]

        # AS (45): GM_eff_m - 패치 v20251122 최소: 전체 Δ(AD) 사용
        # MD 파일 공식: GM_eff = GM - FSE / Δ
        # U{row}: gm_m (원래 GM) - sdsdds.md: Q열 추가로 T(20) → U(21)로 이동
        # AD{row}: disp_t (Total Displacement)
        # FSE_t_m: FSE (t·m), 현재는 0으로 단순화 (향후 탱크별 계산으로 확장 가능)
        fse_value = 0  # Simplified, can be enhanced with tank-by-tank calculation
        ws.cell(row=row, column=45).value = (
            f'=IF(OR($A{row_str}="", U{row_str}="", AD{row_str}="", AD{row_str}=0), "", '
            f"U{row_str} - {fse_value} / AD{row_str})"
        )
        ws.cell(row=row, column=45).number_format = number_format
        ws.cell(row=row, column=45).font = styles["normal_font"]

        # Ramp Angle & Pin Stress 컬럼 (AT-AW) - sdsdds.md: Q열 추가로 AS(45) → AT(46)로 이동
        # AT (46): Ramp_Angle_deg
        # Z{row}: Phys_Freeboard_m - sdsdds.md: Q열 추가로 Y(25) → Z(26)로 이동
        ws.cell(row=row, column=46).value = (
            f'=IF(Z{row_str}="","",DEGREES(ASIN((Z{row_str}-Calc!$E$35)/Calc!$E$33)))'
        )
        ws.cell(row=row, column=46).number_format = number_format
        ws.cell(row=row, column=46).font = styles["normal_font"]

        # AU (47): Ramp_Angle_Check - sdsdds.md: Q열 추가로 AT(46) → AU(47)로 이동
        ws.cell(row=row, column=47).value = (
            f'=IF(AT{row_str}="","",IF(AT{row_str}<=10,"OK","NG"))'
        )
        ws.cell(row=row, column=47).font = styles["normal_font"]

        # AV (48): Pin_Stress_N_mm2 - sdsdds.md: Q열 추가로 AU(47) → AV(48)로 이동
        ws.cell(row=row, column=48).value = (
            f'=IF(AH{row_str}="","",(AH{row_str}/4)/Calc!$E$36*9.81/1000)'
        )
        ws.cell(row=row, column=48).number_format = number_format
        ws.cell(row=row, column=48).font = styles["normal_font"]

        # AW (49): Von_Mises_Check - sdsdds.md: Q열 추가로 AV(48) → AW(49)로 이동
        ws.cell(row=row, column=49).value = (
            f'=IF(AV{row_str}="","",IF(AV{row_str}<=188,"OK","NG"))'
        )
        ws.cell(row=row, column=49).font = styles["normal_font"]

        # AX (50): Required_Tide_m
        # Required_Tide_m = IF(Phys_Freeboard_m>=0, 0, ABS(Phys_Freeboard_m) + 0.30)
        # Z{row}: Phys_Freeboard_m (column 26)
        ws.cell(row=row, column=50).value = (
            f'=IF(Z{row_str}="", "", IF(Z{row_str}>=0, 0, ABS(Z{row_str})+0.30))'
        )
        ws.cell(row=row, column=50).number_format = number_format
        ws.cell(row=row, column=50).font = styles["normal_font"]

        # AY (51): Tide_OK
        # Tide_OK = IF(Tide_ref >= Required_Tide_m, "OK", "CHECK")
        # Tide_ref는 RORO_Stage_Scenarios!B5
        ws.cell(row=row, column=51).value = (
            f'=IF(AX{row_str}="", "", IF($B$5>=AX{row_str}, "OK", "CHECK"))'
        )
        ws.cell(row=row, column=51).font = styles["normal_font"]

    # 컬럼 너비 설정 - Opt C Tide 컬럼 추가로 범위 조정
    for col in range(32, 52):  # AF(32) ~ AY(51)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 11
    ws.column_dimensions["AQ"].width = 18  # Heel_deg
    ws.column_dimensions["AR"].width = 15  # GM_eff_m
    ws.column_dimensions["AT"].width = (
        15  # Ramp Angle - sdsdds.md: Q열 추가로 AS(45) → AT(46)로 이동
    )
    ws.column_dimensions["AV"].width = (
        15  # Pin Stress - sdsdds.md: Q열 추가로 AU(47) → AV(48)로 이동
    )
    ws.column_dimensions["AX"].width = 15  # Required_Tide_m
    ws.column_dimensions["AY"].width = 12  # Tide_OK

    print(
        "  [OK] Structural Strength & Option 1 Ballast Fix Check columns added to RORO_Stage_Scenarios sheet"
    )
    print("  [OK] Hinge Rx 자동 계산 적용")
    print(
        "  [OK] Dynamic Load Case B/C (AL-AM), Option 1 moved (AN-AQ), Heel/FSE (AR-AS), Ramp/Stress moved (AT-AW) (sdsdds.md: Q열 추가)"
    )


def extend_precision_columns(ws, first_data_row, num_stages):
    """
    RORO 시트에 정밀 계산 적용 (v4.0)
    - G, H 컬럼을 직접 수정: FWD_precise, AFT_precise (Fr_stage 추가로 컬럼 인덱스 변경)
    - Fr_stage 컬럼 추가로 인한 컬럼 구조:
      - Column 6: Trim_cm (F)
      - Column 7: FWD_precise_m (G)
      - Column 8: AFT_precise_m (H)
    - 순환 참조 방지: Tmean을 Baseline draft ($B$6) 기반으로 계산
    """
    styles = get_styles()
    # 숫자 포맷 통일: 천단위 구분, 소수점 2자리
    number_format = "#,##0.00"

    # 데이터 행에 수식 추가
    for row in range(first_data_row, first_data_row + num_stages):
        row_str = str(row)

        # G (7): FWD_precise_m - LCF 기반 정밀 Forward Draft
        # MD 파일 공식: Dfwd = Tmean - Trim_m * (1 - LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        # Note: MD 파일의 LCF는 F.P. 기준, Calc!$E$41은 LCF_from_mid_m (midship 기준)
        # 좌표계 변환: LCF_from_FP = LCF_from_mid_m + LBP/2
        # r = LCF_from_FP / LBP = (LCF_from_mid_m + LBP/2) / LBP = LCF_from_mid_m/LBP + 1/2
        # F: 1 - r = 1 - (LCF_from_mid_m/LBP + 1/2) = 1/2 - LCF_from_mid_m/LBP
        ws.cell(row=row, column=7).value = (
            f'=IF($A{row_str}="", "", '
            f"$B$6 - (F{row_str}/100) * (0.5 - Calc!$E$41 / Calc!$E$40))"
        )
        ws.cell(row=row, column=7).number_format = number_format
        ws.cell(row=row, column=7).font = styles["normal_font"]

        # H (8): AFT_precise_m - LCF 기반 정밀 Aft Draft
        # MD 파일 공식: Daft = Tmean + Trim_m * (LCF/LBP)
        # Tmean = Baseline draft ($B$6), Trim_m = F/100 (Column 6 = Trim_cm)
        # r = LCF_from_FP / LBP = LCF_from_mid_m/LBP + 1/2
        ws.cell(row=row, column=8).value = (
            f'=IF($A{row_str}="", "", '
            f"$B$6 + (F{row_str}/100) * (Calc!$E$41 / Calc!$E$40 + 0.5))"
        )
        ws.cell(row=row, column=8).number_format = number_format
        ws.cell(row=row, column=8).font = styles["normal_font"]

    print("  [OK] F, G columns modified to FWD_precise, AFT_precise (v4.0)")


def create_ballast_tanks_sheet(wb):
    """Ballast_Tanks 시트 생성 (tank_coordinates.json + tank_data.json 기반)"""
    ws = wb.create_sheet("Ballast_Tanks")
    styles = get_styles()

    headers = ["TankName", "x_from_mid_m", "max_t", "SG", "use_flag", "air_vent_mm"]

    # 1) JSON 병합: tank_coordinates.json + tank_data.json
    tank_lookup = build_tank_lookup()

    # 2) 이 프로젝트에서 실제로 쓸 Ballast 탱크 목록 + 기본 use_flag
    #    (필요 시 여기에 DO, FO, 기타 탱크 추가 가능)
    target_tanks = [
        ("FWB1.P", "Y"),
        ("FWB1.S", "Y"),
        ("FWB2.P", "Y"),
        ("FWB2.S", "Y"),
        ("FWCARGO1.P", "N"),  # 선택 사용
        ("FWCARGO1.S", "N"),
        ("FWCARGO2.P", "N"),
        ("FWCARGO2.S", "N"),
    ]

    # 3) Fallback 값 (JSON 없거나, 특정 탱크 키가 비어 있을 때 사용)
    fallback = {
        "FWB1.P": {"x": 57.52, "max_t": 50.57, "SG": 1.025, "air_vent_mm": 80},
        "FWB1.S": {"x": 57.52, "max_t": 50.57, "SG": 1.025, "air_vent_mm": 80},
        "FWB2.P": {"x": 50.04, "max_t": 109.98, "SG": 1.025, "air_vent_mm": 80},
        "FWB2.S": {"x": 50.04, "max_t": 109.98, "SG": 1.025, "air_vent_mm": 80},
        "FWCARGO1.P": {"x": 42.75, "max_t": 148.35, "SG": 1.000, "air_vent_mm": 125},
        "FWCARGO1.S": {"x": 42.75, "max_t": 148.35, "SG": 1.000, "air_vent_mm": 125},
        "FWCARGO2.P": {"x": 35.25, "max_t": 148.36, "SG": 1.000, "air_vent_mm": 125},
        "FWCARGO2.S": {"x": 35.25, "max_t": 148.36, "SG": 1.000, "air_vent_mm": 125},
    }

    # 4) 헤더 작성
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # 5) 데이터 행 작성
    for row_idx, (tank_name, use_flag) in enumerate(target_tanks, start=2):
        info = tank_lookup.get(tank_name, {})
        fb = fallback.get(tank_name, {})

        x_val = info.get("x_from_mid_m", fb.get("x"))
        max_t = info.get("max_t", fb.get("max_t"))
        sg = info.get("SG", fb.get("SG", 1.0))
        air_vent_mm = info.get("air_vent_mm", fb.get("air_vent_mm", ""))

        row_data = [tank_name, x_val, max_t, sg, use_flag, air_vent_mm]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.font = styles["normal_font"]
            if c >= 2:  # 숫자 열
                cell.number_format = "0.00"
            cell.border = Border(
                left=styles["thin_border"],
                right=styles["thin_border"],
                top=styles["thin_border"],
                bottom=styles["thin_border"],
            )

    # 6) 컬럼 폭 설정
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14

    if tank_lookup:
        print(
            "  [OK] Ballast_Tanks updated with tank_coordinates.json + tank_data.json (2025-11-18)"
        )
    else:
        print("  [WARN] Ballast_Tanks used fallback hard-coded data (JSON not found)")


def create_hydro_table_sheet(wb):
    """Hydro_Table 시트 생성"""
    ws = wb.create_sheet("Hydro_Table")
    styles = get_styles()

    headers = ["Disp_t", "Tmean_m", "Trim_m", "GM_m", "Draft_FWD", "Draft_AFT"]

    # JSON에서 데이터 로드 시도
    json_data = _load_json("data/hydro_table.json")
    if json_data:
        # JSON이 dict 리스트인 경우 배열로 변환
        if (
            isinstance(json_data, list)
            and len(json_data) > 0
            and isinstance(json_data[0], dict)
        ):
            data = []
            for entry in json_data:
                data.append(
                    [
                        entry.get("Disp_t", 0.0),
                        entry.get("Tmean_m", 0.0),
                        entry.get("Trim_m", 0.0),
                        entry.get("GM_m", 0.0),
                        entry.get("Draft_FWD", 0.0),
                        entry.get("Draft_AFT", 0.0),
                    ]
                )
        else:
            # 이미 배열 형식인 경우
            data = json_data
        print(f"  [OK] Hydro_Table loaded from JSON ({len(data)} points)")
    else:
        # Fallback: 기존 4점 데이터
        print("  [FALLBACK] Using built-in 4 points")
        data = [
            [2991.25, 2.20, 0.20, 2.85, 2.10, 2.30],
            [3208.25, 3.18, -0.53, 1.68, 2.92, 3.45],
            [3265.25, 3.00, 0.60, 1.88, 2.68, 3.28],
            [3425.25, 3.00, 0.70, 1.85, 2.65, 3.35],
        ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(
                row=r,
                column=c,
                value=float(val) if isinstance(val, (int, float)) else val,
            )
            cell.font = styles["normal_font"]
            cell.number_format = "0.00"

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 12

    print("  [OK] Hydro_Table sheet created")


def create_frame_table_sheet(wb):
    """Frame_to_x_Table 시트 생성"""
    ws = wb.create_sheet("Frame_to_x_Table")
    styles = get_styles()

    headers = ["Fr", "x_from_mid_m", "비고"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border = Border(
            left=styles["thin_border"],
            right=styles["thin_border"],
            top=styles["thin_border"],
            bottom=styles["thin_border"],
        )

    # JSON 파일 로드 (상대 경로 사용)
    frame_data = _load_json("data/Frame_x_from_mid_m.json")

    if frame_data:
        try:
            # frame_data가 리스트인지 확인
            if isinstance(frame_data, list):
                for idx, entry in enumerate(frame_data, start=0):
                    row = 2 + idx

                    cell_a = ws.cell(row=row, column=1)
                    cell_a.value = entry.get("Fr", 0.0)
                    cell_a.font = styles["normal_font"]
                    cell_a.number_format = "0.00"

                    cell_b = ws.cell(row=row, column=2)
                    cell_b.value = entry.get("x_from_mid_m", 0.0)
                    cell_b.font = styles["normal_font"]
                    cell_b.number_format = "0.00"

                    cell_c = ws.cell(row=row, column=3)
                    cell_c.value = entry.get("비고", "")
                    cell_c.font = styles["normal_font"]

                print(
                    f"  [OK] Frame_to_x_Table sheet created with {len(frame_data)} rows"
                )
            else:
                print(f"  [WARNING] Frame data is not a list. Creating empty sheet.")
        except Exception as e:
            print(
                f"  [WARNING] Error processing frame data: {e}. Creating empty sheet."
            )
            frame_data = None

    if not frame_data:
        print(
            f"  [WARNING] JSON file not found. Creating empty Frame_to_x_Table sheet."
        )
        for row in range(2, 123):  # 121 rows + header
            ws.cell(row=row, column=1).font = styles["normal_font"]
            ws.cell(row=row, column=1).number_format = "0.00"
            ws.cell(row=row, column=2).font = styles["normal_font"]
            ws.cell(row=row, column=2).number_format = "0.00"
            ws.cell(row=row, column=3).font = styles["normal_font"]

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20


# ============================================================================
# Main Orchestration Function
# ============================================================================


def create_workbook_from_scratch():
    """워크북을 처음부터 생성 (BACKUP PLAN integrated)"""
    print("=" * 80)
    print("LCT_BUSHRA_AGI_TR.xlsx Creation from Scratch (BACKUP PLAN enabled)")
    print("=" * 80)

    # BACKUP PLAN: Pre-flight check
    print("\n[PRE-FLIGHT CHECK]")
    issues = preflight_check()

    # PHASE 0: Tank JSON auto-generation
    try:
        from src.tank_data_manager import ensure_tank_jsons

        logging.info("[PRE-FLIGHT] Checking tank data files")
        success, msg = ensure_tank_jsons("Tank Capacity_Plan.xlsx", "data/")
        if success:
            logging.info(f"[TANK] {msg}")
            print(f"  [OK] {msg}")
        else:
            issues.append(f"WARNING: {msg}")
            print(f"  [WARNING] {msg}")
    except ImportError:
        issues.append("INFO: Tank auto-generation module not available")
        print("  [INFO] Tank auto-generation module not available")
    except Exception as e:
        issues.append(f"WARNING: Tank JSON generation failed: {e}")
        print(f"  [WARNING] Tank JSON generation failed: {e}")

    for issue in issues:
        print(f"  {issue}")
    if any("ERROR" in i for i in issues):
        print("\n[ABORT] Critical issues found. Exiting.")
        sys.exit(1)

    output_dir = os.path.dirname(OUTPUT_FILE)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[OK] Created output directory: {output_dir}")

    final_output_file = OUTPUT_FILE
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE, "r+b"):
                pass
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(OUTPUT_FILE)[0]
            final_output_file = f"{base_name}_{timestamp}.xlsx"
            print(f"[WARNING] Original file is open. Saving as: {final_output_file}")

    # BACKUP PLAN: Setup logging
    print(f"\n[1/9] Setting up logging and workbook")
    log_file = setup_logging(final_output_file)
    logging.info("[1/9] Workbook creation started")

    wb = Workbook()
    wb.remove(wb.active)

    # BACKUP PLAN: Safe sheet creation with error recovery
    print(f"\n[2/9] Creating sheets (with error recovery):")
    logging.info("[2/9] Sheet creation phase started")

    safe_sheet_creation(wb, create_calc_sheet, "Calc")
    safe_sheet_creation(wb, create_tide_sheet, "December_Tide_2025")
    safe_sheet_creation(wb, create_hourly_sheet, "Hourly_FWD_AFT_Heights")

    result = safe_sheet_creation(wb, create_roro_sheet, "RORO_Stage_Scenarios")
    if result:
        if len(result) == 3:
            stages, first_data_row, total_rows = result
        else:
            # 이전 버전 호환성
            stages, first_data_row = result
            total_rows = len(stages)
    else:
        # BACKUP: Provide fallback values if RORO sheet creation fails
        logging.warning("[BACKUP] RORO sheet failed, using defaults")
        stages = []
        first_data_row = 19
        total_rows = 0

    safe_sheet_creation(wb, create_ballast_tanks_sheet, "Ballast_Tanks")
    safe_sheet_creation(wb, create_hydro_table_sheet, "Hydro_Table")
    safe_sheet_creation(wb, create_frame_table_sheet, "Frame_to_x_Table")

    if "RORO_Stage_Scenarios" in wb.sheetnames and stages:
        roro_ws = wb["RORO_Stage_Scenarios"]
        logging.info("[3/9] Extending RORO sheet with additional columns")
        print(f"\n[3/9] Extending RORO sheet")

        try:
            # 일반 Stage만 처리 (Optional Tuning Stages는 자체 수식 보유)
            extend_roro_captain_req(roro_ws, first_data_row, len(stages))
            extend_roro_structural_opt1(roro_ws, first_data_row, len(stages))
            extend_precision_columns(roro_ws, first_data_row, len(stages))
        except Exception as e:
            logging.error(f"[BACKUP] RORO extension failed: {e}")
            print(f"  [BACKUP] Warning: RORO extension failed: {e}")

        # Create Excel Table after all columns are added
        from openpyxl.worksheet.table import Table, TableStyleInfo

        header_row = 18
        last_col = 51
        last_col_letter = get_column_letter(last_col)

        # Verify all headers are strings before creating table
        for col in range(1, last_col + 1):
            cell = roro_ws.cell(row=header_row, column=col)
            if cell.value is not None and not isinstance(cell.value, str):
                cell.value = str(cell.value)
            elif cell.value is None or cell.value == "":
                cell.value = f"Unused_{col}"

        try:
            table = Table(
                displayName="Stages",
                ref=f"A{header_row}:{last_col_letter}{first_data_row + len(stages) - 1}",
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            roro_ws.add_table(table)
            logging.info("[OK] Excel Table created")
            print("  [OK] Excel Table created successfully")
        except Exception as e:
            logging.warning(f"[BACKUP] Excel Table creation failed: {e}")
            print(f"  [BACKUP] Warning: Could not create Excel Table: {e}")

    # Create OPERATION SUMMARY sheet
    if stages:
        logging.info("[4/9] Creating OPERATION SUMMARY sheet")
        print(f"\n[4/9] Creating OPERATION SUMMARY")
        safe_sheet_creation(
            wb, create_captain_report_sheet, "OPERATION SUMMARY", stages, first_data_row
        )

    # Save workbook
    logging.info(f"[5/9] Saving workbook: {final_output_file}")
    print(f"\n[5/9] Saving workbook: {final_output_file}")
    try:
        wb.save(final_output_file)
        logging.info("[OK] File saved successfully")
        print(f"  [OK] File saved successfully")
    except Exception as e:
        logging.error(f"[ERROR] Failed to save: {e}")
        print(f"  [ERROR] Failed to save: {e}")
        sys.exit(1)

    wb.close()

    # BACKUP PLAN: Create backup after successful save
    print(f"\n[6/9] Creating backup")
    logging.info("[6/9] Creating backup file")
    backup_path = create_backup_file(final_output_file)

    # Verification
    logging.info("[7/9] Verification")
    print(f"\n[7/9] Verification:")
    if os.path.exists(final_output_file):
        file_size = os.path.getsize(final_output_file) / 1024
        logging.info(f"File created: {final_output_file}, Size: {file_size:.2f} KB")
        print(f"  [OK] File created: {final_output_file}")
        print(f"  [OK] File size: {file_size:.2f} KB")
        print(f"  [OK] Sheets: {len(wb.sheetnames)}")
        if backup_path:
            print(f"  [OK] Backup: {os.path.basename(backup_path)}")
        print(f"  [OK] Log: {os.path.basename(log_file)}")
    else:
        logging.error("[ERROR] Output file was not created")
        print(f"  [ERROR] Output file was not created")
        sys.exit(1)

    print("\n" + "=" * 80)
    print("[SUCCESS] Workbook creation complete! (BACKUP PLAN active)")
    print("=" * 80)
    logging.info("[SUCCESS] Workbook creation complete")


# ============================================================================
# Stage Evaluation Functions (wewewewe.md 가이드)
# ============================================================================

# 타입 alias
StageDict = Dict[str, Any]

# GMGrid 예시 (실전에서는 Aries/BV GM 테이블 로드)
# Δ–Trim–GM (m) 테이블
GM_GRID_EXAMPLE: GMGrid = {
    # Δ = 1227.59 t 일 때 Trim–GM
    1227.59: {
        0.00: 1.60,
        0.50: 1.58,
        1.00: 1.55,
    },
    # Δ = 1658.71 t 일 때 Trim–GM
    1658.71: {
        0.00: 1.55,
        0.50: 1.53,
        1.00: 1.50,
    },
}

LCF_M = (
    30.91  # m (AP 기준 LCF) - BUSHRA verified: Loaded condition avg LCG 31.45m from AP
)
LBP_M = 60.302  # m (LBP)

# Stage 입력 예시 (엑셀 Row → dict 로 나왔다고 가정)
# - 실제로는 CSV/엑셀에서 읽어서 채우면 됨
STAGES_EXAMPLE: List[StageDict] = [
    {
        "name": "Stage 5A-2",
        "Tmean_m": 2.85,
        "Trim_cm": -96.50,  # 선수침(-), 선미침(+)
        "Disp_t": 1658.71,  # Δ
        "W_stage_t": 217.00,  # 해당 Stage 주요 카고/편심 하중
        "Y_offset_m": 2.50,  # 중심선 기준 횡 편심 (SPMT 등)
        "FSE_t_m": 85.00,  # Free Surface Moment (t·m)
        "Share_Load_t": 210.00,  # Ramp share load (정적)
        "Pin_Stress_MPa": 120.00,  # Pin stress (정적)
        "LoadCase": "B",  # A=STATIC, B=DYNAMIC, C=BRAKING
    },
    {
        "name": "Stage 6B",
        "Tmean_m": 3.10,
        "Trim_cm": -120.00,
        "Disp_t": 1227.59,
        "W_stage_t": 434.00,
        "Y_offset_m": 1.50,
        "FSE_t_m": 40.00,
        "Share_Load_t": 250.00,
        "Pin_Stress_MPa": 135.00,
        "LoadCase": "C",
    },
]


def evaluate_stages(
    stages: List[StageDict],
    gm_grid: GMGrid,
    lcf_m: float = LCF_M,
    lbp_m: float = LBP_M,
) -> List[StageDict]:
    """
    각 Stage에 대해:
      - Dfwd_precise_m / Daft_precise_m (LCF 기반 Draft)
      - GM_calc_m (Δ–Trim 2D 보간)
      - Heel_deg / GM_eff_m / Heel_OK / GM_OK
      - Share_Load_dyn_t / Pin_Stress_dyn_MPa
    필드를 추가해서 반환.
    """

    result: List[StageDict] = []

    for stage in stages:
        # ---- 1) LCF 기반 정밀 Draft ----
        tmean_m = float(stage.get("Tmean_m", 0.0))
        trim_cm = float(stage.get("Trim_cm", 0.0))

        dfwd_m, daft_m = calc_draft_with_lcf(
            tmean_m=tmean_m,
            trim_cm=trim_cm,
            lcf_m=lcf_m,
            lbp_m=lbp_m,
        )
        stage["Dfwd_precise_m"] = round(dfwd_m, 3)
        stage["Daft_precise_m"] = round(daft_m, 3)

        # ---- 2) Δ–Trim 2D GM 보간 ----
        disp_t = float(stage.get("Disp_t", 0.0))
        trim_m = trim_cm / 100.0

        gm_m = get_gm_bilinear(
            disp_t=disp_t,
            trim_m=trim_m,
            gm_grid=gm_grid,
        )
        stage["GM_calc_m"] = round(gm_m, 3)

        # ---- 3) Heel + FSE 반영 GM_eff ----
        weight_t = float(stage.get("W_stage_t", 0.0))
        y_offset_m = float(stage.get("Y_offset_m", 0.0))
        fse_t_m = float(stage.get("FSE_t_m", 0.0))

        heel_deg, gm_eff, heel_ok, gm_ok = heel_and_gm_check(
            weight_t=weight_t,
            y_offset_m=y_offset_m,
            disp_t=disp_t,
            gm_m=gm_m,
            fse_t_m=fse_t_m,
            heel_limit_deg=3.0,
            gm_min_m=1.50,
        )
        stage["Heel_deg"] = round(heel_deg, 3)
        stage["GM_eff_m"] = round(gm_eff, 3)
        stage["Heel_OK"] = heel_ok
        stage["GM_OK"] = gm_ok

        # ---- 4) 동적 / 제동 Load Case ----
        share_static = float(stage.get("Share_Load_t", 0.0))
        pin_static = float(stage.get("Pin_Stress_MPa", 0.0))

        # LoadCase 문자열 → Enum 매핑
        lc_raw = str(stage.get("LoadCase", "A")).upper()
        if lc_raw in ("A", "STATIC"):
            lc = LoadCase.STATIC
        elif lc_raw in ("B", "DYNAMIC"):
            lc = LoadCase.DYNAMIC
        elif lc_raw in ("C", "BRAKE", "BRAKING"):
            lc = LoadCase.BRAKING
        else:
            lc = LoadCase.STATIC

        share_dyn, pin_dyn = apply_dynamic_loads(
            share_load_t=share_static,
            pin_stress_mpa=pin_static,
            load_case=lc,
        )
        stage["Share_Load_dyn_t"] = round(share_dyn, 2)
        stage["Pin_Stress_dyn_MPa"] = round(pin_dyn, 2)
        stage["LoadCase_used"] = lc.name  # "STATIC" / "DYNAMIC" / "BRAKING"

        result.append(stage)

    return result


if __name__ == "__main__":
    import sys

    _init_frame_mapping()

    # Self-check: Frame 축 패치 검증
    print("=" * 60)
    for label, fr in {
        "AP approx": 0.0,
        "Midship (Lpp/2)": 30.151,
        "FP approx": 60.30,
        "FWB1/2 center 55": 55.0,
    }.items():
        x = fr_to_x(fr)
        print(
            f"{label:25}  Fr={fr:6.2f}  →  x={x:8.3f} m  ({'FWD' if x < 0 else 'AFT'})"
        )
    print("=" * 60)

    if len(sys.argv) > 1 and sys.argv[1] == "debug":
        debug_frame_mapping()
    elif len(sys.argv) > 1 and sys.argv[1] == "demo":
        # Stage 평가 데모 실행
        enriched_stages = evaluate_stages(STAGES_EXAMPLE, GM_GRID_EXAMPLE)

        from pprint import pprint

        print("\n=== Stage 계산 결과 요약 ===")
        for s in enriched_stages:
            print(f"\n[{s['name']}]")
            pprint(
                {
                    "Dfwd_precise_m": s["Dfwd_precise_m"],
                    "Daft_precise_m": s["Daft_precise_m"],
                    "GM_calc_m": s["GM_calc_m"],
                    "GM_eff_m": s["GM_eff_m"],
                    "Heel_deg": s["Heel_deg"],
                    "Heel_OK": s["Heel_OK"],
                    "GM_OK": s["GM_OK"],
                    "Share_Load_dyn_t": s["Share_Load_dyn_t"],
                    "Pin_Stress_dyn_MPa": s["Pin_Stress_dyn_MPa"],
                    "LoadCase_used": s["LoadCase_used"],
                }
            )
    else:
        create_workbook_from_scratch()
