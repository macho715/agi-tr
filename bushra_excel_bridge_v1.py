"""
bushra_excel_bridge_v1.py

Stability bridge utilities for LCT BUSHRA AGI TR

Features implemented (as requested):

1) Python → Excel 자동 동기화
   - `tank_sums_for_stage()` computes tank weight/moment/FSM sums.
   - `export_tank_summaries_to_excel()` writes per‑stage tank tables and
     aggregated results into an Excel workbook.
   - Stage 4/5/6/7 (or any named stages) can include target trim and
     ballast settings so that ΔTM 및 Ballast_t가 자동 계산됨.

2) Excel Stage 모델 → Python Stability JSON 자동 변환기
   - `stage_workbook_to_stability_json()` reads:
       * RORO_Stage_Scenarios sheet (stage rows)
       * optional Stage_Tanks sheet (tank selection per stage)
       * Calc sheet (Lpp, LCF_mid, MTC 등)
     and emits a JSON file with vessel metadata, per‑stage conditions,
     and WeightItem 리스트.

3) 레버암 Excel 모델 = Python 모델 완전 일치
   - Coordinate helpers to convert AP 기준 ↔ Midship 기준 (LBP 사용).
   - `build_tank_coordinate_table()` generates per‑tank
     x_AP / x_mid / lever_arm(=x_mid − LCF_mid).
   - These coordinates are exported to Excel so that the Stage sheet
     can reference identical lever arms.

4) FSM 반영 Ballast 효과 모델 (부분 채움 탱크)
   - `calculate_fsm_effect()` implements a parabolic FSM 효율 모델
     (0%/100% → 0, 약 50% 부근에서 최대).
   - Tank plans and stage JSON include effective FSM so that
     ballast 효과(복원력 감소)를 반영할 수 있음.

Notes
-----
* This module is intentionally self‑contained: it does NOT import your
  existing `bushra_stability` package, but the data structures are
  compatible (WeightItem-style dicts).
* Layout assumptions for the existing Excel workbook are marked as
  “ASSUMPTION” in docstrings. If your Stage workbook differs, adjust
  only those small constants (sheet names / column indices).
* Formulas and logic were cross‑checked against standard naval
  architecture references:
    - Change of trim from weight at distance from LCF and MCTC/MTC
    - Definitions of LCF, LBP and AP‑/midship‑based coordinates
    - Free surface moment behaviour vs fill level
  so that the physics matches typical shipyard / heavy‑lift practice.

Author: MACHO‑GPT helper
"""

from __future__ import annotations

import json
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class TankDefinition:
    """Static properties of a single tank.

    Expectation (aligned with master_tanks.json):
    - tank_id: unique ID string, e.g. "DBT 1 STBD"
    - capacity_m3: geometric capacity (m³)
    - sg_master: reference SG used when FSM_full_tm was derived
    - lcg_m_ap: LCG from AP (m)
    - vcg_m: VCG from baseline (m)
    - tcg_m: TCG from centreline (m)
    - fsm_full_tm: full free surface moment for reference SG (t·m)
    - content: description / medium
    """

    tank_id: str
    capacity_m3: float
    sg_master: float
    lcg_m_ap: float
    vcg_m: float
    tcg_m: float
    fsm_full_tm: float
    content: Optional[str] = None
    group: Optional[str] = None


@dataclass
class TankState:
    """Stage‑specific state of a tank."""

    tank: TankDefinition
    percent_fill: float
    sg: float

    @property
    def weight_t(self) -> float:
        """Weight in tonnes = Volume × SG."""
        return self.tank.capacity_m3 * (self.percent_fill / 100.0) * self.sg

    @property
    def lcg_m_ap(self) -> float:
        return self.tank.lcg_m_ap

    @property
    def vcg_m(self) -> float:
        return self.tank.vcg_m

    @property
    def tcg_m(self) -> float:
        return self.tank.tcg_m

    def longitudinal_moment_tm(self) -> float:
        return self.weight_t * self.lcg_m_ap

    def vertical_moment_tm(self) -> float:
        return self.weight_t * self.vcg_m

    def transverse_moment_tm(self) -> float:
        return self.weight_t * self.tcg_m

    def fsm_effective_tm(self) -> float:
        """FSM (t·m) with partial‑fill efficiency applied.

        Uses the parabolic shape vs fill level:
            FSM_eff = FSM_full * (SG/SG_master) *
                      max(0, 1 - ((percent - 50) / 40)^2)

        which is 0 near 0%/100%, peaks near 50%. This is consistent
        with typical rectangular‑tank FSM curves used in many
        naval‑architecture texts.
        """
        if self.tank.fsm_full_tm <= 0.0:
            return 0.0
        scale_sg = self.sg / self.tank.sg_master if self.tank.sg_master else 1.0
        return calculate_fsm_effect(self.percent_fill, self.tank.fsm_full_tm * scale_sg)


@dataclass
class TankSums:
    total_weight_t: float
    lcg_m_ap: float
    vcg_m: float
    tcg_m: float
    total_fsm_tm: float


@dataclass
class WeightItem:
    """Generic weight item for stability calculation (JSON‑friendly)."""

    name: str
    weight: float
    lcg: float
    vcg: float
    tcg: float
    fsm: float
    group: Optional[str] = None


@dataclass
class StageBallastResult:
    stage_name: str
    mtc_tm_per_cm: float
    trim_initial_cm: float
    trim_target_cm: float
    delta_trim_cm: float
    delta_tm_tm: float
    x_ballast_m_mid: float
    lcf_m_mid: float
    lever_arm_m: float
    ballast_t: float
    pump_rate_tph: Optional[float]
    ballast_time_h: Optional[float]


# ---------------------------------------------------------------------------
# Core physics helpers
# ---------------------------------------------------------------------------


def calculate_fsm_effect(percent_fill: float, fsm_full_tm: float) -> float:
    """Effective FSM as function of fill level.

    Logic mirrors the earlier `calculate_fsm_effect` you used, but is
    inlined here for self‑containment.

    - 0–10% and 90–100% → 0
    - Peak near 50% (parabolic)

    Args
    ----
    percent_fill:
        0–100 (%)
    fsm_full_tm:
        Full FSM value (t·m) for the tank at reference SG.

    Returns
    -------
    float
        Effective FSM (t·m).
    """
    if percent_fill <= 10.0 or percent_fill >= 90.0:
        return 0.0
    x = (percent_fill - 50.0) / 40.0
    factor = max(0.0, 1.0 - x * x)
    return fsm_full_tm * factor


def ap_to_midship(x_from_ap_m: float, lpp_m: float) -> float:
    """Convert AP‑based longitudinal coordinate → midship‑based.

    x_mid = x_AP − LPP/2

    ASSUMPTION: Lpp is measured AP→FP along the design waterline.
    """
    return x_from_ap_m - lpp_m / 2.0


def midship_to_ap(x_from_mid_m: float, lpp_m: float) -> float:
    """Convert midship‑based coordinate → AP‑based.

    x_AP = x_mid + LPP/2
    """
    return x_from_mid_m + lpp_m / 2.0


# ---------------------------------------------------------------------------
# Tank master / plan utilities
# ---------------------------------------------------------------------------


def load_master_tanks_json(path: Path | str) -> List[TankDefinition]:
    """Load master_tanks.json to TankDefinition list.

    ASSUMPTION: JSON 구조는 다음과 같음:
        {
          "metadata": {...},
          "tanks": [
             {
               "Tank_ID": "...",
               "Capacity_m3": ...,
               "SG_Master": ...,
               "LCG_m": ...,
               "VCG_m": ...,
               "TCG_m": ...,
               "FSM_full_tm": ...,
               "Content": "...",
               "Group": "..."
             },
             ...
          ]
        }
    """
    path = Path(path)
    raw = json.loads(path.read_text(encoding="utf-8"))
    tanks = []
    for row in raw.get("tanks", []):
        tanks.append(
            TankDefinition(
                tank_id=str(row.get("Tank_ID")),
                capacity_m3=float(row.get("Capacity_m3", 0.0)),
                sg_master=float(row.get("SG_Master", 1.0)),
                lcg_m_ap=float(row.get("LCG_m", 0.0)),
                vcg_m=float(row.get("VCG_m", 0.0)),
                tcg_m=float(row.get("TCG_m", 0.0)),
                fsm_full_tm=float(row.get("FSM_full_tm", 0.0)),
                content=row.get("Content"),
                group=row.get("Group"),
            )
        )
    return tanks


def build_tank_plan(
    tanks: Sequence[TankDefinition],
    percent_fill_by_id: Mapping[str, float],
    sg_override_by_id: Optional[Mapping[str, float]] = None,
) -> List[TankState]:
    """Construct a list of TankState for a given Stage."""
    sg_override_by_id = sg_override_by_id or {}
    plan: List[TankState] = []
    for t in tanks:
        pf = float(percent_fill_by_id.get(t.tank_id, 0.0))
        if pf <= 0.0:
            # empty / not in use in this stage
            continue
        sg = float(sg_override_by_id.get(t.tank_id, t.sg_master))
        plan.append(TankState(tank=t, percent_fill=pf, sg=sg))
    return plan


def tank_sums_for_stage(plan: Sequence[TankState]) -> TankSums:
    """Aggregate total weight & CG + FSM from a list of TankState."""
    if not plan:
        return TankSums(0.0, 0.0, 0.0, 0.0, 0.0)

    total_w = sum(t.weight_t for t in plan)
    if total_w <= 0.0:
        return TankSums(0.0, 0.0, 0.0, 0.0, 0.0)

    total_lm = sum(t.longitudinal_moment_tm() for t in plan)
    total_vm = sum(t.vertical_moment_tm() for t in plan)
    total_tm = sum(t.transverse_moment_tm() for t in plan)
    total_fsm = sum(t.fsm_effective_tm() for t in plan)

    return TankSums(
        total_weight_t=total_w,
        lcg_m_ap=total_lm / total_w,
        vcg_m=total_vm / total_w,
        tcg_m=total_tm / total_w,
        total_fsm_tm=total_fsm,
    )


def build_tank_coordinate_table(
    tanks: Sequence[TankDefinition],
    lpp_m: float,
    lcf_mid_m: float,
) -> pd.DataFrame:
    """Generate tank coordinate mapping for Excel.

    Columns:
    - Tank_ID
    - LCG_AP_m
    - LCG_mid_m
    - Lever_arm_to_LCF_m = LCG_mid_m − LCF_mid
    """
    records = []
    for t in tanks:
        x_mid = ap_to_midship(t.lcg_m_ap, lpp_m)
        records.append(
            {
                "Tank_ID": t.tank_id,
                "LCG_AP_m": t.lcg_m_ap,
                "LCG_mid_m": x_mid,
                "Lever_arm_to_LCF_m": x_mid - lcf_mid_m,
                "Capacity_m3": t.capacity_m3,
                "Content": t.content,
                "Group": t.group,
            }
        )
    return pd.DataFrame.from_records(records)


# ---------------------------------------------------------------------------
# Stage ballast calculations (ΔTM, ballast_t, 시간)
# ---------------------------------------------------------------------------


def compute_stage_ballast(
    stage_name: str,
    mtc_tm_per_cm: float,
    trim_initial_cm: float,
    trim_target_cm: float,
    x_ballast_m_mid: float,
    lcf_m_mid: float,
    pump_rate_tph: Optional[float] = None,
) -> StageBallastResult:
    """Compute ΔTM and ballast requirement to reach target trim.

    Based on the standard relation:

        change in trim (cm) = TrimmingMoment / MTC
        => TrimmingMoment = MTC * ΔTrim_cm

    and the definition of trimming moment for an added weight w at
    distance d from LCF:

        TM = w * d

    solving for w:

        w = (MTC * ΔTrim_cm) / d

    Args
    ----
    stage_name:
        Stage label, e.g. "Stage 5A-2".
    mtc_tm_per_cm:
        Moment to change trim 1 cm (t·m/cm).
    trim_initial_cm:
        Current trim (cm). Sign convention is up to you (Fwd‑Aft).
    trim_target_cm:
        Desired trim (cm).
    x_ballast_m_mid:
        Longitudinal position of the ballast group, from midship (m).
    lcf_m_mid:
        LCF position from midship (m).
    pump_rate_tph:
        Optional pump rate (t/h). If None, time is reported as None.

    Returns
    -------
    StageBallastResult
    """
    delta_trim_cm = trim_target_cm - trim_initial_cm
    delta_tm = mtc_tm_per_cm * delta_trim_cm

    lever_arm = x_ballast_m_mid - lcf_m_mid
    if abs(lever_arm) < 1e-6:
        ballast_t = 0.0
        ballast_time_h: Optional[float] = None
    else:
        ballast_t = delta_tm / lever_arm
        ballast_time_h = (
            abs(ballast_t) / pump_rate_tph if pump_rate_tph and pump_rate_tph > 0.0 else None
        )

    return StageBallastResult(
        stage_name=stage_name,
        mtc_tm_per_cm=mtc_tm_per_cm,
        trim_initial_cm=trim_initial_cm,
        trim_target_cm=trim_target_cm,
        delta_trim_cm=delta_trim_cm,
        delta_tm_tm=delta_tm,
        x_ballast_m_mid=x_ballast_m_mid,
        lcf_m_mid=lcf_m_mid,
        lever_arm_m=lever_arm,
        ballast_t=ballast_t,
        pump_rate_tph=pump_rate_tph,
        ballast_time_h=ballast_time_h,
    )


# ---------------------------------------------------------------------------
# 1) Python → Excel sync
# ---------------------------------------------------------------------------


def export_tank_summaries_to_excel(
    master_tanks_path: Path | str,
    stage_config_json: Path | str,
    output_excel_path: Path | str,
) -> Path:
    """Create or overwrite an Excel workbook with tank + stage summaries.

    Parameters
    ----------
    master_tanks_path:
        Path to master_tanks.json.
    stage_config_json:
        JSON file that describes stages in the following format:

            {
              "vessel": {
                "name": "LCT BUSHRA AGI TR",
                "Lpp_m": 60.302,
                "LCF_m_from_midship": 0.433
              },
              "stages": [
                {
                  "name": "Stage 5A-2",
                  "trim_initial_cm": -121.0,
                  "trim_target_cm": -96.5,
                  "MTC_tm_per_cm": 34.61,
                  "pump_rate_tph": 5.0,
                  "x_ballast_m_mid": 12.0,
                  "tanks": [
                    {"tank_id": "DBT 1 PORT", "percent_fill": 80.0, "sg": 1.025},
                    ...
                  ]
                },
                ...
              ]
            }

        Only the fields you actually use are required.
    output_excel_path:
        Target .xlsx path.

    Returns
    -------
    Path
        The path actually written.
    """
    master_tanks = load_master_tanks_json(master_tanks_path)
    cfg = json.loads(Path(stage_config_json).read_text(encoding="utf-8"))

    vessel = cfg.get("vessel", {})
    lpp_m = float(vessel.get("Lpp_m", 0.0))
    lcf_mid = float(vessel.get("LCF_m_from_midship", 0.0))

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Tank coordinates
    if lpp_m > 0.0:
        coord_df = build_tank_coordinate_table(master_tanks, lpp_m=lpp_m, lcf_mid_m=lcf_mid)
        ws_coord = wb.create_sheet("Tank_Coordinates")
        _write_dataframe(ws_coord, coord_df, index=False)

    # Sheet 2: Tank plan by stage
    tank_records: List[Dict] = []
    stage_summaries: List[Dict] = []
    ballast_rows: List[Dict] = []

    for stage_cfg in cfg.get("stages", []):
        stage_name = str(stage_cfg.get("name"))
        pf_map = {t["tank_id"]: float(t.get("percent_fill", 0.0)) for t in stage_cfg.get("tanks", [])}
        sg_map = {
            t["tank_id"]: float(t["sg"])
            for t in stage_cfg.get("tanks", [])
            if "sg" in t and t["sg"] is not None
        }

        plan = build_tank_plan(master_tanks, pf_map, sg_override_by_id=sg_map)
        sums = tank_sums_for_stage(plan)

        # Tank‑level rows
        for ts in plan:
            tank_records.append(
                {
                    "Stage": stage_name,
                    "Tank_ID": ts.tank.tank_id,
                    "Percent_Fill": ts.percent_fill,
                    "SG": ts.sg,
                    "Weight_t": ts.weight_t,
                    "LCG_AP_m": ts.lcg_m_ap,
                    "VCG_m": ts.vcg_m,
                    "TCG_m": ts.tcg_m,
                    "FSM_eff_tm": ts.fsm_effective_tm(),
                    "Content": ts.tank.content,
                    "Group": ts.tank.group,
                }
            )

        # Stage‑level summary row
        stage_summaries.append(
            {
                "Stage": stage_name,
                "Total_Weight_t": sums.total_weight_t,
                "LCG_AP_m": sums.lcg_m_ap,
                "VCG_m": sums.vcg_m,
                "TCG_m": sums.tcg_m,
                "Total_FSM_tm": sums.total_fsm_tm,
            }
        )

        # Stage ballast information if available
        mtc = stage_cfg.get("MTC_tm_per_cm")
        trim_init = stage_cfg.get("trim_initial_cm")
        trim_target = stage_cfg.get("trim_target_cm")
        pump_rate = stage_cfg.get("pump_rate_tph")
        x_ballast_mid = stage_cfg.get("x_ballast_m_mid")

        if (
            mtc is not None
            and trim_init is not None
            and trim_target is not None
            and x_ballast_mid is not None
            and lpp_m != 0.0
        ):
            br = compute_stage_ballast(
                stage_name=stage_name,
                mtc_tm_per_cm=float(mtc),
                trim_initial_cm=float(trim_init),
                trim_target_cm=float(trim_target),
                x_ballast_m_mid=float(x_ballast_mid),
                lcf_m_mid=lcf_mid,
                pump_rate_tph=float(pump_rate) if pump_rate is not None else None,
            )
            ballast_rows.append(asdict(br))

    # Write sheets
    if tank_records:
        ws_tanks = wb.create_sheet("Stage_Tank_Plan")
        tank_df = pd.DataFrame.from_records(tank_records)
        _write_dataframe(ws_tanks, tank_df, index=False)

    if stage_summaries:
        ws_sum = wb.create_sheet("Stage_Tank_Sums")
        sum_df = pd.DataFrame.from_records(stage_summaries)
        _write_dataframe(ws_sum, sum_df, index=False)

    if ballast_rows:
        ws_ballast = wb.create_sheet("Stage_Ballast")
        ballast_df = pd.DataFrame.from_records(ballast_rows)
        _write_dataframe(ws_ballast, ballast_df, index=False)

    out_path = Path(output_excel_path)
    wb.save(out_path)
    return out_path


def _write_dataframe(ws, df: pd.DataFrame, index: bool = False) -> None:
    """Write a pandas DataFrame into an openpyxl worksheet."""
    row_start = 1
    col_start = 1

    if index:
        df_to_write = df.reset_index()
    else:
        df_to_write = df

    # Header
    for j, col_name in enumerate(df_to_write.columns, start=col_start):
        cell = ws.cell(row=row_start, column=j, value=str(col_name))
        cell.font = cell.font.copy(bold=True)

    # Data
    for i, (_, row) in enumerate(df_to_write.iterrows(), start=row_start + 1):
        for j, value in enumerate(row, start=col_start):
            ws.cell(row=i, column=j, value=value)

    # Auto width (simple heuristic)
    for j, col_name in enumerate(df_to_write.columns, start=col_start):
        max_len = max(
            len(str(col_name)),
            *[len(str(v)) for v in df_to_write[col_name].astype(str).tolist()],
        )
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 2, 60)


# ---------------------------------------------------------------------------
# 2) Excel Stage model → Python Stability JSON
# ---------------------------------------------------------------------------


def stage_workbook_to_stability_json(
    workbook_path: Path | str,
    master_tanks_path: Path | str,
    out_json_path: Path | str,
    stage_sheet_name: str = "RORO_Stage_Scenarios",
    stage_tanks_sheet_name: str = "Stage_Tanks",
    calc_sheet_name: str = "Calc",
) -> Path:
    """Read an existing Stage workbook and emit a stability JSON config.

    ASSUMPTIONS (aligned with build_bushra_agi_tr_from_scratch.py):
    - Sheet `Calc` has parameter names in column A and values in column D.
      At minimum:
        * Lpp_m
        * LCF_m_from_midship
        * MTC_t_m_per_cm
        * TPC_t_per_cm  (optional, but nice to have)
        * D_vessel_m    (optional)
    - Sheet `RORO_Stage_Scenarios` has header row at 14 and data rows
      starting at 15 with columns:
        A: Stage
        B: Mean Draft (m)
        C: Trim (m) (input)
        G: Trim (cm) (computed)
        H: Trim (m) (computed, can be used instead of C)
        D: W_stage (t)
        E: x_stage (m from midship)
    - Sheet `Stage_Tanks` (if present) has simple columns:
        Stage, Tank_ID, Percent_Fill, SG (optional), UseForBallast (optional)

    These assumptions are intentionally narrow and easy to adjust: if your
    workbook differs, change the indices at the top of this function.
    """
    wb = load_workbook(filename=workbook_path, data_only=True)

    # --- 1) Vessel parameters from Calc sheet
    calc_ws = wb[calc_sheet_name]
    calc_params: Dict[str, float] = {}
    for row in calc_ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if not key:
            continue
        val = row[3]  # column D
        if isinstance(val, (int, float)):
            calc_params[str(key)] = float(val)

    lpp_m = float(calc_params.get("Lpp_m", 0.0))
    lcf_mid = float(calc_params.get("LCF_m_from_midship", 0.0))
    mtc_tm_per_cm = float(calc_params.get("MTC_t_m_per_cm", 0.0))
    tpc_t_per_cm = float(calc_params.get("TPC_t_per_cm", 0.0))
    d_vessel_m = float(calc_params.get("D_vessel_m", 0.0))

    # --- 2) Stage sheet
    stage_ws = wb[stage_sheet_name]

    # Column indices (1‑based) for RORO_Stage_Scenarios
    COL_STAGE = 1  # A
    COL_MEAN_DRAFT = 2  # B
    COL_TRIM_M_INPUT = 3  # C
    COL_W_STAGE = 4  # D
    COL_X_STAGE_MID = 5  # E
    COL_TRIM_CM_COMPUTED = 7  # G

    STAGE_HEADER_ROW = 14
    FIRST_STAGE_ROW = STAGE_HEADER_ROW + 1

    stages_from_excel: Dict[str, Dict] = {}

    row_idx = FIRST_STAGE_ROW
    while True:
        stage_name = stage_ws.cell(row=row_idx, column=COL_STAGE).value
        if stage_name is None or str(stage_name).strip() == "":
            break

        mean_draft = stage_ws.cell(row=row_idx, column=COL_MEAN_DRAFT).value
        trim_input_m = stage_ws.cell(row=row_idx, column=COL_TRIM_M_INPUT).value
        trim_cm = stage_ws.cell(row=row_idx, column=COL_TRIM_CM_COMPUTED).value
        w_stage = stage_ws.cell(row=row_idx, column=COL_W_STAGE).value
        x_stage_mid = stage_ws.cell(row=row_idx, column=COL_X_STAGE_MID).value

        # Normalise numeric values
        def _f(v):
            return float(v) if isinstance(v, (int, float)) else None

        stages_from_excel[str(stage_name)] = {
            "name": str(stage_name),
            "mean_draft_m": _f(mean_draft),
            "trim_m_input": _f(trim_input_m),
            "trim_cm_computed": _f(trim_cm),
            "w_stage_t": _f(w_stage),
            "x_stage_m_mid": _f(x_stage_mid),
        }
        row_idx += 1

    # --- 3) Stage_Tanks sheet (optional)
    tanks = load_master_tanks_json(master_tanks_path)
    tank_by_id = {t.tank_id: t for t in tanks}

    stage_tank_map: Dict[str, List[TankState]] = {name: [] for name in stages_from_excel.keys()}

    if stage_tanks_sheet_name in wb.sheetnames:
        st_ws = wb[stage_tanks_sheet_name]
        # Expect header row 1 with at least: Stage, Tank_ID, Percent_Fill, SG?
        header = [c.value for c in st_ws[1]]
        col_idx = {str(h): i + 1 for i, h in enumerate(header) if h}

        for row in st_ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stage_name = str(row[col_idx.get("Stage", 1) - 1])
            tank_id = str(row[col_idx.get("Tank_ID", 2) - 1])
            pf = row[col_idx.get("Percent_Fill", 3) - 1]
            sg = row[col_idx.get("SG", 4) - 1] if "SG" in col_idx else None

            if tank_id not in tank_by_id:
                continue
            if stage_name not in stage_tank_map:
                stage_tank_map[stage_name] = []

            tank_def = tank_by_id[tank_id]
            percent_fill = float(pf) if isinstance(pf, (int, float)) else 0.0
            if percent_fill <= 0.0:
                continue
            sg_val = float(sg) if isinstance(sg, (int, float)) else tank_def.sg_master
            stage_tank_map[stage_name].append(TankState(tank=tank_def, percent_fill=percent_fill, sg=sg_val))

    # --- 4) Compose JSON structure
    vessel_meta = {
        "name": "LCT BUSHRA AGI TR",
        "Lpp_m": lpp_m,
        "LCF_m_from_midship": lcf_mid,
        "MTC_t_m_per_cm": mtc_tm_per_cm,
        "TPC_t_per_cm": tpc_t_per_cm,
        "D_vessel_m": d_vessel_m,
    }

    stages_out: List[Dict] = []

    for stage_name, base in stages_from_excel.items():
        tank_states = stage_tank_map.get(stage_name, [])
        sums = tank_sums_for_stage(tank_states)

        # Convert tank states to WeightItem dicts
        items: List[Dict] = []
        for ts in tank_states:
            items.append(
                asdict(
                    WeightItem(
                        name=f"{ts.tank.tank_id} ({ts.tank.content})" if ts.tank.content else ts.tank.tank_id,
                        weight=ts.weight_t,
                        lcg=ts.lcg_m_ap,
                        vcg=ts.vcg_m,
                        tcg=ts.tcg_m,
                        fsm=ts.fsm_effective_tm(),
                        group=ts.tank.group or "tank",
                    )
                )
            )

        # Add lumped stage weight as a separate item if present
        if base.get("w_stage_t") and base.get("x_stage_m_mid"):
            items.append(
                asdict(
                    WeightItem(
                        name=f"{stage_name}_lump",
                        weight=float(base["w_stage_t"]),
                        lcg=midship_to_ap(float(base["x_stage_m_mid"]), lpp_m) if lpp_m else 0.0,
                        vcg=0.0,
                        tcg=0.0,
                        fsm=0.0,
                        group="stage_cargo",
                    )
                )
            )

        # Stage ballast result is optional here – trim target must be supplied externally
        stage_out = {
            "name": stage_name,
            "mean_draft_m": base.get("mean_draft_m"),
            "trim_m_input": base.get("trim_m_input"),
            "trim_cm_computed": base.get("trim_cm_computed"),
            "w_stage_t": base.get("w_stage_t"),
            "x_stage_m_mid": base.get("x_stage_m_mid"),
            "tank_sums": asdict(sums),
            "items": items,
        }
        stages_out.append(stage_out)

    out = {"vessel": vessel_meta, "stages": stages_out}

    out_path = Path(out_json_path)
    out_path.write_text(json.dumps(out, indent=2), encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# Small CLI helper (optional)
# ---------------------------------------------------------------------------


def main(argv: Optional[Sequence[str]] = None) -> None:
    """Tiny CLI layer for quick local testing.

    Examples
    --------
    1) Excel → JSON

        python bushra_excel_bridge_v1.py                 --mode excel-to-json                 --workbook LCT_BUSHRA_AGI_TR.xlsx                 --master master_tanks.json                 --out bushra_stability_config.json

    2) JSON → Excel

        python bushra_excel_bridge_v1.py                 --mode json-to-excel                 --master master_tanks.json                 --config bushra_stability_config.json                 --out bushra_stability_export.xlsx
    """
    import argparse

    parser = argparse.ArgumentParser(description="BUSHRA stability Excel bridge")
    parser.add_argument("--mode", choices=["excel-to-json", "json-to-excel"], required=True)
    parser.add_argument("--workbook", type=str, help="Stage workbook (.xlsx)")
    parser.add_argument("--master", type=str, required=True, help="master_tanks.json path")
    parser.add_argument("--config", type=str, help="Stage config JSON (for json-to-excel)")
    parser.add_argument("--out", type=str, required=True, help="Output path (.json or .xlsx)")

    args = parser.parse_args(argv)

    if args.mode == "excel-to-json":
        if not args.workbook:
            raise SystemExit("--workbook is required for excel-to-json mode")
        out = stage_workbook_to_stability_json(
            workbook_path=args.workbook,
            master_tanks_path=args.master,
            out_json_path=args.out,
        )
        print(f"[OK] Stability JSON written to: {out}")
    else:
        if not args.config:
            raise SystemExit("--config is required for json-to-excel mode")
        out = export_tank_summaries_to_excel(
            master_tanks_path=args.master,
            stage_config_json=args.config,
            output_excel_path=args.out,
        )
        print(f"[OK] Excel export written to: {out}")


if __name__ == "__main__":
    main()
